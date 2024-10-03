import numpy as np
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.ticker import FuncFormatter
from openpyxl import load_workbook
import matplotlib.patches as patches
import pylab
import tkinter as tk

year = 2024
Kategorie = "Strom"

def getValueOfKategorieAllMonths_diagram_bottom_right(year, kategorie):
    KategorieInp = kategorie

    wb = load_workbook("Finanzen" + str(year) + ".xlsx")  # Erstellt NEUES, bei nur hinzufügen/abändern nutze load_workbook
    ws = wb["Finanzen"]

    for row in range(2, ws.max_row + 1):
        print(row)
        if KategorieInp == ws.cell(row=row, column=1).value:
            print("Gefunden")
            global KategorieRow
            KategorieRow = row

    global AllValuesTogether
    AllValuesTogether = 0
    global Beträge
    Beträge = []
    Monate = []

    for col in range(2, ws.max_column + 1):
        if col == int("1"):
            col = "A"
        if col == int("2"):
            col = "B"
        if col == int("3"):
            col = "C"
        if col == int("4"):
            col = "D"
        if col == int("5"):
            col = "E"
        if col == int("6"):
            col = "F"
        if col == int("7"):
            col = "G"
        if col == int("8"):
            col = "H"
        if col == int("9"):
            col = "I"
        if col == int("10"):
            col = "J"
        if col == int("11"):
            col = "K"
        if col == int("12"):
            col = "L"
        if col == int("13"):
            col = "M"

        if str(ws[str(col) + str(KategorieRow)].value) == "None":
            ws[str(col) + str(KategorieRow)].value = 0

        print(str(ws[str(col) + "1"].value) + ":")
        Monate.insert(int(Monate.__len__()) + 1, ws[str(col) + "1"].value)
        print(ws[str(col) + str(KategorieRow)].value)
        AllValuesTogether += round(float(ws[col + str(KategorieRow)].value), 2)
        Beträge.insert(int(Monate.__len__()) + 1, ws[str(col) + str(KategorieRow)].value)


    print("---------------")
    print("Zusammen:")
    AllValuesTogether = round(AllValuesTogether, 2)
    print(AllValuesTogether)

    print(Monate)
    print(Beträge)

    global AusgabenListe
    AusgabenListe = Beträge

def plot_eineKategorie(year, Kategorie):

    # Set up the plot

    plt.rcParams['font.family'] = 'sans-serif'
    plt.rcParams['font.sans-serif'] = ['Arial']
    fig, ax = plt.subplots(figsize=(16, 9), facecolor='#f0f0f0')
    ax.set_facecolor('#ffffff')
    fig.canvas.manager.set_window_title('Mein Finanzüberblick ' + str(year))


    getValueOfKategorieAllMonths_diagram_bottom_right(year, Kategorie)

    # Data
    months = ["Januar", "Februar", "März", "April", "Mai", "Juni", "Juli", "August", "September", "Oktober", "November", "Dezember"]
    expenses = AusgabenListe


    # place a text box in upper left in axes coords
    props = dict(boxstyle='round', facecolor='white', alpha=0.7)

    ax.text(0.992, 0.98, "Komplette Ausgaben in der Kategorie: " + str(AllValuesTogether), 
            transform=ax.transAxes, 
            fontsize=14, 
            verticalalignment='top', 
            horizontalalignment='right',
            bbox=props)

    
    # Add annotations for highest and lowest points
    '''
    max_expense = max(expenses)
    min_expense = min(expenses)

    # Annotation für das Maximum
    ax.annotate(f'Maximum: {max_expense}€', 
                xy=(months[expenses.index(max_expense)], max_expense),
                xytext=(50, 50), textcoords='offset points', 
                ha='left', va='bottom', fontsize=10,
                bbox=dict(boxstyle='round,pad=0.5', fc='yellow', alpha=0.5),
                arrowprops=dict(arrowstyle='->', connectionstyle='arc3,rad=0.3'))

    # Annotation für das Minimum
    ax.annotate(f'Minimum: {min_expense}€', 
                xy=(months[expenses.index(min_expense)], min_expense),
                xytext=(50, -50), textcoords='offset points', 
                ha='left', va='top', fontsize=10,
                bbox=dict(boxstyle='round,pad=0.5', fc='yellow', alpha=0.5),
                arrowprops=dict(arrowstyle='->', connectionstyle='arc3,rad=0.3'))
    '''




    # Set the width of each bar
    bar_width = 0.35

    # Set the positions of the bars on the x-axis
    r1 = np.arange(len(months))
    #r2 = [x + bar_width for x in r1]

    # Create the bar chart
    expense_bars = ax.bar(r1, expenses, color='#FF6B6B', width=bar_width, label='Ausgaben (€)', alpha=0.8)

    # Max-/min einnahmen/ausgaben Legende
    """
    yellow_frame = [0,0,0,0,0,0,0,0,0,0,0,0]
    yellow_label = ax.bar(r1, yellow_frame, color="#fcf52f", width=bar_width, label="Max./Min. Ausgaben", alpha=0.8)"""

    # Customize the plot
    ax.set_title("Ausgaben der Kategorie <" + str(Kategorie) + "> im Jahresverlauf " + str(year), fontsize=24, fontweight='bold', pad=20)
    ax.set_xlabel('Monate', fontsize=16, fontweight="bold", labelpad=15)
    ax.set_ylabel('Betrag (€)', fontsize=16, labelpad=15, fontweight="bold")

    
    # Calculate and plot trendline

    x_numeric = np.arange(len(months))
    z = np.polyfit(x_numeric, expenses, 1)
    p = np.poly1d(z)
    ax.plot(months, p(x_numeric), linestyle='', color='red', linewidth=2) # label weg und linestyle auf none da sond ein fehler mit 'april' auftritt wenn ich den gesamten bereich auskommentiere


    # Set x-axis ticks
    #ax.set_xticks([r + bar_width/2 for r in range(len(months))])
    ax.set_xticklabels(months, rotation=45, ha='right', fontsize=12)


    # Add grid
    ax.grid(True, linestyle='--', alpha=0.7, axis='y')



    # Add value labels on top of each bar
    def add_value_labels(bars):
        for bar in bars:
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2., height,
                    f'{height:,.0f}€',##########################################################Hier mit '2f' die nachkommastellen bearbeiten
                    ha='center', va='bottom', fontsize=10, fontweight='bold')

    add_value_labels(expense_bars)

    # Add a subtle box around the plot area
    for spine in ax.spines.values():
        spine.set_edgecolor('#CCCCCC')


    # Add a background color gradient
    gradient = np.linspace(0, 1, 256).reshape(1, -1)
    gradient = np.vstack((gradient, gradient))
    ax.imshow(gradient, extent=[ax.get_xlim()[0], ax.get_xlim()[1], ax.get_ylim()[0], ax.get_ylim()[1]], 
            aspect='auto', alpha=0.1, cmap='coolwarm')


    # Add a title box
    title_box = patches.Rectangle((0, 1.02), 1, 0.08, transform=ax.transAxes, fill=True,
                                facecolor='lightgray', edgecolor='none', alpha=0.8)
    ax.add_patch(title_box)


    # Customize legend
    ax.legend(fontsize=14, loc='upper left', frameon=True, facecolor='white', edgecolor='gray')

    # Add a watermark
    fig.text(0.95, 0, 'Jonas Gaiser', fontsize=12, color='gray', ha='right', va='bottom', alpha=0.5)
    
    # Adjust layout and display
    plt.tight_layout()
    plt.show()

plot_eineKategorie(year, Kategorie)