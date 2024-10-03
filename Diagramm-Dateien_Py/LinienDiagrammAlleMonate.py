import numpy as np
import matplotlib.pyplot as plt

import os
import sys
import time
import tkinter as tk
import file
from tkinter import messagebox, filedialog, colorchooser, ttk
from tkinter import *
from tkinter.ttk import *
import pyautogui
import keyboard
from openpyxl import Workbook, load_workbook
import datetime
from PIL import Image, ImageTk
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

import pandas as pd
def getAusgaben(year):
    wb = load_workbook("Finanzen" + year + ".xlsx")
    ws = wb["Finanzen"]

    wbGehalt = load_workbook("Gehalt" + year + ".xlsx")
    wsGehalt = wbGehalt["Gehalt"]

    global GehaltListe
    GehaltListe = []

    gehaltAll = 0
    for col in range(2, wsGehalt.max_column + 1):
        if col == int("1"):
            col = "A"
        if col == int("2"):
            col = "B"
        if col == int("3"):
            col = "C"
        if col == int("4"):
            col = "D"
        if col == int("5"):
            col = "E"
        if col == int("6"):
            col = "F"
        if col == int("7"):
            col = "G"
        if col == int("8"):
            col = "H"
        if col == int("9"):
            col = "I"
        if col == int("10"):
            col = "J"
        if col == int("11"):
            col = "K"
        if col == int("12"):
            col = "L"
        if col == int("13"):
            col = "M"

        gehaltDieserMonat = 0
        for row in range(2, wsGehalt.max_row + 1):
            print(row)
            gehaltDieserMonat += round(float(wsGehalt[str(col) + str(row)].value), 2)
        print("Gehalt dieser Monat: " + str(gehaltDieserMonat))

        GehaltListe.append(gehaltDieserMonat)

        gehaltAll += gehaltDieserMonat
        gehaltAll = round(gehaltAll, 2)
        if str(gehaltDieserMonat) == "0.0":
            print("Gehalt noch nicht eingetragen für den Monat " + str(wsGehalt[str(col) + str(1)].value))
            #showText("Das Gehalt für den Monat " + str(wsGehalt[str(col) + str(
                #1)].value) + " wurde noch nicht eingetragen. Gehe dafür bitte zum Hauptfenster zurück und füge dein Gehalt links unten, unter <Gehalt> hinzu.")
        print("Gehalt all:" + str(gehaltAll))

    global AllMonthsValuesTogether
    AllMonthsValuesTogether = 0

    

    global AusgabenL
    AusgabenL = []

    for col in range(2, ws.max_column + 1):
        if col == int("1"):
            col = "A"
        if col == int("2"):
            col = "B"
        if col == int("3"):
            col = "C"
        if col == int("4"):
            col = "D"
        if col == int("5"):
            col = "E"
        if col == int("6"):
            col = "F"
        if col == int("7"):
            col = "G"
        if col == int("8"):
            col = "H"
        if col == int("9"):
            col = "I"
        if col == int("10"):
            col = "J"
        if col == int("11"):
            col = "K"
        if col == int("12"):
            col = "L"
        if col == int("13"):
            col = "M"
        print(col)
        MonthValue = 0

        for row in range(2, ws.max_row + 1):
            print(row)
            MonthValue += float(ws[str(col) + str(row)].value)


        print("MonthValue: " + str(MonthValue))
        AllMonthsValuesTogether += float(MonthValue)
        print(round(AllMonthsValuesTogether, 2), f"= round({AllMonthsValuesTogether}, 2)")
        AllMonthsValuesTogether = round(AllMonthsValuesTogether, 2)

        AusgabenL.append(MonthValue)


    Relation = float(gehaltAll) - float(AllMonthsValuesTogether)
    print(Relation)
    Relation = round(Relation, 2)

    print(str(GehaltListe))
    print(str(AusgabenL))
    global AusgabenListe
    AusgabenListe = AusgabenL

# Set the style for a more modern look
plt.style.use('ggplot')

# Create the figure and axis objects
fig, ax = plt.subplots(figsize=(16, 9))
getAusgaben("2024")
# Data
months = ["Januar", "Februar", "März", "April", "Mai", "Juni", "Juli", "August", "September", "Oktober", "November", "Dezember"]
expenses = AusgabenListe
salary = GehaltListe

# Plot the original data
ax.plot(months, expenses, marker='o', linewidth=2, markersize=8, label='Ausgaben (€)', color='#1f77b4')
ax.plot(months, salary, marker='o', linewidth=2, markersize=8, label='Einkommen (€)', color='green')

# Calculate and plot trendline
x_numeric = np.arange(len(months))
z = np.polyfit(x_numeric, expenses, 1)
p = np.poly1d(z)
ax.plot(months, p(x_numeric), linestyle='--', color='red', linewidth=2, label='Trendlinie (Ausgaben)')

x_numeric2 = np.arange(len(months))
z2 = np.polyfit(x_numeric2, salary, 1)
p2 = np.poly1d(z2)
ax.plot(months, p2(x_numeric2), linestyle='--', color='blue', linewidth=2, label='Trendlinie (Gehalt)')

# Customize the plot
ax.set_title("Monatliche Ausgaben im Jahresverlauf", fontsize=20, pad=20)
ax.set_xlabel('Monate', fontsize=14, labelpad=10)
ax.set_ylabel('Ausgaben (€)', fontsize=14, labelpad=10)

# Rotate x-axis labels and adjust their alignment
plt.xticks(rotation=45, ha='right')

# Add grid
ax.grid(True, linestyle=':', alpha=0.7)

# Customize legend
ax.legend(fontsize=12, loc='upper left')

# Add annotations for highest and lowest points
max_expense = max(expenses)
min_expense = min(expenses)
ax.annotate(f'Maximum: {max_expense}€', 
            xy=(months[expenses.index(max_expense)], max_expense),
            xytext=(10, 10), textcoords='offset points', 
            ha='left', va='bottom', fontsize=10,
            bbox=dict(boxstyle='round,pad=0.5', fc='yellow', alpha=0.5),
            arrowprops=dict(arrowstyle='->', connectionstyle='arc3,rad=0'))

ax.annotate(f'Minimum: {min_expense}€', 
            xy=(months[expenses.index(min_expense)], min_expense),
            xytext=(10, -10), textcoords='offset points', 
            ha='left', va='top', fontsize=10,
            bbox=dict(boxstyle='round,pad=0.5', fc='yellow', alpha=0.5),
            arrowprops=dict(arrowstyle='->', connectionstyle='arc3,rad=0'))


max_salary = max(salary)
min_salary = min(salary)
ax.annotate(f'Maximum: {max_salary}€', 
            xy=(months[salary.index(max_salary)], max_salary),
            xytext=(10, 10), textcoords='offset points', 
            ha='left', va='bottom', fontsize=10,
            bbox=dict(boxstyle='round,pad=0.5', fc='purple', alpha=0.5),
            arrowprops=dict(arrowstyle='->', connectionstyle='arc3,rad=0'))

ax.annotate(f'Minimum: {min_salary}€', 
            xy=(months[salary.index(min_salary)], min_salary),
            xytext=(10, -10), textcoords='offset points', 
            ha='left', va='top', fontsize=10,
            bbox=dict(boxstyle='round,pad=0.5', fc='purple', alpha=0.5),
            arrowprops=dict(arrowstyle='->', connectionstyle='arc3,rad=0'))


gradient = np.linspace(0, 1, 256).reshape(1, -1)
gradient = np.vstack((gradient, gradient))
ax.imshow(gradient, extent=[ax.get_xlim()[0], ax.get_xlim()[1], ax.get_ylim()[0], ax.get_ylim()[1]], 
          aspect='auto', alpha=0.1, cmap='coolwarm')

# Adjust layout and display
plt.tight_layout()
plt.show()