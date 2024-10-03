import numpy as np
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.ticker import FuncFormatter
from openpyxl import load_workbook
import matplotlib.patches as patches
import pylab

def getAusgabenGehalt_diagram_top_left(year):
    wb = load_workbook("Finanzen" + year + ".xlsx")
    ws = wb["Finanzen"]

    wbGehalt = load_workbook("Gehalt" + year + ".xlsx")
    wsGehalt = wbGehalt["Gehalt"]

    global GehaltListe
    GehaltListe = []

    global gehaltAll
    gehaltAll = 0
    for col in range(2, wsGehalt.max_column + 1):
        if col == int("1"):
            col = "A"
        if col == int("2"):
            col = "B"
        if col == int("3"):
            col = "C"
        if col == int("4"):
            col = "D"
        if col == int("5"):
            col = "E"
        if col == int("6"):
            col = "F"
        if col == int("7"):
            col = "G"
        if col == int("8"):
            col = "H"
        if col == int("9"):
            col = "I"
        if col == int("10"):
            col = "J"
        if col == int("11"):
            col = "K"
        if col == int("12"):
            col = "L"
        if col == int("13"):
            col = "M"

        gehaltDieserMonat = 0
        for row in range(2, wsGehalt.max_row + 1):
            print(row)
            gehaltDieserMonat += round(float(wsGehalt[str(col) + str(row)].value), 2)
        print("Gehalt dieser Monat: " + str(gehaltDieserMonat))
        GehaltListe.append(gehaltDieserMonat)

        gehaltAll += float(gehaltDieserMonat)
        #gehaltAll = round(gehaltAll, 0)
        if str(gehaltDieserMonat) == "0.0":
            print("Gehalt noch nicht eingetragen für den Monat " + str(wsGehalt[str(col) + str(1)].value))
            #showText("Das Gehalt für den Monat " + str(wsGehalt[str(col) + str(
                #1)].value) + " wurde noch nicht eingetragen. Gehe dafür bitte zum Hauptfenster zurück und füge dein Gehalt links unten, unter <Gehalt> hinzu.")
        print("Gehalt all:" + str(gehaltAll))

    global AllMonthsValuesTogether
    AllMonthsValuesTogether = 0

    

    global AusgabenL
    AusgabenL = []

    for col in range(2, ws.max_column + 1):
        if col == int("1"):
            col = "A"
        if col == int("2"):
            col = "B"
        if col == int("3"):
            col = "C"
        if col == int("4"):
            col = "D"
        if col == int("5"):
            col = "E"
        if col == int("6"):
            col = "F"
        if col == int("7"):
            col = "G"
        if col == int("8"):
            col = "H"
        if col == int("9"):
            col = "I"
        if col == int("10"):
            col = "J"
        if col == int("11"):
            col = "K"
        if col == int("12"):
            col = "L"
        if col == int("13"):
            col = "M"
        print(col)
        MonthValue = 0

        for row in range(2, ws.max_row + 1):
            print(row)
            MonthValue += float(ws[str(col) + str(row)].value)


        print("MonthValue: " + str(MonthValue))
        AllMonthsValuesTogether += float(MonthValue)
        print(round(AllMonthsValuesTogether, 2), f"= round({AllMonthsValuesTogether}, 2)")
        AllMonthsValuesTogether = round(AllMonthsValuesTogether, 2)

        AusgabenL.append(MonthValue)

    global Relation
    Relation = float(gehaltAll) - float(AllMonthsValuesTogether)
    print(Relation)
    Relation = round(Relation, 2)

    print(str(GehaltListe))
    print(str(AusgabenL))
    global AusgabenListe
    AusgabenListe = AusgabenL


def plot_Months(year):
    # Set up the plot

    plt.rcParams['font.family'] = 'sans-serif'
    plt.rcParams['font.sans-serif'] = ['Arial']
    fig, ax = plt.subplots(figsize=(16, 9), facecolor='#f0f0f0')
    ax.set_facecolor('#ffffff')
    fig.canvas.manager.set_window_title(f'Mein Finanzüberblick {year}')


    getAusgabenGehalt_diagram_top_left(year)

    # Data
    months = ["Januar", "Februar", "März", "April", "Mai", "Juni", "Juli", "August", "September", "Oktober", "November", "Dezember"]
    expenses = AusgabenListe
    salary = GehaltListe

    
    # Add annotations for highest and lowest points
    """
    max_expense = max(expenses)
    min_expense = min(expenses)

    # Annotation für das Maximum
    ax.annotate(f'Maximum: {max_expense}€', 
                xy=(months[expenses.index(max_expense)], max_expense),
                xytext=(50, 50), textcoords='offset points', 
                ha='left', va='bottom', fontsize=10,
                bbox=dict(boxstyle='round,pad=0.5', fc='yellow', alpha=0.5),
                arrowprops=dict(arrowstyle='->', connectionstyle='arc3,rad=0.3'))

    # Annotation für das Minimum
    ax.annotate(f'Minimum: {min_expense}€', 
                xy=(months[expenses.index(min_expense)], min_expense),
                xytext=(50, -50), textcoords='offset points', 
                ha='left', va='top', fontsize=10,
                bbox=dict(boxstyle='round,pad=0.5', fc='yellow', alpha=0.5),
                arrowprops=dict(arrowstyle='->', connectionstyle='arc3,rad=0.3'))


    max_salary = max(salary)
    min_salary = min(salary)
    # Annotation für das Maximum
    ax.annotate(f'Maximum: {max_salary}€', 
                xy=(months[salary.index(max_salary)], max_salary),
                xytext=(50, 50), textcoords='offset points', 
                ha='left', va='bottom', fontsize=10,
                bbox=dict(boxstyle='round,pad=0.5', fc='purple', alpha=0.5),
                arrowprops=dict(arrowstyle='->', connectionstyle='arc3,rad=0.3'))

    # Annotation für das Minimum
    ax.annotate(f'Minimum: {min_salary}€', 
                xy=(months[salary.index(min_salary)], min_salary),
                xytext=(50, -50), textcoords='offset points', 
                ha='left', va='top', fontsize=10,
                bbox=dict(boxstyle='round,pad=0.5', fc='purple', alpha=0.5),
                arrowprops=dict(arrowstyle='->', connectionstyle='arc3,rad=0.3'))
    """


    # Set the width of each bar
    bar_width = 0.35

    # Set the positions of the bars on the x-axis
    r1 = np.arange(len(months))
    r2 = [x + bar_width for x in r1]

    # Create the bar chart
    expense_bars = ax.bar(r1, expenses, color='#FF6B6B', width=bar_width, label='Ausgaben (€)', alpha=0.8)
    salary_bars = ax.bar(r2, salary, color='#4ECDC4', width=bar_width, label='Einnahmen (€)', alpha=0.8)

    # Max-/Min Einkommen/Ausgaben Legende
    """purple_frame = [0,0,0,0,0,0,0,0,0,0,0,0]
    purple_label = ax.bar(r1, purple_frame, color="#b555b5", width=bar_width, label="Max./Min. Einnahmen", alpha=0.8)
    yellow_frame = [0,0,0,0,0,0,0,0,0,0,0,0]
    yellow_label = ax.bar(r1, yellow_frame, color="#fcf52f", width=bar_width, label="Max./Min. Ausgaben", alpha=0.8)"""

    # Customize the plot
    ax.set_title("Monatliche Ausgaben und Einnahmen im Jahresverlauf 2024", fontsize=24, fontweight='bold', pad=20)
    ax.set_xlabel('Monate', fontsize=16, labelpad=15, fontweight="bold")
    ax.set_ylabel('Betrag (€)', fontsize=16, labelpad=15, fontweight="bold")

    # place a text box in upper left in axes coords
    props = dict(boxstyle='round', facecolor='white', alpha=0.7)

    ax.text(0.992, 0.98, "Komplette Einnahmen: " + str(gehaltAll) + '\n' + "Komplette Ausgaben: " + str(AllMonthsValuesTogether) + '\n' + '\n' + "Differenz: " + str(Relation), 
            transform=ax.transAxes, 
            fontsize=14, 
            verticalalignment='top', 
            horizontalalignment='right',
            bbox=props)
    
    #Trendlinie
    """# Calculate and plot trendline
    x_numeric = np.arange(len(months))
    z = np.polyfit(x_numeric, expenses, 1)
    p = np.poly1d(z)
    ax.plot(months, p(x_numeric), linestyle='--', color='red', linewidth=2, label='Trendlinie (Ausgaben)')

    x_numeric2 = np.arange(len(months))
    z2 = np.polyfit(x_numeric2, salary, 1)
    p2 = np.poly1d(z2)
    ax.plot(months, p2(x_numeric2), linestyle='--', color='blue', linewidth=2, label='Trendlinie (Gehalt)')"""

    # Set x-axis ticks
    ax.set_xticks([r + bar_width/2 for r in range(len(months))])
    ax.set_xticklabels(months, rotation=45, ha='right', fontsize=12)

    # Customize y-axis
    ax.yaxis.set_major_formatter(FuncFormatter(lambda x, p: f'{x:,.0f}'))
    ax.set_ylim(0, max(max(expenses), max(salary)) * 1.1)

    # Add grid
    ax.grid(True, linestyle='--', alpha=0.7, axis='y')



    # Add value labels on top of each bar
    def add_value_labels(bars):
        for bar in bars:
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2., height,
                    f'{height:,.0f}€',##########################################################Hier mit '2f' die nachkommastellen bearbeiten
                    ha='center', va='bottom', fontsize=10, fontweight='bold')

    add_value_labels(expense_bars)
    add_value_labels(salary_bars)

    # Add a subtle box around the plot area
    for spine in ax.spines.values():
        spine.set_edgecolor('#CCCCCC')

    # Add a background color gradient
    gradient = np.linspace(0, 1, 256).reshape(1, -1)
    gradient = np.vstack((gradient, gradient))
    ax.imshow(gradient, extent=[ax.get_xlim()[0], ax.get_xlim()[1], ax.get_ylim()[0], ax.get_ylim()[1]], 
            aspect='auto', alpha=0.1, cmap='coolwarm')

    # Add some visual enhancements
    for bar in expense_bars + salary_bars:
        bar.set_edgecolor('white')
        bar.set_linewidth(0.5)

    # Add a title box
    title_box = patches.Rectangle((0, 1.02), 1, 0.08, transform=ax.transAxes, fill=True,
                                facecolor='lightgray', edgecolor='none', alpha=0.8)
    ax.add_patch(title_box)


    # Customize legend
    ax.legend(fontsize=14, loc='upper left', frameon=True, facecolor='white', edgecolor='gray')

    # Add a watermark
    fig.text(0.95, 0, 'Jonas Gaiser', fontsize=12, color='gray', ha='right', va='bottom', alpha=0.5)


    # Adjust layout and display
    plt.tight_layout()
    plt.show()

plot_Months("2024")