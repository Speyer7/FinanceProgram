import tkinter as tk
from tkinter import messagebox, ttk
from tkinter import *
from tkinter.ttk import *
from openpyxl import load_workbook
import datetime
import tkinter as tk
from tkinter import messagebox, ttk
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.ticker import FuncFormatter
from openpyxl import load_workbook
import matplotlib.patches as patches
from datetime import datetime
import os
import sys
import pandas as pd




def get_log_file_path():
    # Aktuelles Datum
    global now
    now = datetime.now()
    year = now.strftime("%Y")
    month = now.strftime("%m")

    # Basisverzeichnis (gleicher Ordner wie das Skript)
    base_dir = os.path.dirname(os.path.abspath(__file__))

    # Hauptlog-Ordner
    log_dir = "Log"#os.path.join(base_dir, "Log")
    if not os.path.exists(log_dir):
        os.makedirs(log_dir)

    # Jahresordner innerhalb des Log-Ordners
    year_dir = os.path.join(log_dir, year)
    if not os.path.exists(year_dir):
        os.makedirs(year_dir)

    # Log-Dateiname (YYYY-MM.log)
    log_filename = f"{year}-{month}.log"
    
    # Vollständiger Pfad zur Log-Datei
    return os.path.join(year_dir, log_filename)

class Logger(object):
    def __init__(self):
        self.terminal = sys.stdout
        self.log_file_path = get_log_file_path()
        self.log = open(self.log_file_path, "a", encoding='utf-8')
        self.last_check = datetime.now()

    def write(self, message):
        self.terminal.write(message)
        
        # Überprüfen, ob ein neuer Monat begonnen hat
        now = datetime.now()
        if now.month != self.last_check.month or now.year != self.last_check.year:
            self.log.close()
            self.log_file_path = get_log_file_path()
            self.log = open(self.log_file_path, "a", encoding='utf-8')
            self.last_check = now

        self.log.write(message)

    def flush(self):
        self.terminal.flush()
        self.log.flush()

# Umleiten von sys.stdout
sys.stdout = Logger()




print("----------------------------------" + str(now) + "----------------------------------")




global current_date
current_date = datetime.today().strftime('%Y-%m-%d')
global current_dateYear
current_dateYear = datetime.today().strftime('%Y')
global current_dateMonth
current_dateMonth = datetime.today().strftime('%m')
global current_dateday
current_dateday =datetime.today().strftime('%d')
print(current_dateday)

if current_dateMonth == "01":
    current_dateMonth = "Januar"
if current_dateMonth == "02":
    current_dateMonth = "Februar"
if current_dateMonth == "03":
    current_dateMonth = "März"
if current_dateMonth == "04":
    current_dateMonth = "April"
if current_dateMonth == "05":
    current_dateMonth = "Mai"
if current_dateMonth == "06":
    current_dateMonth = "Juni"
if current_dateMonth == "07":
    current_dateMonth = "Juli"
if current_dateMonth == "08":
    current_dateMonth = "August"
if current_dateMonth == "09":
    current_dateMonth = "September"
if current_dateMonth == "10":
    current_dateMonth = "Oktober"
if current_dateMonth == "11":
    current_dateMonth = "November"
if current_dateMonth == "12":
    current_dateMonth = "Dezember"
print(current_dateMonth)






def showText(text):
    messagebox.showinfo("Info", text)

def showWarn(text):
    messagebox.showwarning("Achtung", text)

def showError(text):
    messagebox.showerror("Error", text)

def PlsReopenSoftware():
    print("For good installing, please open this software again.")
    messagebox.showwarning("Info", "For good installing, please open this software again.")
    exit(1)






def BeforeStart():
    print("Load...")

    try:
        with open("FirstStart.txt", "r") as FirstStartFile:
            FirstStart = FirstStartFile.read()
            if FirstStart == "1" or FirstStart == "": # First Start und grundeinstellungen
                print("FirstStart")

                with open("FirstStart.txt", 'w') as file:
                    file.write('0')


                return 1

            elif FirstStart == "0":
                print("initializing...") # Nicht erster Start trotzdem hier noch ein paar checks

                def speichern_liste(dateiname, liste):
                    with open(dateiname, 'w') as file:
                        for element in liste:
                            file.write(f"{element}\n")

                def laden_liste(dateiname):
                    liste = []
                    try:
                        with open(dateiname, 'r') as file:
                            liste = [zeile.strip() for zeile in file]
                    except FileNotFoundError:
                        print(f"Die Datei {dateiname} wurde nicht gefunden und wird neu erstellt.")
                        with open(dateiname, 'w') as file:
                            file.write('')
                        PlsReopenSoftware()
                    return liste

                def hinzufuegen_element(dateiname, element):
                    liste = laden_liste(dateiname)
                    liste.append(element)
                    speichern_liste(dateiname, liste)

                KategorieListName = "KategorieList.txt"
                KategorieList = laden_liste(KategorieListName)  # Checkt ob Kategorie Liste vorhanden ist
                print("Aktuelle KategorieListe:", KategorieList)

                YearsListName = "YearsList.txt"
                YearsList = laden_liste(YearsListName)  #Checkt ob Years Liste vorhanden ist
                print("Aktuelle YearsListe:", YearsList)

                EinnahmequelleList = "EinnahmequelleList.txt"
                EinnahmequelleList = laden_liste(EinnahmequelleList)  # Checkt ob Years Liste vorhanden ist
                print("Aktuelle EinnahmequelleListe:", EinnahmequelleList)


                if YearsList.__contains__(str(current_dateYear)):
                    print("Up to date (year)")
                else:
                    hinzufuegen_element(YearsListName, current_dateYear)
                    from openpyxl import Workbook
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Finanzen"

                    ws.append(["", "Januar", "Februar", "März", "April", "Mai", "Juni", "Juli", "August", "September",
                               "Oktober", "November", "Dezember"])  # Fügt eine reihe hinzu

                    wb.save("Finanzen" + str(current_dateYear) + ".xlsx")


                    for KL_length in range(0, KategorieList.__len__()):
                        KatetegorieListItem = KategorieList.__getitem__(KL_length)


                        NewKategorieInp = KatetegorieListItem

                        from openpyxl import load_workbook
                        wb = load_workbook("Finanzen" + str(current_dateYear) + ".xlsx")  # Erstellt NEUES, bei nur hinzufügen/abändern nutze load_workbook
                        ws = wb["Finanzen"]

                        # ws["A" + str(1)].value = NewKategorieInp
                        ws["A" + str((int(ws.max_row) + 1))].value = NewKategorieInp
                        wb.save("Finanzen" + str(current_dateYear) + ".xlsx")

                        wb = load_workbook(
                            "Finanzen" + str(current_dateYear) + ".xlsx")  # Erstellt NEUES, bei nur hinzufügen/abändern nutze load_workbook
                        ws = wb["Finanzen"]
                        print("clearing Tabel...")
                        for col in range(1, ws.max_column + 1):
                            print("---Col: " + str(col))
                            if col == int("1"):
                                col = "A"
                            if col == int("2"):
                                col = "B"
                            if col == int("3"):
                                col = "C"
                            if col == int("4"):
                                col = "D"
                            if col == int("5"):
                                col = "E"
                            if col == int("6"):
                                col = "F"
                            if col == int("7"):
                                col = "G"
                            if col == int("8"):
                                col = "H"
                            if col == int("9"):
                                col = "I"
                            if col == int("10"):
                                col = "J"
                            if col == int("11"):
                                col = "K"
                            if col == int("12"):
                                col = "L"
                            if col == int("13"):
                                col = "M"
                            for row in range(2, ws.max_row + 1):
                                print(row)
                                if str(ws[str(col) + str(row)].value) == "None":
                                    ws[str(col) + str(row)].value = 0
                                    print("Changend")

                        wb.save("Finanzen" + str(current_dateYear) + ".xlsx")

                    wbGehalt = Workbook()
                    wsGehalt = wbGehalt.active
                    wsGehalt.title = "Gehalt"
                    wsGehalt.append(
                        ["", "Januar", "Februar", "März", "April", "Mai", "Juni", "Juli", "August", "September",
                            "Oktober",
                            "November", "Dezember"])  # Fügt eine reihe hinzu
                    #wsGehalt.append(["Gehalt:", "0", "0", "0", "0", "0", "0", "0", "0", "0", "0", "0", "0"])
                    wbGehalt.save("Gehalt" + str(current_dateYear) + ".xlsx")

                    for EQL_length in range(0, EinnahmequelleList.__len__()):
                        EQListItem = EinnahmequelleList.__getitem__(EQL_length)

                        from openpyxl import load_workbook
                        wb = load_workbook("Gehalt" + str(
                            current_dateYear) + ".xlsx")  # Erstellt NEUES, bei nur hinzufügen/abändern nutze load_workbook
                        ws = wb["Gehalt"]

                        ws["A" + str((int(ws.max_row) + 1))].value = EQListItem
                        wb.save("Gehalt" + str(current_dateYear) + ".xlsx")

                        wb = load_workbook(
                            "Gehalt" + str(
                                current_dateYear) + ".xlsx")  # Erstellt NEUES, bei nur hinzufügen/abändern nutze load_workbook
                        ws = wb["Gehalt"]
                        print("clearing Tabel...")
                        for col in range(1, ws.max_column + 1):
                            print("---Col: " + str(col))
                            if col == int("1"):
                                col = "A"
                            if col == int("2"):
                                col = "B"
                            if col == int("3"):
                                col = "C"
                            if col == int("4"):
                                col = "D"
                            if col == int("5"):
                                col = "E"
                            if col == int("6"):
                                col = "F"
                            if col == int("7"):
                                col = "G"
                            if col == int("8"):
                                col = "H"
                            if col == int("9"):
                                col = "I"
                            if col == int("10"):
                                col = "J"
                            if col == int("11"):
                                col = "K"
                            if col == int("12"):
                                col = "L"
                            if col == int("13"):
                                col = "M"
                            for row in range(2, ws.max_row + 1):
                                print(row)
                                if str(ws[str(col) + str(row)].value) == "None":
                                    ws[str(col) + str(row)].value = 0
                                    print("Changend")
                                    wb.save("Gehalt" + str(current_dateYear) + ".xlsx")

                    PlsReopenSoftware()


                return 0

    except FileNotFoundError:
        with open("FirstStart.txt", 'w') as file:
            print("File not found")
            file.write('1')
        from openpyxl import Workbook, load_workbook

        wb = Workbook()  # Erstellt NEUES, bei nur hinzufügen/abändern nutze load_workbook
        ws = wb.active
        ws.title = "Finanzen"

        ws.append(["", "Januar", "Februar", "März", "April", "Mai", "Juni", "Juli", "August", "September", "Oktober",
                   "November", "Dezember"])  # Fügt eine reihe hinzu

        wb.save("Finanzen" + str(current_dateYear) + ".xlsx")


        wbGehalt = Workbook()
        wsGehalt = wbGehalt.active
        wsGehalt.title = "Gehalt"
        wsGehalt.append(["", "Januar", "Februar", "März", "April", "Mai", "Juni", "Juli", "August", "September", "Oktober",
                   "November", "Dezember"])  # Fügt eine reihe hinzu
        #wsGehalt.append(["Gehalt:", "0", "0", "0", "0", "0", "0", "0", "0", "0", "0", "0", "0"])
        wbGehalt.save("Gehalt" + str(current_dateYear) + ".xlsx")

        return 1

if BeforeStart() == 1:
    PlsReopenSoftware()
else:
    print("Not First Start")





# ALLGEMEIN

root = tk.Tk()
root.title('Finanzen')

#root.iconphoto(False, tk.PhotoImage(file='icon.png'))


# Bildschirmgröße ermitteln
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()

root.geometry(f'{screen_width//2}x{screen_height//2}')
root.minsize(1200,900)





# GEHALT FÜR DIESEN MONAT SCHON EINGETRAGEN?
def checkMonatGehaltInsert():
    from openpyxl import load_workbook
    wbGehalt = load_workbook("Gehalt" + str(current_dateYear) + ".xlsx")
    wsGehalt = wbGehalt["Gehalt"]

    if current_dateMonth == "Januar":
        current_dateMonthIA = "B"
    if current_dateMonth == "Februar":
        current_dateMonthIA = "C"
    if current_dateMonth == "März":
        current_dateMonthIA = "D"
    if current_dateMonth == "April":
        current_dateMonthIA = "E"
    if current_dateMonth == "Mai":
        current_dateMonthIA = "F"
    if current_dateMonth == "Juni":
        current_dateMonthIA = "G"
    if current_dateMonth == "Juli":
        current_dateMonthIA = "H"
    if current_dateMonth == "August":
        current_dateMonthIA = "I"
    if current_dateMonth == "September":
        current_dateMonthIA = "J"
    if current_dateMonth == "Oktober":
        current_dateMonthIA = "K"
    if current_dateMonth == "November":
        current_dateMonthIA = "M"
    if current_dateMonth == "Dezember":
        current_dateMonthIA = "P"

    
    # Check ob überhaupt möglich
    if str(wsGehalt[current_dateMonthIA + str("2")].value) == "None":
        print("CheckMonatsGehaltInsert() nicht möglich. Es wurde noch keine Einnahmequelle definiert. Empfohlen es dringend zu machen.")
    else:
        print(str(wsGehalt[current_dateMonthIA + "2"].value))
        print(current_dateMonthIA + "2")
        gehaltdieserMonat = 0
        for row in range(2, wsGehalt.max_row + 1):
            print(row)
            gehaltdieserMonat += float(wsGehalt[current_dateMonthIA + str(row)].value) # gehaltdieserMonat += int(wsGehalt[current_dateMonthIA + str(row)].value)

        if gehaltdieserMonat == 0:
            showWarn("Einnahmen für diesen Monat muss noch eingetragen werden.")    

        #if str(wsGehalt[current_dateMonthIA + str(2)].value) == "0":
            #showText("Gehalt für diesen Monat muss noch eingetragen werden.")

        wbGehalt.save("Gehalt" + str(current_dateYear) + ".xlsx")
checkMonatGehaltInsert()





# FUNKTIONEN

def speichern_liste(dateiname, liste):
    with open(dateiname, 'w') as file:
        for element in liste:
            file.write(f"{element}\n")


def laden_liste(dateiname):
    liste = []
    try:
        with open(dateiname, 'r') as file:
            liste = [zeile.strip() for zeile in file]
    except FileNotFoundError:
        print(f"Die Datei {dateiname} wurde nicht gefunden und wird neu erstellt.")
        with open(dateiname, 'w') as file:
            file.write('')
        PlsReopenSoftware()
    return liste


def hinzufuegen_element(dateiname, element):
    liste = laden_liste(dateiname)
    liste.append(element)
    speichern_liste(dateiname, liste)

def is_valid_number(string):
    # Entfernen Sie Leerzeichen am Anfang und Ende des Strings
    string = string.strip()
    
    # Überprüfen Sie, ob der String leer ist
    if not string:
        return False
    
    # Teilen Sie den String am Punkt (falls vorhanden)
    parts = string.split('.')
    
    # Überprüfen Sie, ob es mehr als einen Punkt gibt
    if len(parts) > 2:
        return False
    
    # Überprüfen Sie den Teil vor dem Punkt (oder den ganzen String, wenn kein Punkt vorhanden)
    if not parts[0].isdigit():
        return False
    
    # Wenn es einen Nachkommateil gibt, überprüfen Sie ihn
    if len(parts) == 2:
        # Überprüfen Sie, ob der Nachkommateil nur aus Ziffern besteht und nicht mehr als 2 Stellen hat
        if not parts[1].isdigit() or len(parts[1]) > 2:
            return False
    
    return True

# TABELLEN

# Openpyxl

wb = load_workbook("Finanzen" + "2024" + ".xlsx") # Erstellt NEUES, bei nur hinzufügen/abändern nutze load_workbook
ws = wb["Finanzen"]

YearsListName = "YearsList.txt"
YearsList = laden_liste(YearsListName)  #Checkt ob Years Liste vorhanden ist


def ReadOutContentOfRow():
    row = input("Enter a row: ")

    print(" " + str(ws["A" + str(row)].value))
    print("Januar: " + str(ws["B" + str(row)].value))
    print("Februar: " + str(ws["C" + str(row)].value))
    print("März: " + str(ws["D" + str(row)].value))
    print("April: " + str(ws["E" + str(row)].value))
    print("Mai: " + str(ws["F" + str(row)].value))
    print("Juni: " + str(ws["G" + str(row)].value))
    print("Juli: " + str(ws["H" + str(row)].value))
    print("August: " + str(ws["I" + str(row)].value))
    print("September: " + str(ws["J" + str(row)].value))
    print("Oktober: " + str(ws["K" + str(row)].value))
    print("November: " + str(ws["L" + str(row)].value))
    print("Dezember: " + str(ws["M" + str(row)].value))


def ReadOutContentOfKategorie():
    KategorieInp = input("Enter a Kategorie: ")

    for row in range(1, ws.max_row + 1):
        print(row)
        if KategorieInp == ws.cell(row=row, column=1).value:
            print("Gefunden")
            global KategorieRowRead
            KategorieRowRead = row

    print("----------------------------------")
    print(str(KategorieRowRead) + " ist die gesuchte Zeile.")
    print("----------------------------------")

    print(" " + str(ws["A" + str(KategorieRowRead)].value))
    print("Januar: " + str(ws["B" + str(KategorieRowRead)].value))
    print("Februar: " + str(ws["C" + str(KategorieRowRead)].value))
    print("März: " + str(ws["D" + str(KategorieRowRead)].value))
    print("April: " + str(ws["E" + str(KategorieRowRead)].value))
    print("Mai: " + str(ws["F" + str(KategorieRowRead)].value))
    print("Juni: " + str(ws["G" + str(KategorieRowRead)].value))
    print("Juli: " + str(ws["H" + str(KategorieRowRead)].value))
    print("August: " + str(ws["I" + str(KategorieRowRead)].value))
    print("September: " + str(ws["J" + str(KategorieRowRead)].value))
    print("Oktober: " + str(ws["K" + str(KategorieRowRead)].value))
    print("November: " + str(ws["L" + str(KategorieRowRead)].value))
    print("Dezember: " + str(ws["M" + str(KategorieRowRead)].value))

    KategorieInp = input("Of which Kategorie do you want to see all values together? ")

    for row in range(2, ws.max_row + 1):
        print(row)
        if KategorieInp == ws.cell(row=row, column=1).value:
            print("Gefunden")
            global KategorieRow
            KategorieRow = row

    test = KategorieRow
    print("zeile der Kategorie: " + str(test))


def ReadOutContentOfMonth(year, month):  # Zahlen weg!

    wb = load_workbook("Finanzen" + str(year) + ".xlsx")  # Erstellt NEUES, bei nur hinzufügen/abändern nutze load_workbook
    ws = wb["Finanzen"]

    MonthInp = month

    for column in range(1, ws.max_column + 1):
        print(column)
        if MonthInp == ws.cell(row=1, column=column).value:
            print("Gefunden")
            global MonthColRead
            MonthColRead = column

    print("----------------------------------")
    print(str(MonthColRead) + " ist die gesuchte Zeile.")
    print("----------------------------------")

    print(MonthInp + ":")
    print("---------")

    for i in range(2, ws.max_row + 1):
        # print(i), print(ws["A" + str(i)].value)

        if MonthColRead == int("1"):
            MonthColRead = "A"
        if MonthColRead == int("2"):
            MonthColRead = "B"
        if MonthColRead == int("3"):
            MonthColRead = "C"
        if MonthColRead == int("4"):
            MonthColRead = "D"
        if MonthColRead == int("5"):
            MonthColRead = "E"
        if MonthColRead == int("6"):
            MonthColRead = "F"
        if MonthColRead == int("7"):
            MonthColRead = "G"
        if MonthColRead == int("8"):
            MonthColRead = "H"
        if MonthColRead == int("9"):
            MonthColRead = "I"
        if MonthColRead == int("10"):
            MonthColRead = "J"
        if MonthColRead == int("11"):
            MonthColRead = "K"
        if MonthColRead == int("12"):
            MonthColRead = "L"
        if MonthColRead == int("13"):
            MonthColRead = "M"

        print(ws["A" + str(i)].value + ": ")
        print(ws[str(MonthColRead) + str(i)].value)  # # zahlen weg!


def NewKategorie(KategorieName):
    #if str(KategorieName).__contains__(",", ".", "*", "~", "#", ":", ";", "/", "1", "2", "3", "4", "5", "6", "7", "8", "9", "0", "!", "§", "$", "²", "³", "%", "&", "(", "{", "[", ")", "]", "=", "}", "?", "ß", "´", "`"):
        #showText("Bitte nutze keine Sonderzeichen.")
    if KategorieList.__contains__(str(KategorieName)):
        showText("Diese Kategorie existiert bereits.")
    else:
        dateiname = "KategorieList.txt"
        element = str(tab3_entry.get())
        hinzufuegen_element(dateiname, element)
        for length in range(0, YearsList.__len__()):
            print(YearsList.__len__())
            YearsListItem = str(YearsList.__getitem__(length))
            print(YearsList.__getitem__(length))

            NewKategorieInp = KategorieName

            wb = load_workbook("Finanzen" + YearsList.__getitem__(length) + ".xlsx")  # Erstellt NEUES, bei nur hinzufügen/abändern nutze load_workbook
            ws = wb["Finanzen"]

            #ws["A" + str(1)].value = NewKategorieInp
            ws["A" + str((int(ws.max_row) + 1))].value = NewKategorieInp
            wb.save("Finanzen" + YearsList.__getitem__(length) + ".xlsx")

            wb = load_workbook(
                "Finanzen" + YearsList.__getitem__(length) + ".xlsx")  # Erstellt NEUES, bei nur hinzufügen/abändern nutze load_workbook
            ws = wb["Finanzen"]
            print("clearing Tabel...")
            for col in range(1, ws.max_column + 1):
                print("---Col: " + str(col))
                if col == int("1"):
                    col = "A"
                if col == int("2"):
                    col = "B"
                if col == int("3"):
                    col = "C"
                if col == int("4"):
                    col = "D"
                if col == int("5"):
                    col = "E"
                if col == int("6"):
                    col = "F"
                if col == int("7"):
                    col = "G"
                if col == int("8"):
                    col = "H"
                if col == int("9"):
                    col = "I"
                if col == int("10"):
                    col = "J"
                if col == int("11"):
                    col = "K"
                if col == int("12"):
                    col = "L"
                if col == int("13"):
                    col = "M"
                for row in range(2, ws.max_row + 1):
                    print(row)
                    if str(ws[str(col) + str(row)].value) == "None":
                        ws[str(col) + str(row)].value = 0
                        print("Changend")

            wb.save("Finanzen" + YearsList.__getitem__(length) + ".xlsx")

        showText("Kategorie ergolgreich hinzugefügt")
        PlsReopenSoftware()


def NewEinnahmeQuelle(QuellenName):
    EinnahmeQuellenListName = "EinnahmequelleList.txt"
    EinnahmeQuelleList = laden_liste(EinnahmeQuellenListName)


    if EinnahmeQuelleList.__contains__(str(QuellenName)):
        showText("Diese Kategorie existiert bereits.")
    #if str(QuellenName).__contains__(",", ".", "*", "~", "#", ":", ";", "/", "1", "2", "3", "4", "5", "6", "7", "8", "9", "0", "!", "§", "$", "²", "³", "%", "&", "(", "{", "[", ")", "]", "=", "}", "?", "ß", "´", "`"):
        #showText("Bitte nutze keine Sonderzeichen.")
    else:
        dateiname = "EinnahmequelleList.txt"
        element = str(QuellenName)
        hinzufuegen_element(dateiname, element)
        for length in range(0, YearsList.__len__()):
            print(YearsList.__len__())
            YearsListItem = str(YearsList.__getitem__(length))
            print(YearsList.__getitem__(length))

            NewQuelleInp = QuellenName

            wb = load_workbook("Gehalt" + YearsList.__getitem__(length) + ".xlsx")
            ws = wb["Gehalt"]
            """print("yyyyyyyyyyyyyyyyyyy" + str(ws["A2"].value))
            if str(ws["A2"].value) == "Gehalt:":
                ws.delete_rows(2)"""

            #ws["A" + str(1)].value = NewKategorieInp
            ws["A" + str((int(ws.max_row) + 1))].value = NewQuelleInp
            wb.save("Gehalt" + YearsList.__getitem__(length) + ".xlsx")

            wb = load_workbook(
                "Gehalt" + YearsList.__getitem__(length) + ".xlsx")  # Erstellt NEUES, bei nur hinzufügen/abändern nutze load_workbook
            ws = wb["Gehalt"]
            print("clearing Tabel...")
            for col in range(1, ws.max_column + 1):
                print("---Col: " + str(col))
                if col == int("1"):
                    col = "A"
                if col == int("2"):
                    col = "B"
                if col == int("3"):
                    col = "C"
                if col == int("4"):
                    col = "D"
                if col == int("5"):
                    col = "E"
                if col == int("6"):
                    col = "F"
                if col == int("7"):
                    col = "G"
                if col == int("8"):
                    col = "H"
                if col == int("9"):
                    col = "I"
                if col == int("10"):
                    col = "J"
                if col == int("11"):
                    col = "K"
                if col == int("12"):
                    col = "L"
                if col == int("13"):
                    col = "M"
                for row in range(2, ws.max_row + 1):
                    print(row)
                    if str(ws[str(col) + str(row)].value) == "None":
                        ws[str(col) + str(row)].value = 0
                        print("Changend")

            wb.save("Gehalt" + YearsList.__getitem__(length) + ".xlsx")

        showText("Einnahme-Quelle ergolgreich hinzugefügt")
        PlsReopenSoftware()


def deleteKategorie(DeleteKategorieName):
    res=messagebox.askquestion('Ausgabekategorie löschen?', f'Soll die Ausgabenkategorie {DeleteKategorieName} wirklich gelöscht werden?')
    if res == 'yes' :
        DeleteKategorieInp = DeleteKategorieName

        dateiname = KategorieListName

        zu_loeschendes_wort = DeleteKategorieName

        # Datei lesen und Inhalt in eine Variable speichern
        with open(dateiname, 'r') as datei:
            zeilen = datei.readlines()

        # Jede Zeile bearbeiten und das zu löschende Wort sowie leere Zeilen entfernen
        with open(dateiname, 'w') as datei:
            for zeile in zeilen:
                # Ersetzt das zu löschende Wort mit nichts, also löscht es
                bearbeitete_zeile = zeile.replace(zu_loeschendes_wort, '')
                # Überprüft, ob die Zeile nach dem Entfernen des Wortes nicht leer ist
                if bearbeitete_zeile.strip():
                    datei.write(bearbeitete_zeile)


        for yearIdx in range(0, YearsList.__len__()):
            with open(YearsListName, "r") as f:
                content = f.read()
                print(content.splitlines()[yearIdx])  # Ausgabe der ersten Zeile
            yearLine = content.splitlines()[yearIdx]
            wb = load_workbook("Finanzen" + str(yearLine) + ".xlsx")  # Erstellt NEUES, bei nur hinzufügen/abändern nutze load_workbook
            ws = wb["Finanzen"]

            for row in range(1, ws.max_row + 1):
                print(row)
                if DeleteKategorieInp == ws.cell(row=row, column=1).value:
                    print("Gefunden")
                    global KategorieRowDelete
                    KategorieRowDelete = row

            ws.delete_rows(KategorieRowDelete)

            wb.save("Finanzen" + str(yearLine) + ".xlsx")

        showText("Kategorie erfolgreich gelöscht")
     
    else :
        print("Abbruch")

    PlsReopenSoftware()
    

def deleteEinnahmequelle(DeleteEinnahmequelle):
    res=messagebox.askquestion('Einnahmequelle löschen?', f'Soll die Einnahmequelle {DeleteEinnahmequelle} wirklich gelöscht werden?')
    if res == 'yes' :
        DeleteEinnahmequelleInp = DeleteEinnahmequelle

        EinnahmeQuellenListName = "EinnahmequelleList.txt"
        EinnahmeQuelleList = laden_liste(EinnahmeQuellenListName)

        dateiname = EinnahmeQuellenListName

        zu_loeschendes_wort = DeleteEinnahmequelle

        # Datei lesen und Inhalt in eine Variable speichern
        with open(dateiname, 'r') as datei:
            zeilen = datei.readlines()

        # Jede Zeile bearbeiten und das zu löschende Wort sowie leere Zeilen entfernen
        with open(dateiname, 'w') as datei:
            for zeile in zeilen:
                # Ersetzt das zu löschende Wort mit nichts, also löscht es
                bearbeitete_zeile = zeile.replace(zu_loeschendes_wort, '')
                # Überprüft, ob die Zeile nach dem Entfernen des Wortes nicht leer ist
                if bearbeitete_zeile.strip():
                    datei.write(bearbeitete_zeile)

        for yearIdx in range(0, YearsList.__len__()):
            with open(YearsListName, "r") as f:
                content = f.read()
                print(content.splitlines()[yearIdx])  # Ausgabe der ersten Zeile
            yearLine = content.splitlines()[yearIdx]
            wb = load_workbook(
                "Gehalt" + str(yearLine) + ".xlsx")  # Erstellt NEUES, bei nur hinzufügen/abändern nutze load_workbook
            ws = wb["Gehalt"]

            for row in range(1, ws.max_row + 1):
                print(row)
                if DeleteEinnahmequelleInp == ws.cell(row=row, column=1).value:
                    print("Gefunden")
                    global KategorieRowDelete
                    KategorieRowDelete = row

            ws.delete_rows(KategorieRowDelete)

            wb.save("Gehalt" + str(yearLine) + ".xlsx")

        showText("Einnahme-Quelle erfolgreich gelöscht")

        PlsReopenSoftware()
    else :
        print("Abbruch")
    

def deleteValue(year, Month, Kategorie):
    res=messagebox.askquestion('Betrag löschen?', f'Soll der Betrag aus der Kategorie {Kategorie} von {Month} {year} wirklich gelöscht werden?')
    if res == 'yes' :
        wb = load_workbook("Finanzen" + str(year) + ".xlsx")  # Erstellt NEUES, bei nur hinzufügen/abändern nutze load_workbook
        ws = wb["Finanzen"]

        DeleteValueOfKategorie = Kategorie
        DeleteValueOfMonth = Month

        print(year)
        print(DeleteValueOfKategorie)
        print(DeleteValueOfMonth)
        for row in range(1, ws.max_row + 1):
            print(row)
            if DeleteValueOfKategorie == ws.cell(row=row, column=1).value:
                print("Gefunden")
                global DeleteValueOfKategorieRow
                DeleteValueOfKategorieRow = row

        for col in range(1, ws.max_column + 1):
            print(col)
            if DeleteValueOfMonth == ws.cell(row=1, column=col).value:
                print("Gefunden")
                global DeleteValueOfMonthCol
                DeleteValueOfMonthCol = col

        if DeleteValueOfMonthCol == int("1"):
            DeleteValueOfMonthCol = "A"
        if DeleteValueOfMonthCol == int("2"):
            DeleteValueOfMonthCol = "B"
        if DeleteValueOfMonthCol == int("3"):
            DeleteValueOfMonthCol = "C"
        if DeleteValueOfMonthCol == int("4"):
            DeleteValueOfMonthCol = "D"
        if DeleteValueOfMonthCol == int("5"):
            DeleteValueOfMonthCol = "E"
        if DeleteValueOfMonthCol == int("6"):
            DeleteValueOfMonthCol = "F"
        if DeleteValueOfMonthCol == int("7"):
            DeleteValueOfMonthCol = "G"
        if DeleteValueOfMonthCol == int("8"):
            DeleteValueOfMonthCol = "H"
        if DeleteValueOfMonthCol == int("9"):
            DeleteValueOfMonthCol = "I"
        if DeleteValueOfMonthCol == int("10"):
            DeleteValueOfMonthCol = "J"
        if DeleteValueOfMonthCol == int("11"):
            DeleteValueOfMonthCol = "K"
        if DeleteValueOfMonthCol == int("12"):
            DeleteValueOfMonthCol = "L"
        if DeleteValueOfMonthCol == int("13"):
            DeleteValueOfMonthCol = "M"

        print(DeleteValueOfKategorieRow, DeleteValueOfMonthCol)


        ws[DeleteValueOfMonthCol + str(DeleteValueOfKategorieRow)].value = 0
        wb.save("Finanzen" + year + ".xlsx")

        showText("Betrag erfolgreich gelöscht.")
    else :
        print("Abbruch")
   

def Insert(year, month, kategorie, value, addOrreset): 
    wb = load_workbook("Finanzen" + year + ".xlsx")  # Erstellt NEUES, bei nur hinzufügen/abändern nutze load_workbook
    ws = wb["Finanzen"]
    InsertToKategorie = kategorie
    InsertToMonth = month
    InsertPrice = value
    InsertOrAdd = addOrreset

    for row in range(1, ws.max_row + 1):
        print(row)
        if InsertToKategorie == ws.cell(row=row, column=1).value:
            print("Gefunden")
            global KategorieRowInsert
            KategorieRowInsert = row

    for col in range(1, ws.max_column + 1):
        print(col)
        if InsertToMonth == ws.cell(row=1, column=col).value:
            print("Gefunden")
            global MonthRowInsert
            MonthRowInsert = col

    print(str(KategorieRowInsert) + " ist die Zeile der Kategorie " + InsertToKategorie)
    print(str(MonthRowInsert) + " ist die Zeile des Monats " + InsertToMonth)

    if MonthRowInsert == int("1"):
        MonthRowInsert = "A"
    if MonthRowInsert == int("2"):
        MonthRowInsert = "B"
    if MonthRowInsert == int("3"):
        MonthRowInsert = "C"
    if MonthRowInsert == int("4"):
        MonthRowInsert = "D"
    if MonthRowInsert == int("5"):
        MonthRowInsert = "E"
    if MonthRowInsert == int("6"):
        MonthRowInsert = "F"
    if MonthRowInsert == int("7"):
        MonthRowInsert = "G"
    if MonthRowInsert == int("8"):
        MonthRowInsert = "H"
    if MonthRowInsert == int("9"):
        MonthRowInsert = "I"
    if MonthRowInsert == int("10"):
        MonthRowInsert = "J"
    if MonthRowInsert == int("11"):
        MonthRowInsert = "K"
    if MonthRowInsert == int("12"):
        MonthRowInsert = "L"
    if MonthRowInsert == int("13"):
        MonthRowInsert = "M"

    print(str(MonthRowInsert) + " ist der Buchstabe des Monats " + InsertToMonth)

    if InsertOrAdd == "Addieren":
        PriceNow = ws[MonthRowInsert + str(KategorieRowInsert)].value
        if str(PriceNow) == "None":
            ws[MonthRowInsert + str(KategorieRowInsert)].value = 0
        Price = int(ws[MonthRowInsert + str(KategorieRowInsert)].value) + float(InsertPrice) #int(InsertPrice)
        ws[MonthRowInsert + str(KategorieRowInsert)].value = Price
    if InsertOrAdd == "Ersetzen":
        ws[MonthRowInsert + str(KategorieRowInsert)].value = InsertPrice

    wb.save("Finanzen" + year + ".xlsx")

    showText("Betrag erfolgreich hinzugefügt")


def getValueOfMonthAllKategories(year, month):
    MonthInp = month

    wb = load_workbook("Finanzen" + str(year) + ".xlsx")  # Erstellt NEUES, bei nur hinzufügen/abändern nutze load_workbook
    ws = wb["Finanzen"]

    wbGehalt = load_workbook("Gehalt" + year + ".xlsx")
    wsGehalt = wbGehalt["Gehalt"]

    # Check ob Gehalt überall schon eingetragen
    gehaltAll = 0
    for colGehalt in range(2, wsGehalt.max_column + 1):
        if colGehalt == int("1"):
            colGehalt = "A"
        if colGehalt == int("2"):
            colGehalt = "B"
        if colGehalt == int("3"):
            colGehalt = "C"
        if colGehalt == int("4"):
            colGehalt = "D"
        if colGehalt == int("5"):
            colGehalt = "E"
        if colGehalt == int("6"):
            colGehalt = "F"
        if colGehalt == int("7"):
            colGehalt = "G"
        if colGehalt == int("8"):
            colGehalt = "H"
        if colGehalt == int("9"):
            colGehalt = "I"
        if colGehalt == int("10"):
            colGehalt = "J"
        if colGehalt == int("11"):
            colGehalt = "K"
        if colGehalt == int("12"):
            colGehalt = "L"
        if colGehalt == int("13"):
            colGehalt = "M"

        gehaltdieserMonat = 0
        for row in range(2, wsGehalt.max_row + 1):
            print(row)
            gehaltdieserMonat += round(float(wsGehalt[colGehalt + str(row)].value), 2)
            gehaltAll += gehaltdieserMonat
        print("gehalt dieser Monat: " + str(gehaltdieserMonat))

        if str(gehaltdieserMonat) == "0.0":
            print("Gehalt noch nicht eingetragen für den Monat " + str(wsGehalt[str(colGehalt) + str(1)].value))
            #showText("Das Gehalt für den Monat " + str(wsGehalt[str(colGehalt) + str(
                #1)].value) + " wurde noch nicht eingetragen. Gehe dafür bitte zum Hauptfenster zurück und füge dein Gehalt links unten, unter <Gehalt> hinzu.")
        print(round(gehaltAll, 2))

    for col in range(1, ws.max_column + 1):
        print(col)
        if MonthInp == ws.cell(row=1, column=col).value:
            print("Gefunden")
            global MonthCol
            MonthCol = col

    if MonthCol == int("1"):
        MonthCol = "A"
    if MonthCol == int("2"):
        MonthCol = "B"
    if MonthCol == int("3"):
        MonthCol = "C"
    if MonthCol == int("4"):
        MonthCol = "D"
    if MonthCol == int("5"):
        MonthCol = "E"
    if MonthCol == int("6"):
        MonthCol = "F"
    if MonthCol == int("7"):
        MonthCol = "G"
    if MonthCol == int("8"):
        MonthCol = "H"
    if MonthCol == int("9"):
        MonthCol = "I"
    if MonthCol == int("10"):
        MonthCol = "J"
    if MonthCol == int("11"):
        MonthCol = "K"
    if MonthCol == int("12"):
        MonthCol = "L"
    if MonthCol == int("13"):
        MonthCol = "M"

    AllValuesTogether = 0

    Kategorien = []
    Beträge = []

    for row in range(2, ws.max_row + 1):
        print(ws["A" + str(row)].value)
        Kategorien.insert(int(Kategorien.__len__()) + 1, ws["A" + str(row)].value)

        if str(ws[MonthCol + str(row)].value) == "None":
            ws[MonthCol + str(row)].value = 0

        print(ws[MonthCol + str(row)].value)
        Beträge.insert(int(Kategorien.__len__()) + 1, ws[MonthCol + str(row)].value)
        AllValuesTogether += float(ws[MonthCol + str(row)].value)


    print("---------------")
    print("Zusammen:")
    print(AllValuesTogether)


    print(Kategorien)
    print(Beträge)

    # NEUES FENSTER WIRD ERSTELLT ZUM ANZEIGEN DER BETRÄGE
    getValueOfMonthAllKategoriesWind = tk.Tk()
    getValueOfMonthAllKategoriesWind.title("Monat ausgelesen: ")


    entryMonat = tk.Label(getValueOfMonthAllKategoriesWind, text=MonthInp, font=("Arial", 20))
    entryMonat.pack()

    for KategorienLen in range(0, Kategorien.__len__()):
        frame = tk.Frame(getValueOfMonthAllKategoriesWind)
        frame.pack()
        entry1 = tk.Entry(frame, font=("Arial", 15))
        entry1.pack(side=tk.LEFT)
        entry1.insert(0, str(Kategorien.__getitem__(KategorienLen)) + ": ")

        entry2 = tk.Entry(frame, font=("Arial", 15))
        entry2.pack(side=tk.RIGHT)
        entry2.insert(0,str(float(Beträge.__getitem__(KategorienLen))))

    entryAll = tk.Entry(getValueOfMonthAllKategoriesWind, font=("Arial", 20), width=30)
    entryAll.pack(pady=10, padx=10)
    entryAll.insert(0, float(AllValuesTogether))
    entryAll.insert(0, "Summe Ausgaben: ")

    for row in range(2, wsGehalt.max_row + 1):
        print(row)
        gehaltdieserMonat += round(float(wsGehalt[MonthCol + str(row)].value), 2)
        gehaltAll += gehaltdieserMonat
    print("Summe Einnahmen: " + str(gehaltdieserMonat))


    entryGehalt = tk.Entry(getValueOfMonthAllKategoriesWind, font=("Arial", 20), width=30)
    entryGehalt.pack(pady=10, padx=10)
    entryGehalt.insert(0, str(float(gehaltdieserMonat)))
    entryGehalt.insert(0, "Summe Einnahmen: ")

    Relation = float(gehaltdieserMonat) - float(AllValuesTogether)
    print(Relation)
    Relation = round(Relation, 2)

    if Relation < 0:
        print("Verlust")
        entryRelation = tk.Entry(getValueOfMonthAllKategoriesWind, font=("Arial", 20), width=30, fg="red")
        entryRelation.pack(pady=10, padx=10)
        entryRelation.insert(0, str(Relation))
        entryRelation.insert(0, "Verlust von ")
    else:  # Auch wenn Relation == 0
        print("Gewinn")
        entryRelation = tk.Entry(getValueOfMonthAllKategoriesWind, font=("Arial", 20), width=30, fg="green")
        entryRelation.pack(pady=10, padx=10)
        entryRelation.insert(0, str(Relation))
        entryRelation.insert(0, "Gewinn von ")



    getValueOfMonthAllKategoriesWind.mainloop()


def getValueOfKategorieAllMonths(year, kategorie):
    KategorieInp = kategorie

    wb = load_workbook("Finanzen" + str(year) + ".xlsx")  # Erstellt NEUES, bei nur hinzufügen/abändern nutze load_workbook
    ws = wb["Finanzen"]

    for row in range(2, ws.max_row + 1):
        print(row)
        if KategorieInp == ws.cell(row=row, column=1).value:
            print("Gefunden")
            global KategorieRow
            KategorieRow = row

    test = KategorieRow
    print("zeile der Kategorie: " + str(test))


    AllValuesTogether = 0
    Beträge = []
    Monate = []

    for col in range(2, ws.max_column + 1):
        if col == int("1"):
            col = "A"
        if col == int("2"):
            col = "B"
        if col == int("3"):
            col = "C"
        if col == int("4"):
            col = "D"
        if col == int("5"):
            col = "E"
        if col == int("6"):
            col = "F"
        if col == int("7"):
            col = "G"
        if col == int("8"):
            col = "H"
        if col == int("9"):
            col = "I"
        if col == int("10"):
            col = "J"
        if col == int("11"):
            col = "K"
        if col == int("12"):
            col = "L"
        if col == int("13"):
            col = "M"

        if str(ws[str(col) + str(KategorieRow)].value) == "None":
            ws[str(col) + str(KategorieRow)].value = 0

        print(str(ws[str(col) + "1"].value) + ":")
        Monate.insert(int(Monate.__len__()) + 1, ws[str(col) + "1"].value)
        print(ws[str(col) + str(KategorieRow)].value)
        AllValuesTogether += float(ws[col + str(KategorieRow)].value)
        Beträge.insert(int(Monate.__len__()) + 1, ws[str(col) + str(KategorieRow)].value)


    print("---------------")
    print("Zusammen:")
    print(AllValuesTogether)

    print(Monate)
    print(Beträge)

    # NEUES FENSTER WIRD ERSTELLT ZUM ANZEIGEN DER BETRÄGE
    getValueOfKategorieAllMonthsWind = tk.Tk()
    getValueOfKategorieAllMonthsWind.title("Kategorie ausgelesen: ")
    getValueOfKategorieAllMonthsWind.minsize(width="200", height="0")

    entryMonat = tk.Label(getValueOfKategorieAllMonthsWind, text=KategorieInp, font=("Arial", 20))
    entryMonat.pack()

    for MonateLen in range(0, Monate.__len__()):

        frame = tk.Frame(getValueOfKategorieAllMonthsWind)
        frame.pack()
        entry1 = tk.Entry(frame, font=("Arial", 15))
        entry1.pack(side=tk.LEFT)
        entry1.insert(0, str(Monate.__getitem__(MonateLen)) + ": ")

        entry2 = tk.Entry(frame, font=("Arial", 15))
        entry2.pack(side=tk.RIGHT)
        entry2.insert(0, str(float(Beträge.__getitem__(MonateLen))))

    entryAll = tk.Entry(getValueOfKategorieAllMonthsWind, font=("Arial", 20), width=30)
    entryAll.pack(pady=10, padx=10)
    entryAll.insert(0, float(AllValuesTogether))
    entryAll.insert(0, "Summe Ausgaben: ")

    getValueOfKategorieAllMonthsWind.mainloop()


def getValueAll(year):
    wb = load_workbook("Finanzen" + year + ".xlsx")  # Erstellt NEUES, bei nur hinzufügen/abändern nutze load_workbook
    ws = wb["Finanzen"]
    global AllValuesTogether
    AllValuesTogether = 0
    for col in range(2, ws.max_column + 1):
        print("-----Col: " + str(col))
        if col == int("1"):
            col = "A"
        if col == int("2"):
            col = "B"
        if col == int("3"):
            col = "C"
        if col == int("4"):
            col = "D"
        if col == int("5"):
            col = "E"
        if col == int("6"):
            col = "F"
        if col == int("7"):
            col = "G"
        if col == int("8"):
            col = "H"
        if col == int("9"):
            col = "I"
        if col == int("10"):
            col = "J"
        if col == int("11"):
            col = "K"
        if col == int("12"):
            col = "L"
        if col == int("13"):
            col = "M"

        for row in range(2, ws.max_row + 1):
            print(row)

            AllValuesTogether += float(ws[str(col) + str(row)].value)

    print("---------------")
    print("Zusammen:")
    AllValuesTogether = round(AllValuesTogether, 2)
    print(AllValuesTogether)
    return AllValuesTogether


def clearTabel(year):
    wb = load_workbook("Finanzen" + year + ".xlsx")  # Erstellt NEUES, bei nur hinzufügen/abändern nutze load_workbook
    ws = wb["Finanzen"]
    print("clearing Tabel...")
    for col in range(1, ws.max_column + 1):
        print("---Col: " + str(col))
        if col == int("1"):
            col = "A"
        if col == int("2"):
            col = "B"
        if col == int("3"):
            col = "C"
        if col == int("4"):
            col = "D"
        if col == int("5"):
            col = "E"
        if col == int("6"):
            col = "F"
        if col == int("7"):
            col = "G"
        if col == int("8"):
            col = "H"
        if col == int("9"):
            col = "I"
        if col == int("10"):
            col = "J"
        if col == int("11"):
            col = "K"
        if col == int("12"):
            col = "L"
        if col == int("13"):
            col = "M"
        for row in range(2, ws.max_row + 1):
            print(row)
            if str(ws[str(col) + str(row)].value) == "None":
                ws[str(col) + str(row)].value = 0
                print("Changend")
            if str(ws[str(col) + str(row)].value) == "0.0":
                ws[str(col) + str(row)].value = 0
                print("Changend")

    wb.save("Finanzen" + year + ".xlsx")

    PlsReopenSoftware()


def GetAllMonths(year):
    wb = load_workbook("Finanzen" + year + ".xlsx")
    ws = wb["Finanzen"]

    wbGehalt = load_workbook("Gehalt" + year + ".xlsx")
    wsGehalt = wbGehalt["Gehalt"]

    global GehaltListe
    GehaltListe = []

    gehaltAll = 0
    for col in range(2, wsGehalt.max_column + 1):
        if col == int("1"):
            col = "A"
        if col == int("2"):
            col = "B"
        if col == int("3"):
            col = "C"
        if col == int("4"):
            col = "D"
        if col == int("5"):
            col = "E"
        if col == int("6"):
            col = "F"
        if col == int("7"):
            col = "G"
        if col == int("8"):
            col = "H"
        if col == int("9"):
            col = "I"
        if col == int("10"):
            col = "J"
        if col == int("11"):
            col = "K"
        if col == int("12"):
            col = "L"
        if col == int("13"):
            col = "M"

        gehaltDieserMonat = 0
        for row in range(2, wsGehalt.max_row + 1):
            print(row)
            gehaltDieserMonat += round(float(wsGehalt[str(col) + str(row)].value), 2)
        print("Gehalt dieser Monat: " + str(gehaltDieserMonat))
        GehaltListe.append(gehaltDieserMonat)

        gehaltAll += gehaltDieserMonat
        gehaltAll = round(gehaltAll, 2)
        if str(gehaltDieserMonat) == "0.0":
            print("Gehalt noch nicht eingetragen für den Monat " + str(wsGehalt[str(col) + str(1)].value))
            #showText("Das Gehalt für den Monat " + str(wsGehalt[str(col) + str(
                #1)].value) + " wurde noch nicht eingetragen. Gehe dafür bitte zum Hauptfenster zurück und füge dein Gehalt links unten, unter <Gehalt> hinzu.")
        print("Gehalt all:" + str(gehaltAll))

    global AllMonthsValuesTogether
    AllMonthsValuesTogether = 0

    getValueOfAllMonthsWind = tk.Tk()
    getValueOfAllMonthsWind.title("Jahr: " + year + ", alle Monate")

    JahrTitle = tk.Label(getValueOfAllMonthsWind, text="Alle Monate aus dem Jahr " + year, font=("Arial", 20))
    JahrTitle.pack()

    frame = tk.Frame(getValueOfAllMonthsWind)
    frame.pack(side="top")
    JahreLabel = tk.Entry(frame, font=("Arial", 15))
    JahreLabel.pack(side="left")
    JahreLabel.insert(0, "Jahre:")
    AusgabenLabel = tk.Entry(frame, font=("Arial", 15))
    AusgabenLabel.pack(side="right")
    AusgabenLabel.insert(0, "Ausgaben:")
    EinnahmenLabel = tk.Entry(frame, font=("Arial", 15))
    EinnahmenLabel.pack(side="right")
    EinnahmenLabel.insert(0, "Einnahmen:")

    for col in range(2, ws.max_column + 1):
        if col == int("1"):
            col = "A"
        if col == int("2"):
            col = "B"
        if col == int("3"):
            col = "C"
        if col == int("4"):
            col = "D"
        if col == int("5"):
            col = "E"
        if col == int("6"):
            col = "F"
        if col == int("7"):
            col = "G"
        if col == int("8"):
            col = "H"
        if col == int("9"):
            col = "I"
        if col == int("10"):
            col = "J"
        if col == int("11"):
            col = "K"
        if col == int("12"):
            col = "L"
        if col == int("13"):
            col = "M"
        print(col)
        MonthValue = 0

        for row in range(2, ws.max_row + 1):
            print(row)
            MonthValue += float(ws[str(col) + str(row)].value)



        frame = tk.Frame(getValueOfAllMonthsWind)
        frame.pack()
        entry1 = tk.Entry(frame, font=("Arial", 15))
        entry1.pack(side=tk.LEFT)
        entry1.insert(0, str(ws[str(col) + str(1)].value) + ": ")

        entry2 = tk.Entry(frame, font=("Arial", 15))
        entry2.pack(side=tk.RIGHT)
        entry2.insert(0, str(MonthValue))

        if col == "A":
            col2 = -1
        if col == "B":
            col2 = 0
        if col == "C":
            col2 = 1
        if col == "D":
            col2 = 2
        if col == "E":
            col2 = 3
        if col == "F":
            col2 = 4
        if col == "G":
            col2 = 5
        if col == "H":
            col2 = 6
        if col == "I":
            col2 = 7
        if col == "J":
            col2 = 8
        if col == "K":
            col2 = 9
        if col == "L":
            col2 = 10
        if col == "M":
            col2 = 11

        entry3 = tk.Entry(frame, font=("Arial", 15))
        entry3.pack(side=tk.RIGHT)
        entry3.insert(0, str(float(GehaltListe.__getitem__(col2))))

        print("MonthValue: " + str(MonthValue))
        AllMonthsValuesTogether += float(MonthValue)
        print(round(AllMonthsValuesTogether, 2), f"= round({AllMonthsValuesTogether}, 2)")
        AllMonthsValuesTogether = round(AllMonthsValuesTogether, 2)

    entryAll = tk.Entry(getValueOfAllMonthsWind, font=("Arial", 20), width=30)
    entryAll.pack(pady=10, padx=10)
    entryAll.insert(0, str(float(AllMonthsValuesTogether)))
    entryAll.insert(0, "Summe Ausgaben: ")



    # Gehalt für das Jahr berechnen und anzeigen -> Relation zu den Ausgaben

    # Gehalt Check = 0 schon oben, damit das eigentliche Fenster sicher öffnet und gleich sichtbar ist

    entryGehalt = tk.Entry(getValueOfAllMonthsWind, font=("Arial", 20), width=30)
    entryGehalt.pack(pady=10, padx=10)
    entryGehalt.insert(0, str(float(gehaltAll)))
    entryGehalt.insert(0, "Summe Einnahmen: ")

    Relation = float(gehaltAll) - float(AllMonthsValuesTogether)
    print(Relation)
    Relation = round(Relation, 2)


    if Relation < 0:
        print("Verlust")
        entryRelation = tk.Entry(getValueOfAllMonthsWind, font=("Arial", 20), width=30, fg="red")
        entryRelation.pack(pady=10, padx=10)
        entryRelation.insert(0, str(Relation))
        entryRelation.insert(0, "Verlust von ")
    else:                  # Auch wenn Relation == 0
        print("Gewinn")
        entryRelation = tk.Entry(getValueOfAllMonthsWind, font=("Arial", 20), width=30, fg="green")
        entryRelation.pack(pady=10, padx=10)
        entryRelation.insert(0, str(Relation))
        entryRelation.insert(0, "Gewinn von ")

    getValueOfAllMonthsWind.mainloop()


def GetAllKategories(year):
    wb = load_workbook("Finanzen" + year + ".xlsx")
    ws = wb["Finanzen"]

    wbGehalt = load_workbook("Gehalt" + year + ".xlsx")
    wsGehalt = wbGehalt["Gehalt"]

    gehaltAll = 0
    for col in range(2, wsGehalt.max_column + 1):
        if col == int("1"):
            col = "A"
        if col == int("2"):
            col = "B"
        if col == int("3"):
            col = "C"
        if col == int("4"):
            col = "D"
        if col == int("5"):
            col = "E"
        if col == int("6"):
            col = "F"
        if col == int("7"):
            col = "G"
        if col == int("8"):
            col = "H"
        if col == int("9"):
            col = "I"
        if col == int("10"):
            col = "J"
        if col == int("11"):
            col = "K"
        if col == int("12"):
            col = "L"
        if col == int("13"):
            col = "M"

        gehaltDieserMonat = 0
        for row in range(2, wsGehalt.max_row + 1):
            print(row)
            gehaltDieserMonat += round(float(wsGehalt[str(col) + str(row)].value), 2)
        print("Gehalt dieser Monat: " + str(gehaltDieserMonat))

        gehaltAll += gehaltDieserMonat
        gehaltAll = round(gehaltAll, 2)
        if str(gehaltDieserMonat) == "0.0":
            print("Gehalt noch nicht eingetragen für den Monat " + str(wsGehalt[str(col) + str(1)].value))
            #showText("Das Gehalt für den Monat " + str(wsGehalt[str(col) + str(
                #1)].value) + " wurde noch nicht eingetragen. Gehe dafür bitte zum Hauptfenster zurück und füge dein Gehalt links unten, unter <Gehalt> hinzu.")
        print("Gehalt all:" + str(gehaltAll))

    global AllKategorienValuesTogether
    AllKategorienValuesTogether = 0

    getValueOfAllKategorienWind = tk.Tk()
    getValueOfAllKategorienWind.title("Jahr: " + year + ", alle Kategorien")

    JahrTitle = tk.Label(getValueOfAllKategorienWind, text="Alle Kategorien aus dem Jahr " + year, font=("Arial", 20))
    JahrTitle.pack()

    for row in range(2, ws.max_row + 1):
        print(row)
        KategorieValue = 0

        for col in range(2, ws.max_column + 1):
            print(col)
            if col == int("1"):
                col = "A"
            if col == int("2"):
                col = "B"
            if col == int("3"):
                col = "C"
            if col == int("4"):
                col = "D"
            if col == int("5"):
                col = "E"
            if col == int("6"):
                col = "F"
            if col == int("7"):
                col = "G"
            if col == int("8"):
                col = "H"
            if col == int("9"):
                col = "I"
            if col == int("10"):
                col = "J"
            if col == int("11"):
                col = "K"
            if col == int("12"):
                col = "L"
            if col == int("13"):
                col = "M"
            KategorieValue += float(ws[str(col) + str(row)].value)

        frame = tk.Frame(getValueOfAllKategorienWind)
        frame.pack()
        entry1 = tk.Entry(frame, font=("Arial", 15))
        entry1.pack(side=tk.LEFT)
        entry1.insert(0, str(ws[str("A") + str(row)].value) + ": ")

        entry2 = tk.Entry(frame, font=("Arial", 15))
        entry2.pack(side=tk.RIGHT)
        entry2.insert(0, str(KategorieValue))

        print("MonthValue: " + str(KategorieValue))
        AllKategorienValuesTogether += float(KategorieValue)
        print(round(AllKategorienValuesTogether, 2), f"= round({AllKategorienValuesTogether}, 2)")
        AllKategorienValuesTogether = round(AllKategorienValuesTogether, 2)

    entryAll = tk.Entry(getValueOfAllKategorienWind, font=("Arial", 20), width=30)
    entryAll.pack(pady=10, padx=10)
    entryAll.insert(0, str(float(AllKategorienValuesTogether)))
    entryAll.insert(0, "Summe Ausgaben: ")

    # Gehalt für das Jahr berechnen und anzeigen -> Relation zu den Ausgaben

    # Gehalt Check = 0 schon oben, damit das eigentliche Fenster sicher öffnet und gleich sichtbar ist

    entryGehalt = tk.Entry(getValueOfAllKategorienWind, font=("Arial", 20), width=30)
    entryGehalt.pack(pady=10, padx=10)
    entryGehalt.insert(0, str(float(gehaltAll)))
    entryGehalt.insert(0, "Summe Einnahmen: ")

    Relation = float(gehaltAll) - float(AllKategorienValuesTogether)
    print(Relation)
    Relation = round(Relation, 2)

    if Relation < 0:
        print("Verlust")
        entryRelation = tk.Entry(getValueOfAllKategorienWind, font=("Arial", 20), width=30, fg="red")
        entryRelation.pack(pady=10, padx=10)
        entryRelation.insert(0, str(Relation))
        entryRelation.insert(0, "Verlust von ")
    else:  # Auch wenn Relation == 0
        print("Gewinn")
        entryRelation = tk.Entry(getValueOfAllKategorienWind, font=("Arial", 20), width=30, fg="green")
        entryRelation.pack(pady=10, padx=10)
        entryRelation.insert(0, str(Relation))
        entryRelation.insert(0, "Gewinn von ")

    getValueOfAllKategorienWind.mainloop()


def showCoder():
    messagebox.showinfo("Info", "Von: Jonas Gaiser"  '\n'
                        "---------------------------------" '\n'
                        "Kontakt:" \
                        "" '\n'
                        "E-Mail: jonas.gaiser@online.de")


def showInfosAboutPythonWorkbook():
    print(wb.path)
    print(wb.index)
    print(wb.properties)
    print(wb.code_name)
    print(wb.calculation)
    print(wb.encoding)
    print(wb.security)
    messagebox.showinfo("Info", "Infos über Python/Workbook"  '\n'
                        "---------------------------------" '\n'
                        "path: " '\n' + str(wb.path) 
                        + "index: " '\n' + str(wb.index)
                        + "properties: " '\n' + str(wb.properties)
                        + "code_name: " '\n' + str(wb.code_name)
                        + "calculation: " '\n' + str(wb.calculation)
                        + "encoding: " '\n' + str(wb.encoding)
                        + "security: " '\n' + str(wb.security))





def empty_fields():
    messagebox.showwarning("Info", "Please fill out all the fields.")


def FarbcodesErklärung():
    messagebox.showinfo("Farbcodes", "Die Farbcodes dienen zur Übersicht für die Nutzer*innen. \n" 
                        "Dabei weisen die unterschiedlichen Farben auf unterschiedliche Funktionsbereiche hin. " \
                        "So gilt grau für die Markierung von Darstellungen, lila zur Markierung von Ausgaben, " \
                        "grün zur Markierung von Einnahmen und blau zur Markierung von Umrandungen von Darstellungen.")


def InterestFacts():
    showText("Fakten zu Diagrammen \n"
             "---------------------------\n"
             "1. Diagramme sind keine Details-Anzeigen. " \
                "Sie dienen jediglich zur übersichtlichen Darstellung " \
                "von Werten. \n"
             "2. Diagramme kannst du auch abspeichern. \n"
                "Sie können durch Aufrufen und anschließendem Drücken des " \
                "Speichern-Symbols unten links bei einem beliebigen Ort abgespeichert werden." )





# Check For Open Entrys

def CheckForOpenEntrys():
    CheckForOpenEntrysWind = tk.Tk()
    CheckForOpenEntrysWind.geometry("600x400")
    CheckForOpenEntrysWind.title("Leere Einträge Suche")

    frameCheckOpenEntry = tk.Frame(CheckForOpenEntrysWind, bg="#606c84")
    frameCheckOpenEntry.pack(fill="both", expand=True)
    CheckFrame = tk.Frame(frameCheckOpenEntry, bg = "white")
    CheckFrame.pack(fill="both", expand=True, padx=10, pady=10)

    title = tk.Label(CheckFrame, text="Suche nach Leere Einträge", font=("Arial", 25), bg="white", fg="black")
    title.pack(side="top")
    title2 = tk.Label(CheckFrame, text="Wenn keine Reaktion nach dem \n Drücken eines der beiden Knöpfe folgt, \n wurde noch keine \nEinnahmequelle/Ausgabenkategorie \n benannt.", font=("Arial", 15), bg="white", fg="black")
    title2.pack(side="top")

    Btn1 = tk.Button(CheckFrame, text="Suche nach Leere \n Einnahmeinträge", font=("Arial", 20), bg="#F25E24", fg="black", command=lambda: [CheckForOpenEntrysWind.destroy(), CheckForSalaryInp(), CheckForOpenEntrys()])
    Btn1.pack(side="left", expand=True, fill="both", padx=7.5, pady=7.5)

    Btn2 = tk.Button(CheckFrame, text="Suche nach Leere \n Ausgabeneinträge", font=("Arial", 20), bg="#F25E24", fg="black", command= lambda: [CheckForOpenEntrysWind.destroy(), CheckForExpensesInp(), CheckForOpenEntrys()])
    Btn2.pack(side="left", expand=True, fill="both", padx=7.5, pady=7.5)

    CheckForOpenEntrysWind.mainloop()


# Check for Open salary inps
def CheckForSalaryInp():
    print(YearsList.__len__())
    for length in range(0, YearsList.__len__()):
        YearsListItem = str(YearsList.__getitem__(length))
        print(YearsList.__getitem__(length))


        wb = load_workbook(
            "Gehalt" + YearsList.__getitem__(length) + ".xlsx")  # Erstellt NEUES, bei nur hinzufügen/abändern nutze load_workbook
        ws = wb["Gehalt"]



        for col in range(1, ws.max_column + 1):
            print("---Col: " + str(col))
            if col == int("1"):
                col = "A"
            if col == int("2"):
                col = "B"
            if col == int("3"):
                col = "C"
            if col == int("4"):
                col = "D"
            if col == int("5"):
                col = "E"
            if col == int("6"):
                col = "F"
            if col == int("7"):
                col = "G"
            if col == int("8"):
                col = "H"
            if col == int("9"):
                col = "I"
            if col == int("10"):
                col = "J"
            if col == int("11"):
                col = "K"
            if col == int("12"):
                col = "L"
            if col == int("13"):
                col = "M"
            for row in range(2, ws.max_row + 1):
                print(row)
                if str(ws[str(col) + str(row)].value) == "None":
                    ws[str(col) + str(row)].value = 0
                    print("Changend")
                if str(ws[str(col) + str(row)].value) == "0" or str(ws[str(col) + str(row)].value) == "0.0":
                    showWarn(f"Einnahmen für den Monat " + str(ws[str(col) + str(1)].value) + " im Jahr " + YearsList.__getitem__(length) + " wurden noch nicht eingetragen. Gehe dafür bitte zum Hauptfenster zurück und füge deine Einnahmen rechts unten, unter <Einnahmen> hinzu.")
                    print("Einnahmen noch nicht eingetragen für den Monat " + str(ws[str(col) + str(1)].value))

        wb.save("Gehalt" + YearsList.__getitem__(length) + ".xlsx")

def CheckForExpensesInp():
    for length in range(0, YearsList.__len__()):
        print(YearsList.__len__())
        YearsListItem = str(YearsList.__getitem__(length))
        print(YearsList.__getitem__(length))


        wb = load_workbook("Finanzen" + YearsList.__getitem__(length) + ".xlsx")  # Erstellt NEUES, bei nur hinzufügen/abändern nutze load_workbook
        ws = wb["Finanzen"]

        wb = load_workbook(
            "Finanzen" + YearsList.__getitem__(length) + ".xlsx")  # Erstellt NEUES, bei nur hinzufügen/abändern nutze load_workbook
        ws = wb["Finanzen"]
        print("clearing Tabel...")
        for col in range(1, ws.max_column + 1):
            print("---Col: " + str(col))
            if col == int("1"):
                col = "A"
            if col == int("2"):
                col = "B"
            if col == int("3"):
                col = "C"
            if col == int("4"):
                col = "D"
            if col == int("5"):
                col = "E"
            if col == int("6"):
                col = "F"
            if col == int("7"):
                col = "G"
            if col == int("8"):
                col = "H"
            if col == int("9"):
                col = "I"
            if col == int("10"):
                col = "J"
            if col == int("11"):
                col = "K"
            if col == int("12"):
                col = "L"
            if col == int("13"):
                col = "M"
            for row in range(2, ws.max_row + 1):
                print(row)
                if str(ws[str(col) + str(row)].value) == "None":
                    ws[str(col) + str(row)].value = 0
                    print("Changend")
                if str(ws[str(col) + str(row)].value) == "0" or str(ws[str(col) + str(row)].value) == "0.0":
                    showWarn(f"Ausgaben für den Monat " + str(ws[str(col) + str(1)].value) + " im Jahr " + YearsList.__getitem__(length) + " wurden noch nicht eingetragen. Gehe dafür bitte zum Hauptfenster zurück und füge deine Ausgaben rechts unten, unter <Ausgaben> hinzu.")
                    print("Ausgaben noch nicht eingetragen für den Monat " + str(ws[str(col) + str(1)].value))

        wb.save("Finanzen" + YearsList.__getitem__(length) + ".xlsx")






# GEHALT

def GehaltInsert(year, month, quelle, value): 
    wb = load_workbook("Gehalt" + year + ".xlsx")
    ws = wb["Gehalt"]

    YearsListName = "YearsList.txt"
    YearsList = laden_liste(YearsListName)  # Checkt ob Years Liste vorhanden ist


    for col in range(1, ws.max_column + 1):
        print(col)
        if month == ws.cell(row=1, column=col).value:
            print("Gefunden")
            global MonthColInsert
            MonthColInsert = col

    print(str(MonthColInsert) + " ist die Zeile des Monats " + month)

    if MonthColInsert == int("1"):
        MonthColInsert = "A"
    if MonthColInsert == int("2"):
        MonthColInsert = "B"
    if MonthColInsert == int("3"):
        MonthColInsert = "C"
    if MonthColInsert == int("4"):
        MonthColInsert = "D"
    if MonthColInsert == int("5"):
        MonthColInsert = "E"
    if MonthColInsert == int("6"):
        MonthColInsert = "F"
    if MonthColInsert == int("7"):
        MonthColInsert = "G"
    if MonthColInsert == int("8"):
        MonthColInsert = "H"
    if MonthColInsert == int("9"):
        MonthColInsert = "I"
    if MonthColInsert == int("10"):
        MonthColInsert = "J"
    if MonthColInsert == int("11"):
        MonthColInsert = "K"
    if MonthColInsert == int("12"):
        MonthColInsert = "L"
    if MonthColInsert == int("13"):
        MonthColInsert = "M"

    for row in range(1, ws.max_row + 1):
        print(row)
        if quelle == ws.cell(row=row, column=1).value:
            print("Gefunden")
            global MonthRowInsert
            MonthRowInsert = row

    print(str(MonthColInsert) + " ist der Buchstabe des Monats " + month)


    ws[str(MonthColInsert) + str(MonthRowInsert)].value = value

    wb.save("Gehalt" + year + ".xlsx")

    showText("Gehalt erfolgreich hinzugefügt")


def GehaltDelete(year, month, quelle):
    res=messagebox.askquestion('Einnahmen Löschen?', f'Sollen die Einnahmen des {quelle} von {month} {year} gelöscht werden?')
    if res == 'yes' :
        wb = load_workbook("Gehalt" + year + ".xlsx")
        ws = wb["Gehalt"]

        YearsListName = "YearsList.txt"
        YearsList = laden_liste(YearsListName)  # Checkt ob Years Liste vorhanden ist

        for col in range(1, ws.max_column + 1):
            print(col)
            if month == ws.cell(row=1, column=col).value:
                print("Gefunden")
                global MonthColInsert
                MonthColInsert = col

        print(str(MonthColInsert) + " ist die Zeile des Monats " + month)

        if MonthColInsert == int("1"):
            MonthColInsert = "A"
        if MonthColInsert == int("2"):
            MonthColInsert = "B"
        if MonthColInsert == int("3"):
            MonthColInsert = "C"
        if MonthColInsert == int("4"):
            MonthColInsert = "D"
        if MonthColInsert == int("5"):
            MonthColInsert = "E"
        if MonthColInsert == int("6"):
            MonthColInsert = "F"
        if MonthColInsert == int("7"):
            MonthColInsert = "G"
        if MonthColInsert == int("8"):
            MonthColInsert = "H"
        if MonthColInsert == int("9"):
            MonthColInsert = "I"
        if MonthColInsert == int("10"):
            MonthColInsert = "J"
        if MonthColInsert == int("11"):
            MonthColInsert = "K"
        if MonthColInsert == int("12"):
            MonthColInsert = "L"
        if MonthColInsert == int("13"):
            MonthColInsert = "M"

        for row in range(1, ws.max_row + 1):
            print(row)
            if quelle == ws.cell(row=row, column=1).value:
                print("Gefunden")
                global MonthRowInsert
                MonthRowInsert = row

        print(str(MonthColInsert) + " ist der Buchstabe des Monats " + month)

        ws[str(MonthColInsert) + str(MonthRowInsert)].value = 0

        wb.save("Gehalt" + year + ".xlsx")

        showText("Gehalt erfolgreich gelöscht.")
    else:
        print("Abbruch")
   



# DIAGRAMME

# Top-Left
def getAusgabenGehalt_diagram_top_left(year):
    wb = load_workbook("Finanzen" + year + ".xlsx")
    ws = wb["Finanzen"]

    wbGehalt = load_workbook("Gehalt" + year + ".xlsx")
    wsGehalt = wbGehalt["Gehalt"]

    global GehaltListe
    GehaltListe = []

    global gehaltAll
    gehaltAll = 0
    for col in range(2, wsGehalt.max_column + 1):
        if col == int("1"):
            col = "A"
        if col == int("2"):
            col = "B"
        if col == int("3"):
            col = "C"
        if col == int("4"):
            col = "D"
        if col == int("5"):
            col = "E"
        if col == int("6"):
            col = "F"
        if col == int("7"):
            col = "G"
        if col == int("8"):
            col = "H"
        if col == int("9"):
            col = "I"
        if col == int("10"):
            col = "J"
        if col == int("11"):
            col = "K"
        if col == int("12"):
            col = "L"
        if col == int("13"):
            col = "M"

        gehaltDieserMonat = 0
        for row in range(2, wsGehalt.max_row + 1):
            print(row)
            gehaltDieserMonat += round(float(wsGehalt[str(col) + str(row)].value), 2)
        print("Gehalt dieser Monat: " + str(gehaltDieserMonat))
        GehaltListe.append(gehaltDieserMonat)

        gehaltAll += float(gehaltDieserMonat)
        #gehaltAll = round(gehaltAll, 0)
        if str(gehaltDieserMonat) == "0.0":
            print("Gehalt noch nicht eingetragen für den Monat " + str(wsGehalt[str(col) + str(1)].value))
            #showText("Das Gehalt für den Monat " + str(wsGehalt[str(col) + str(
                #1)].value) + " wurde noch nicht eingetragen. Gehe dafür bitte zum Hauptfenster zurück und füge dein Gehalt links unten, unter <Gehalt> hinzu.")
        print("Gehalt all:" + str(gehaltAll))

    global AllMonthsValuesTogether
    AllMonthsValuesTogether = 0

    

    global AusgabenL
    AusgabenL = []

    for col in range(2, ws.max_column + 1):
        if col == int("1"):
            col = "A"
        if col == int("2"):
            col = "B"
        if col == int("3"):
            col = "C"
        if col == int("4"):
            col = "D"
        if col == int("5"):
            col = "E"
        if col == int("6"):
            col = "F"
        if col == int("7"):
            col = "G"
        if col == int("8"):
            col = "H"
        if col == int("9"):
            col = "I"
        if col == int("10"):
            col = "J"
        if col == int("11"):
            col = "K"
        if col == int("12"):
            col = "L"
        if col == int("13"):
            col = "M"
        print(col)
        MonthValue = 0

        for row in range(2, ws.max_row + 1):
            print(row)
            MonthValue += float(ws[str(col) + str(row)].value)


        print("MonthValue: " + str(MonthValue))
        AllMonthsValuesTogether += float(MonthValue)
        print(round(AllMonthsValuesTogether, 2), f"= round({AllMonthsValuesTogether}, 2)")
        AllMonthsValuesTogether = round(AllMonthsValuesTogether, 2)

        AusgabenL.append(MonthValue)

    global Relation
    Relation = float(gehaltAll) - float(AllMonthsValuesTogether)
    print(Relation)
    Relation = round(Relation, 2)

    print(str(GehaltListe))
    print(str(AusgabenL))
    global AusgabenListe
    AusgabenListe = AusgabenL

def plot_Months(year):
    # Set up the plot

    #For remote control
    global pltTL
    pltTL = plt

    pltTL.rcParams['font.family'] = 'sans-serif'
    pltTL.rcParams['font.sans-serif'] = ['Arial']
    fig, ax = pltTL.subplots(figsize=(16, 9), facecolor='#f0f0f0')
    ax.set_facecolor('#ffffff')
    fig.canvas.manager.set_window_title(f'Mein Finanzüberblick {year}')


    getAusgabenGehalt_diagram_top_left(year)

    
    # Data
    months = ["Januar", "Februar", "März", "April", "Mai", "Juni", "Juli", "August", "September", "Oktober", "November", "Dezember"]
    expenses = AusgabenListe
    salary = GehaltListe

    
    # Add annotations for highest and lowest points
    """
    max_expense = max(expenses)
    min_expense = min(expenses)

    # Annotation für das Maximum
    ax.annotate(f'Maximum: {max_expense}€', 
                xy=(months[expenses.index(max_expense)], max_expense),
                xytext=(50, 50), textcoords='offset points', 
                ha='left', va='bottom', fontsize=10,
                bbox=dict(boxstyle='round,pad=0.5', fc='yellow', alpha=0.5),
                arrowprops=dict(arrowstyle='->', connectionstyle='arc3,rad=0.3'))

    # Annotation für das Minimum
    ax.annotate(f'Minimum: {min_expense}€', 
                xy=(months[expenses.index(min_expense)], min_expense),
                xytext=(50, -50), textcoords='offset points', 
                ha='left', va='top', fontsize=10,
                bbox=dict(boxstyle='round,pad=0.5', fc='yellow', alpha=0.5),
                arrowprops=dict(arrowstyle='->', connectionstyle='arc3,rad=0.3'))


    max_salary = max(salary)
    min_salary = min(salary)
    # Annotation für das Maximum
    ax.annotate(f'Maximum: {max_salary}€', 
                xy=(months[salary.index(max_salary)], max_salary),
                xytext=(50, 50), textcoords='offset points', 
                ha='left', va='bottom', fontsize=10,
                bbox=dict(boxstyle='round,pad=0.5', fc='purple', alpha=0.5),
                arrowprops=dict(arrowstyle='->', connectionstyle='arc3,rad=0.3'))

    # Annotation für das Minimum
    ax.annotate(f'Minimum: {min_salary}€', 
                xy=(months[salary.index(min_salary)], min_salary),
                xytext=(50, -50), textcoords='offset points', 
                ha='left', va='top', fontsize=10,
                bbox=dict(boxstyle='round,pad=0.5', fc='purple', alpha=0.5),
                arrowprops=dict(arrowstyle='->', connectionstyle='arc3,rad=0.3'))
    """


    # Set the width of each bar
    bar_width = 0.35

    # Set the positions of the bars on the x-axis
    r1 = np.arange(len(months))
    r2 = [x + bar_width for x in r1]

    # Create the bar chart
    expense_bars = ax.bar(r1, expenses, color='#FF6B6B', width=bar_width, label='Ausgaben (€)', alpha=0.8)
    salary_bars = ax.bar(r2, salary, color='#4ECDC4', width=bar_width, label='Einnahmen (€)', alpha=0.8)

    # Max-/Min Einkommen/Ausgaben Legende
    """purple_frame = [0,0,0,0,0,0,0,0,0,0,0,0]
    purple_label = ax.bar(r1, purple_frame, color="#b555b5", width=bar_width, label="Max./Min. Einnahmen", alpha=0.8)
    yellow_frame = [0,0,0,0,0,0,0,0,0,0,0,0]
    yellow_label = ax.bar(r1, yellow_frame, color="#fcf52f", width=bar_width, label="Max./Min. Ausgaben", alpha=0.8)"""

    # Customize the plot
    ax.set_title(f"Monatliche Ausgaben und Einnahmen im Jahresverlauf {year}", fontsize=24, fontweight='bold', pad=20)
    ax.set_xlabel('Monate', fontsize=16, labelpad=15, fontweight="bold")
    ax.set_ylabel('Betrag (€)', fontsize=16, labelpad=15, fontweight="bold")

    # place a text box in upper left in axes coords
    props = dict(boxstyle='round', facecolor='white', alpha=0.7)

    ax.text(0.992, 0.98, "Komplette Einnahmen: " + str(gehaltAll) + '\n' + "Komplette Ausgaben: " + str(AllMonthsValuesTogether) + '\n' + '\n' + "Differenz: " + str(Relation), 
            transform=ax.transAxes, 
            fontsize=14, 
            verticalalignment='top', 
            horizontalalignment='right',
            bbox=props)
    
    #Trendlinie
    """# Calculate and plot trendline
    x_numeric = np.arange(len(months))
    z = np.polyfit(x_numeric, expenses, 1)
    p = np.poly1d(z)
    ax.plot(months, p(x_numeric), linestyle='--', color='red', linewidth=2, label='Trendlinie (Ausgaben)')

    x_numeric2 = np.arange(len(months))
    z2 = np.polyfit(x_numeric2, salary, 1)
    p2 = np.poly1d(z2)
    ax.plot(months, p2(x_numeric2), linestyle='--', color='blue', linewidth=2, label='Trendlinie (Gehalt)')"""

    # Set x-axis ticks
    ax.set_xticks([r + bar_width/2 for r in range(len(months))])
    ax.set_xticklabels(months, rotation=45, ha='right', fontsize=12)

    # Customize y-axis
    ax.yaxis.set_major_formatter(FuncFormatter(lambda x, p: f'{x:,.0f}'))
    ax.set_ylim(0, max(max(expenses), max(salary)) * 1.1)

    # Add grid
    ax.grid(True, linestyle='--', alpha=0.7, axis='y')



    # Add value labels on top of each bar
    def add_value_labels(bars):
        for bar in bars:
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2., height,
                    f'{height:,.0f}€',##########################################################Hier mit '2f' die nachkommastellen bearbeiten
                    ha='center', va='bottom', fontsize=10, fontweight='bold')

    add_value_labels(expense_bars)
    add_value_labels(salary_bars)

    # Add a subtle box around the plot area
    for spine in ax.spines.values():
        spine.set_edgecolor('#CCCCCC')

    # Add a background color gradient
    gradient = np.linspace(0, 1, 256).reshape(1, -1)
    gradient = np.vstack((gradient, gradient))
    ax.imshow(gradient, extent=[ax.get_xlim()[0], ax.get_xlim()[1], ax.get_ylim()[0], ax.get_ylim()[1]], 
            aspect='auto', alpha=0.1, cmap='coolwarm')

    # Add some visual enhancements
    for bar in expense_bars + salary_bars:
        bar.set_edgecolor('white')
        bar.set_linewidth(0.5)

    # Add a title box
    title_box = patches.Rectangle((0, 1.02), 1, 0.08, transform=ax.transAxes, fill=True,
                                facecolor='lightgray', edgecolor='none', alpha=0.8)
    ax.add_patch(title_box)


    # Customize legend
    ax.legend(fontsize=14, loc='upper left', frameon=True, facecolor='white', edgecolor='gray')

    # Add a watermark
    fig.text(0.95, 0, 'Jonas Gaiser', fontsize=12, color='gray', ha='right', va='bottom', alpha=0.5)


    # Adjust layout and display
    pltTL.tight_layout()
    pltTL.show()

# Top-Right
def GetAllKategories_diagram_top_right(year):
    wb = load_workbook("Finanzen" + year + ".xlsx")
    ws = wb["Finanzen"]

    wbGehalt = load_workbook("Gehalt" + year + ".xlsx")
    wsGehalt = wbGehalt["Gehalt"]

    global gehaltAll
    gehaltAll = 0
    for col in range(2, wsGehalt.max_column + 1):
        if col == int("1"):
            col = "A"
        if col == int("2"):
            col = "B"
        if col == int("3"):
            col = "C"
        if col == int("4"):
            col = "D"
        if col == int("5"):
            col = "E"
        if col == int("6"):
            col = "F"
        if col == int("7"):
            col = "G"
        if col == int("8"):
            col = "H"
        if col == int("9"):
            col = "I"
        if col == int("10"):
            col = "J"
        if col == int("11"):
            col = "K"
        if col == int("12"):
            col = "L"
        if col == int("13"):
            col = "M"

        gehaltDieserMonat = 0
        for row in range(2, wsGehalt.max_row + 1):
            print(row)
            gehaltDieserMonat += round(float(wsGehalt[str(col) + str(row)].value), 2)
        print("Gehalt dieser Monat: " + str(gehaltDieserMonat))

        gehaltAll += gehaltDieserMonat
        gehaltAll = round(gehaltAll, 2)
        if str(gehaltDieserMonat) == "0.0":
            print("Gehalt noch nicht eingetragen für den Monat " + str(wsGehalt[str(col) + str(1)].value))
            #showText("Das Gehalt für den Monat " + str(wsGehalt[str(col) + str(
                #1)].value) + " wurde noch nicht eingetragen. Gehe dafür bitte zum Hauptfenster zurück und füge dein Gehalt links unten, unter <Gehalt> hinzu.")
        print("Gehalt all:" + str(gehaltAll))

    global AllKategorienValuesTogether
    AllKategorienValuesTogether = 0

    global KategorienListe
    KategorienListe = []
    global BeträgeListe
    BeträgeListe = []
    for row in range(2, ws.max_row + 1):
        print(row)
        KategorieValue = 0
        KategorienListe.append(ws[str("A") + str(row)].value)
        for col in range(2, ws.max_column + 1):
            print(col)
            if col == int("1"):
                col = "A"
            if col == int("2"):
                col = "B"
            if col == int("3"):
                col = "C"
            if col == int("4"):
                col = "D"
            if col == int("5"):
                col = "E"
            if col == int("6"):
                col = "F"
            if col == int("7"):
                col = "G"
            if col == int("8"):
                col = "H"
            if col == int("9"):
                col = "I"
            if col == int("10"):
                col = "J"
            if col == int("11"):
                col = "K"
            if col == int("12"):
                col = "L"
            if col == int("13"):
                col = "M"
            KategorieValue += float(ws[str(col) + str(row)].value)
        BeträgeListe.append(str(KategorieValue))


        print("MonthValue: " + str(KategorieValue))
        AllKategorienValuesTogether += float(KategorieValue)
        print(round(AllKategorienValuesTogether, 2), f"= round({AllKategorienValuesTogether}, 2)")
        AllKategorienValuesTogether = round(AllKategorienValuesTogether, 2)

    global Relation
    Relation = float(gehaltAll) - float(AllKategorienValuesTogether)
    print(Relation)
    Relation = round(Relation, 2)

    print("All: " + str(AllKategorienValuesTogether))
    print(KategorienListe)
    print(BeträgeListe)
    print(gehaltAll)
    print(Relation)

def plot_kategorien(year):
    GetAllKategories_diagram_top_right(year)  # Ruft Ihre Funktion auf, um die Daten zu erhalten


    #For remote control
    global pltTR
    pltTR = plt


    # Konvertieren Sie die Beträge in Floats
    betraege = [float(betrag) for betrag in BeträgeListe]

    # Erstellen Sie das Diagramm
    fig, ax = pltTR.subplots(figsize=(16, 9), facecolor='#f0f0f0')
    ax.set_facecolor('#ffffff')
    fig.canvas.manager.set_window_title('Mein Finanzüberblick ' + str(year))

    # Balkenbreite und Positionen
    bar_width = 0.7
    r = np.arange(len(KategorienListe))

    # Erstellen Sie die Balken
    bars = ax.bar(r, betraege, color='#FF6B6B', width=bar_width, edgecolor='white', linewidth=0.7, alpha=0.8, label="Ausgaben")
    ax.legend(fontsize=14, loc='upper left', frameon=True, facecolor='white', edgecolor='gray')

    # Beschriftungen und Titel
    ax.set_xlabel('Kategorien', fontsize=16, fontweight='bold', labelpad=15)
    ax.set_ylabel('Beträge (€)', fontsize=16, fontweight='bold', labelpad=15)
    ax.set_title(f'Ausgaben aller Kategorien im Jahresverlauf {year}', fontsize=24, fontweight='bold', pad=20)

    # Rotieren Sie die x-Achsen-Beschriftungen für bessere Lesbarkeit
    pltTR.xticks(r, KategorienListe, rotation=45, ha='right', fontsize=12)

    # Formatieren Sie die y-Achse als Währung
    def currency_formatter(x, p):
        return f"{x:,.2f} €"
    ax.yaxis.set_major_formatter(FuncFormatter(currency_formatter))

    # Fügen Sie Werte über den Balken hinzu
    for bar in bars:
        height = bar.get_height()
        ax.text(bar.get_x() + bar.get_width()/2., height,
                f'{height:,.0f} €',##########################################################Hier mit '2f' die nachkommastellen bearbeiten
                ha='center', va='bottom', fontsize=10, fontweight='bold')

    # Fügen Sie eine Textbox mit Gesamtinformationen hinzu
    props = dict(boxstyle='round,pad=0.5', facecolor='white', alpha=0.7, edgecolor='gray')
    info_text = (f"Komplette Einnahmen: {gehaltAll:,.2f} €\n"
                 f"Komplette Ausgaben: {AllKategorienValuesTogether:,.2f} €\n"
                 f"Differenz: {Relation:,.2f} €")
    ax.text(0.992, 0.98, info_text, transform=ax.transAxes, 
            fontsize=14, verticalalignment='top', horizontalalignment='right',
            bbox=props)

    # Fügen Sie ein Gitternetz hinzu
    ax.grid(True, linestyle='--', alpha=0.7, axis='y')

    # Entfernen Sie die obere und rechte Achse
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)

    # Fügen Sie einen Rahmen um den Plotbereich hinzu
    for spine in ax.spines.values():
        spine.set_edgecolor('#CCCCCC')

    # Fügen Sie einen Farbverlauf als Hintergrund hinzu
    gradient = np.linspace(0, 1, 256).reshape(1, -1)
    gradient = np.vstack((gradient, gradient))
    ax.imshow(gradient, extent=[ax.get_xlim()[0], ax.get_xlim()[1], ax.get_ylim()[0], ax.get_ylim()[1]], 
              aspect='auto', alpha=0.1, cmap='coolwarm')

    # Fügen Sie ein Titelfeld hinzu
    title_box = patches.Rectangle((0, 1.02), 1, 0.08, transform=ax.transAxes, fill=True,
                                  facecolor='lightgray', edgecolor='none', alpha=0.8)
    ax.add_patch(title_box)

    # Add a watermark
    fig.text(0.95, 0, 'Jonas Gaiser', fontsize=12, color='gray', ha='right', va='bottom', alpha=0.5)

    # Passen Sie das Layout an
    pltTR.tight_layout()

    # Zeigen Sie das Diagramm an
    pltTR.show()

# Bottom-Left

def GetAllKategoriesMonth_diagram_bottom_left(year, MonthInp):

    wb = load_workbook("Finanzen" + str(year) + ".xlsx")  # Erstellt NEUES, bei nur hinzufügen/abändern nutze load_workbook
    ws = wb["Finanzen"]

    wbGehalt = load_workbook("Gehalt" + year + ".xlsx")
    wsGehalt = wbGehalt["Gehalt"]

    gehaltAll = 0
    global gehaltMonth
    gehaltMonth = 0

    MonthCol = "A"
    print("---------------" + MonthInp)
    if MonthInp == "Januar":
        MonthCol = "B"
    if MonthInp == "Februar":
        MonthCol = "C"
    if MonthInp == "März":
        MonthCol = "D"
    if MonthInp == "April":
        MonthCol = "E"
    if MonthInp == "Mai":
        MonthCol = "F"
    if MonthInp == "Juni":
        MonthCol = "G"
    if MonthInp == "Juli":
        MonthCol = "H"
    if MonthInp == "August":
        MonthCol = "I"
    if MonthInp == "September":
        MonthCol = "J"
    if MonthInp == "Oktober":
        MonthCol = "K"
    if MonthInp == "November":
        MonthCol = "L"
    if MonthInp == "Dezember":
        MonthCol = "M"

    print("---------------------" + MonthCol)
    
    global gehaltdieserMonat
    gehaltdieserMonat = 0
    for row in range(2, wsGehalt.max_row + 1):
        print(row)
        gehaltdieserMonat += round(float(wsGehalt[str(MonthCol) + str(row)].value), 2)
        gehaltAll += gehaltdieserMonat
    print("gehalt dieser Monat: " + str(gehaltdieserMonat))


    

    global AllValuesTogether
    AllValuesTogether = 0

    global Kategorien
    Kategorien = []
    global Beträge
    Beträge = []

    for row in range(2, ws.max_row + 1):
        print(ws["A" + str(row)].value)
        Kategorien.insert(int(Kategorien.__len__()) + 1, ws["A" + str(row)].value)

        if str(ws[MonthCol + str(row)].value) == "None":
            ws[MonthCol + str(row)].value = 0

        print(ws[MonthCol + str(row)].value)
        Beträge.insert(int(Kategorien.__len__()) + 1, ws[MonthCol + str(row)].value)
        AllValuesTogether += float(ws[MonthCol + str(row)].value)


    print("---------------")
    print("Zusammen:")
    print(AllValuesTogether)

    global KategorienListe
    KategorienListe = Kategorien
    print(KategorienListe)
    global BeträgeListe
    BeträgeListe = Beträge
    print(BeträgeListe)







    for row in range(2, wsGehalt.max_row + 1):
        print(row)
        gehaltdieserMonat += round(float(wsGehalt[MonthCol + str(row)].value), 2)
        gehaltAll += gehaltdieserMonat
        
    gehaltdieserMonat = gehaltdieserMonat / 2 # Fehlerbehebung?
    print("gehalt dieser Monat : " + str(gehaltdieserMonat))

    global Relation
    Relation = float(gehaltdieserMonat) - float(AllValuesTogether)
    print(Relation)
    Relation = round(Relation, 2)

def plot_kategorienMonth(year, month):
    GetAllKategoriesMonth_diagram_bottom_left(year, month)  # Ruft Ihre Funktion auf, um die Daten zu erhalten
    
    #For remote control
    global pltBL
    pltBL = plt

    # Konvertieren Sie die Beträge in Floats
    betraege = [float(betrag) for betrag in BeträgeListe]

    # Erstellen Sie das Diagramm
    fig, ax = pltBL.subplots(figsize=(16, 9), facecolor='#f0f0f0')
    ax.set_facecolor('#ffffff')
    fig.canvas.manager.set_window_title(f"Mein Finanzüberblick {month} {year}")

    # Balkenbreite und Positionen
    bar_width = 0.7
    r = np.arange(len(KategorienListe))

    # Erstellen Sie die Balken
    bars = ax.bar(r, betraege, color='#FF6B6B', width=bar_width, edgecolor='white', linewidth=0.7, alpha=0.8, label="Ausgaben")
    ax.legend(fontsize=14, loc='upper left', frameon=True, facecolor='white', edgecolor='gray')

    # Beschriftungen und Titel
    ax.set_xlabel('Kategorien', fontsize=16, fontweight='bold', labelpad=15)
    ax.set_ylabel('Beträge (€)', fontsize=16, fontweight='bold', labelpad=15)
    ax.set_title(f'Ausgaben aller Kategorien des Monats {month} im Jahr {year}', fontsize=24, fontweight='bold', pad=20)

    # Rotieren Sie die x-Achsen-Beschriftungen für bessere Lesbarkeit
    pltBL.xticks(r, KategorienListe, rotation=45, ha='right', fontsize=12)

    # Formatieren Sie die y-Achse als Währung
    def currency_formatter(x, p):
        return f"{x:,.0f} €" ##########################################################Hier mit '2f' die nachkommastellen bearbeiten
    ax.yaxis.set_major_formatter(FuncFormatter(currency_formatter))

    # Fügen Sie Werte über den Balken hinzu
    for bar in bars:
        height = bar.get_height()
        ax.text(bar.get_x() + bar.get_width()/2., height,
                f'{height:,.0f} €',##########################################################Hier mit '2f' die nachkommastellen bearbeiten
                ha='center', va='bottom', fontsize=10, fontweight='bold')

    # Fügen Sie eine Textbox mit Gesamtinformationen hinzu
    props = dict(boxstyle='round,pad=0.5', facecolor='white', alpha=0.7, edgecolor='gray')
    info_text = (f"Einnahmen im Monat: {gehaltdieserMonat:,.2f} €\n"
                 f"Ausgaben im Monat: {AllValuesTogether:,.2f} €\n"
                 f"Differenz: {Relation:,.2f} €")
    ax.text(0.992, 0.98, info_text, transform=ax.transAxes, 
            fontsize=14, verticalalignment='top', horizontalalignment='right',
            bbox=props)

    # Fügen Sie ein Gitternetz hinzu
    ax.grid(True, linestyle='--', alpha=0.7, axis='y')

    # Entfernen Sie die obere und rechte Achse
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)

    # Fügen Sie einen Rahmen um den Plotbereich hinzu
    for spine in ax.spines.values():
        spine.set_edgecolor('#CCCCCC')

    # Fügen Sie einen Farbverlauf als Hintergrund hinzu
    gradient = np.linspace(0, 1, 256).reshape(1, -1)
    gradient = np.vstack((gradient, gradient))
    ax.imshow(gradient, extent=[ax.get_xlim()[0], ax.get_xlim()[1], ax.get_ylim()[0], ax.get_ylim()[1]], 
              aspect='auto', alpha=0.1, cmap='coolwarm')

    # Fügen Sie ein Titelfeld hinzu
    title_box = patches.Rectangle((0, 1.02), 1, 0.08, transform=ax.transAxes, fill=True,
                                  facecolor='lightgray', edgecolor='none', alpha=0.8)
    ax.add_patch(title_box)

    # Add a watermark
    fig.text(0.95, 0, 'Jonas Gaiser', fontsize=12, color='gray', ha='right', va='bottom', alpha=0.5)

    # Passen Sie das Layout an
    pltBL.tight_layout()

    # Zeigen Sie das Diagramm an
    pltBL.show()

# Bottom-Right

def getValueOfKategorieAllMonths_diagram_bottom_right(year, Kategorie):
    KategorieInp = Kategorie

    wb = load_workbook("Finanzen" + str(year) + ".xlsx")  # Erstellt NEUES, bei nur hinzufügen/abändern nutze load_workbook
    ws = wb["Finanzen"]

    for row in range(2, ws.max_row + 1):
        print(row)
        if KategorieInp == ws.cell(row=row, column=1).value:
            print("Gefunden")
            global KategorieRow
            KategorieRow = row

    global AllValuesTogether
    AllValuesTogether = 0
    global Beträge
    Beträge = []
    Monate = []

    for col in range(2, ws.max_column + 1):
        if col == int("1"):
            col = "A"
        if col == int("2"):
            col = "B"
        if col == int("3"):
            col = "C"
        if col == int("4"):
            col = "D"
        if col == int("5"):
            col = "E"
        if col == int("6"):
            col = "F"
        if col == int("7"):
            col = "G"
        if col == int("8"):
            col = "H"
        if col == int("9"):
            col = "I"
        if col == int("10"):
            col = "J"
        if col == int("11"):
            col = "K"
        if col == int("12"):
            col = "L"
        if col == int("13"):
            col = "M"

        if str(ws[str(col) + str(KategorieRow)].value) == "None":
            ws[str(col) + str(KategorieRow)].value = 0

        print(str(ws[str(col) + "1"].value) + ":")
        Monate.insert(int(Monate.__len__()) + 1, ws[str(col) + "1"].value)
        print(ws[str(col) + str(KategorieRow)].value)
        AllValuesTogether += round(float(ws[col + str(KategorieRow)].value), 2)
        Beträge.insert(int(Monate.__len__()) + 1, ws[str(col) + str(KategorieRow)].value)


    print("---------------")
    print("Zusammen:")
    AllValuesTogether = round(AllValuesTogether, 2)
    print(AllValuesTogether)

    print(Monate)
    print(Beträge)

    global AusgabenListe
    AusgabenListe = Beträge

def plot_eineKategorie(year, Kategorie):

    #For remote control
    global pltBR
    pltBR = plt

    # Set up the plot

    pltBR.rcParams['font.family'] = 'sans-serif'
    pltBR.rcParams['font.sans-serif'] = ['Arial']
    fig, ax = pltBR.subplots(figsize=(16, 9), facecolor='#f0f0f0')
    ax.set_facecolor('#ffffff')
    fig.canvas.manager.set_window_title('Mein Finanzüberblick ' + str(year))


    getValueOfKategorieAllMonths_diagram_bottom_right(year, Kategorie)

    # Data
    months = ["Januar", "Februar", "März", "April", "Mai", "Juni", "Juli", "August", "September", "Oktober", "November", "Dezember"]
    expenses = AusgabenListe


    # place a text box in upper left in axes coords
    props = dict(boxstyle='round', facecolor='white', alpha=0.7)

    ax.text(0.992, 0.98, "Komplette Ausgaben in der Kategorie: " + str(AllValuesTogether), 
            transform=ax.transAxes, 
            fontsize=14, 
            verticalalignment='top', 
            horizontalalignment='right',
            bbox=props)

    
    # Add annotations for highest and lowest points
    '''
    max_expense = max(expenses)
    min_expense = min(expenses)

    # Annotation für das Maximum
    ax.annotate(f'Maximum: {max_expense}€', 
                xy=(months[expenses.index(max_expense)], max_expense),
                xytext=(50, 50), textcoords='offset points', 
                ha='left', va='bottom', fontsize=10,
                bbox=dict(boxstyle='round,pad=0.5', fc='yellow', alpha=0.5),
                arrowprops=dict(arrowstyle='->', connectionstyle='arc3,rad=0.3'))

    # Annotation für das Minimum
    ax.annotate(f'Minimum: {min_expense}€', 
                xy=(months[expenses.index(min_expense)], min_expense),
                xytext=(50, -50), textcoords='offset points', 
                ha='left', va='top', fontsize=10,
                bbox=dict(boxstyle='round,pad=0.5', fc='yellow', alpha=0.5),
                arrowprops=dict(arrowstyle='->', connectionstyle='arc3,rad=0.3'))
    '''




    # Set the width of each bar
    bar_width = 0.35

    # Set the positions of the bars on the x-axis
    r1 = np.arange(len(months))
    #r2 = [x + bar_width for x in r1]

    # Create the bar chart
    expense_bars = ax.bar(r1, expenses, color='#FF6B6B', width=bar_width, label='Ausgaben (€)', alpha=0.8)

    # Max-/min einnahmen/ausgaben Legende
    """
    yellow_frame = [0,0,0,0,0,0,0,0,0,0,0,0]
    yellow_label = ax.bar(r1, yellow_frame, color="#fcf52f", width=bar_width, label="Max./Min. Ausgaben", alpha=0.8)"""

    # Customize the plot
    ax.set_title("Ausgaben der Kategorie <" + str(Kategorie) + "> im Jahresverlauf " + str(year), fontsize=24, fontweight='bold', pad=20)
    ax.set_xlabel('Monate', fontsize=16, fontweight="bold", labelpad=15)
    ax.set_ylabel('Betrag (€)', fontsize=16, labelpad=15, fontweight="bold")

    
    # Calculate and plot trendline

    x_numeric = np.arange(len(months))
    z = np.polyfit(x_numeric, expenses, 1)
    p = np.poly1d(z)
    ax.plot(months, p(x_numeric), linestyle='', color='red', linewidth=2) # label weg und linestyle auf none da sond ein fehler mit 'april' auftritt wenn ich den gesamten bereich auskommentiere


    # Set x-axis ticks
    #ax.set_xticks([r + bar_width/2 for r in range(len(months))])
    ax.set_xticklabels(months, rotation=45, ha='right', fontsize=12)


    # Add grid
    ax.grid(True, linestyle='--', alpha=0.7, axis='y')



    # Add value labels on top of each bar
    def add_value_labels(bars):
        for bar in bars:
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2., height,
                    f'{height:,.0f}€',##########################################################Hier mit '2f' die nachkommastellen bearbeiten
                    ha='center', va='bottom', fontsize=10, fontweight='bold')

    add_value_labels(expense_bars)

    # Add a subtle box around the plot area
    for spine in ax.spines.values():
        spine.set_edgecolor('#CCCCCC')


    # Add a background color gradient
    gradient = np.linspace(0, 1, 256).reshape(1, -1)
    gradient = np.vstack((gradient, gradient))
    ax.imshow(gradient, extent=[ax.get_xlim()[0], ax.get_xlim()[1], ax.get_ylim()[0], ax.get_ylim()[1]], 
            aspect='auto', alpha=0.1, cmap='coolwarm')


    # Add a title box
    title_box = patches.Rectangle((0, 1.02), 1, 0.08, transform=ax.transAxes, fill=True,
                                facecolor='lightgray', edgecolor='none', alpha=0.8)
    ax.add_patch(title_box)


    # Customize legend
    ax.legend(fontsize=14, loc='upper left', frameon=True, facecolor='white', edgecolor='gray')

    # Add a watermark
    fig.text(0.95, 0, 'Jonas Gaiser', fontsize=12, color='gray', ha='right', va='bottom', alpha=0.5)
    
    # Adjust layout and display
    pltBR.tight_layout()
    pltBR.show()





## ALL YEARS

def getValueAll_AllYears():
    global AllValuesTogether
    AllValuesTogether = 0
    global Jahre
    Jahre = []
    global BeträgeProJahr
    BeträgeProJahr = []
    for length in range(0, YearsList.__len__()):
        print(YearsList.__len__())
        YearsListItem = str(YearsList.__getitem__(length))
        print(YearsList.__getitem__(length))

        wb = load_workbook("Finanzen" + YearsListItem + ".xlsx")  # Erstellt NEUES, bei nur hinzufügen/abändern nutze load_workbook
        ws = wb["Finanzen"]
        JahrCursor = 0
        for col in range(2, ws.max_column + 1):
            print("-----Col: " + str(col))
            if col == int("1"):
                col = "A"
            if col == int("2"):
                col = "B"
            if col == int("3"):
                col = "C"
            if col == int("4"):
                col = "D"
            if col == int("5"):
                col = "E"
            if col == int("6"):
                col = "F"
            if col == int("7"):
                col = "G"
            if col == int("8"):
                col = "H"
            if col == int("9"):
                col = "I"
            if col == int("10"):
                col = "J"
            if col == int("11"):
                col = "K"
            if col == int("12"):
                col = "L"
            if col == int("13"):
                col = "M"

            for row in range(2, ws.max_row + 1):
                print(row)
                JahrCursor += float(ws[str(col) + str(row)].value)
                AllValuesTogether += float(ws[str(col) + str(row)].value)
        
        BeträgeProJahr.append(JahrCursor)
        Jahre.append(YearsListItem)


    global GehaltListe
    GehaltListe = []
    global gehaltAll
    gehaltAll = 0

    for length in range(0, YearsList.__len__()):
        print(YearsList.__len__())
        YearsListItem = str(YearsList.__getitem__(length))
        print(YearsList.__getitem__(length))

        wbGehalt = load_workbook("Gehalt" + YearsListItem + ".xlsx")
        wsGehalt = wbGehalt["Gehalt"]


        JahrCursorGehalt = 0
        for col in range(2, wsGehalt.max_column + 1):
            if col == int("1"):
                col = "A"
            if col == int("2"):
                col = "B"
            if col == int("3"):
                col = "C"
            if col == int("4"):
                col = "D"
            if col == int("5"):
                col = "E"
            if col == int("6"):
                col = "F"
            if col == int("7"):
                col = "G"
            if col == int("8"):
                col = "H"
            if col == int("9"):
                col = "I"
            if col == int("10"):
                col = "J"
            if col == int("11"):
                col = "K"
            if col == int("12"):
                col = "L"
            if col == int("13"):
                col = "M"

            
            for row in range(2, wsGehalt.max_row + 1):
                print(row)
                JahrCursorGehalt += round(float(wsGehalt[str(col) + str(row)].value), 2)
                gehaltAll += round(float(wsGehalt[str(col) + str(row)].value), 2)

        GehaltListe.append(JahrCursorGehalt)

    print("---------------")
    print("Zusammen: (Ausgaben)")
    AllValuesTogether = round(AllValuesTogether, 2)
    print(AllValuesTogether)
    print("Zusammen: (Einnahmen)")
    gehaltAll = round(gehaltAll, 2)
    print(gehaltAll)
    print(Jahre)
    print(BeträgeProJahr)
    print(GehaltListe)
    

def getValueAll_AllYears_List():
    global AllValuesTogether
    AllValuesTogether = 0
    global Jahre
    Jahre = []
    global BeträgeProJahr
    BeträgeProJahr = []
    for length in range(0, YearsList.__len__()):
        print(YearsList.__len__())
        YearsListItem = str(YearsList.__getitem__(length))
        print(YearsList.__getitem__(length))

        wb = load_workbook("Finanzen" + YearsListItem + ".xlsx")  # Erstellt NEUES, bei nur hinzufügen/abändern nutze load_workbook
        ws = wb["Finanzen"]
        JahrCursor = 0
        for col in range(2, ws.max_column + 1):
            print("-----Col: " + str(col))
            if col == int("1"):
                col = "A"
            if col == int("2"):
                col = "B"
            if col == int("3"):
                col = "C"
            if col == int("4"):
                col = "D"
            if col == int("5"):
                col = "E"
            if col == int("6"):
                col = "F"
            if col == int("7"):
                col = "G"
            if col == int("8"):
                col = "H"
            if col == int("9"):
                col = "I"
            if col == int("10"):
                col = "J"
            if col == int("11"):
                col = "K"
            if col == int("12"):
                col = "L"
            if col == int("13"):
                col = "M"

            for row in range(2, ws.max_row + 1):
                print(row)
                JahrCursor += float(ws[str(col) + str(row)].value)
                AllValuesTogether += float(ws[str(col) + str(row)].value)
        
        BeträgeProJahr.append(JahrCursor)
        Jahre.append(YearsListItem)


    global GehaltListe
    GehaltListe = []
    global gehaltAll
    gehaltAll = 0

    for length in range(0, YearsList.__len__()):
        print(YearsList.__len__())
        YearsListItem = str(YearsList.__getitem__(length))
        print(YearsList.__getitem__(length))

        wbGehalt = load_workbook("Gehalt" + YearsListItem + ".xlsx")
        wsGehalt = wbGehalt["Gehalt"]


        JahrCursorGehalt = 0
        for col in range(2, wsGehalt.max_column + 1):
            if col == int("1"):
                col = "A"
            if col == int("2"):
                col = "B"
            if col == int("3"):
                col = "C"
            if col == int("4"):
                col = "D"
            if col == int("5"):
                col = "E"
            if col == int("6"):
                col = "F"
            if col == int("7"):
                col = "G"
            if col == int("8"):
                col = "H"
            if col == int("9"):
                col = "I"
            if col == int("10"):
                col = "J"
            if col == int("11"):
                col = "K"
            if col == int("12"):
                col = "L"
            if col == int("13"):
                col = "M"

            
            for row in range(2, wsGehalt.max_row + 1):
                print(row)
                JahrCursorGehalt += round(float(wsGehalt[str(col) + str(row)].value), 2)
                gehaltAll += round(float(wsGehalt[str(col) + str(row)].value), 2)

        GehaltListe.append(JahrCursorGehalt)

    print("---------------")
    print("Zusammen: (Ausgaben)")
    AllValuesTogether = round(AllValuesTogether, 2)
    print(AllValuesTogether)
    print("Zusammen: (Einnahmen)")
    gehaltAll = round(gehaltAll, 2)
    print(gehaltAll)
    print(Jahre)
    print(BeträgeProJahr)
    print(GehaltListe)

    #WINDOW

    getValueAll_AllYears_ListWind = tk.Tk()
    getValueAll_AllYears_ListWind.title("Alle Jahre ausgelesen: ")


    title = tk.Label(getValueAll_AllYears_ListWind, text="Alle Jahre", font=("Arial", 20))
    title.pack()

    frame = tk.Frame(getValueAll_AllYears_ListWind)
    frame.pack(side="top")
    JahreLabel = tk.Entry(frame, font=("Arial", 15))
    JahreLabel.pack(side="left")
    JahreLabel.insert(0, "Jahre:")
    AusgabenLabel = tk.Entry(frame, font=("Arial", 15))
    AusgabenLabel.pack(side="right")
    AusgabenLabel.insert(0, "Ausgaben:")
    EinnahmenLabel = tk.Entry(frame, font=("Arial", 15))
    EinnahmenLabel.pack(side="right")
    EinnahmenLabel.insert(0, "Einnahmen:")
    


    for length in range(0, YearsList.__len__()):
        print(YearsList.__len__())
        YearsListItem = str(YearsList.__getitem__(length))
        print(YearsList.__getitem__(length))

        frame = tk.Frame(getValueAll_AllYears_ListWind)
        frame.pack()
        entry1 = tk.Entry(frame, font=("Arial", 15))
        entry1.pack(side=tk.LEFT)
        entry1.insert(0, str(YearsListItem) + ": ")

        entry2 = tk.Entry(frame, font=("Arial", 15))
        entry2.pack(side=tk.RIGHT)
        entry2.insert(0,str(float(BeträgeProJahr.__getitem__(length))))

        entry3 = tk.Entry(frame, font=("Arial", 15))
        entry3.pack(side=tk.RIGHT)
        entry3.insert(0,str(float(GehaltListe.__getitem__(length))))

    entryAll = tk.Entry(getValueAll_AllYears_ListWind, font=("Arial", 20), width=30)
    entryAll.pack(pady=10, padx=10)
    entryAll.insert(0, float(AllValuesTogether))
    entryAll.insert(0, "Summe Ausgaben: ")



    entryGehalt = tk.Entry(getValueAll_AllYears_ListWind, font=("Arial", 20), width=30)
    entryGehalt.pack(pady=10, padx=10)
    entryGehalt.insert(0, str(float(gehaltAll)))
    entryGehalt.insert(0, "Summe Einnahmen: ")

    Relation = float(gehaltAll) - float(AllValuesTogether)
    print(Relation)
    Relation = round(Relation, 2)

    if Relation < 0:
        print("Verlust")
        entryRelation = tk.Entry(getValueAll_AllYears_ListWind, font=("Arial", 20), width=30, fg="red")
        entryRelation.pack(pady=10, padx=10)
        entryRelation.insert(0, str(Relation))
        entryRelation.insert(0, "Verlust von ")
    else:  # Auch wenn Relation == 0
        print("Gewinn")
        entryRelation = tk.Entry(getValueAll_AllYears_ListWind, font=("Arial", 20), width=30, fg="green")
        entryRelation.pack(pady=10, padx=10)
        entryRelation.insert(0, str(Relation))
        entryRelation.insert(0, "Gewinn von ")



    getValueAll_AllYears_ListWind.mainloop()


def getValueAll_AllYears_Diagram():
    global AllValuesTogether
    AllValuesTogether = 0
    global Jahre
    Jahre = []
    global BeträgeProJahr
    BeträgeProJahr = []
    for length in range(0, YearsList.__len__()):
        print(YearsList.__len__())
        YearsListItem = str(YearsList.__getitem__(length))
        print(YearsList.__getitem__(length))

        wb = load_workbook("Finanzen" + YearsListItem + ".xlsx")  # Erstellt NEUES, bei nur hinzufügen/abändern nutze load_workbook
        ws = wb["Finanzen"]
        JahrCursor = 0
        for col in range(2, ws.max_column + 1):
            print("-----Col: " + str(col))
            if col == int("1"):
                col = "A"
            if col == int("2"):
                col = "B"
            if col == int("3"):
                col = "C"
            if col == int("4"):
                col = "D"
            if col == int("5"):
                col = "E"
            if col == int("6"):
                col = "F"
            if col == int("7"):
                col = "G"
            if col == int("8"):
                col = "H"
            if col == int("9"):
                col = "I"
            if col == int("10"):
                col = "J"
            if col == int("11"):
                col = "K"
            if col == int("12"):
                col = "L"
            if col == int("13"):
                col = "M"

            for row in range(2, ws.max_row + 1):
                print(row)
                JahrCursor += float(ws[str(col) + str(row)].value)
                AllValuesTogether += float(ws[str(col) + str(row)].value)
        
        BeträgeProJahr.append(JahrCursor)
        Jahre.append(YearsListItem)


    global GehaltListe
    GehaltListe = []
    global gehaltAll
    gehaltAll = 0

    for length in range(0, YearsList.__len__()):
        print(YearsList.__len__())
        YearsListItem = str(YearsList.__getitem__(length))
        print(YearsList.__getitem__(length))

        wbGehalt = load_workbook("Gehalt" + YearsListItem + ".xlsx")
        wsGehalt = wbGehalt["Gehalt"]


        JahrCursorGehalt = 0
        for col in range(2, wsGehalt.max_column + 1):
            if col == int("1"):
                col = "A"
            if col == int("2"):
                col = "B"
            if col == int("3"):
                col = "C"
            if col == int("4"):
                col = "D"
            if col == int("5"):
                col = "E"
            if col == int("6"):
                col = "F"
            if col == int("7"):
                col = "G"
            if col == int("8"):
                col = "H"
            if col == int("9"):
                col = "I"
            if col == int("10"):
                col = "J"
            if col == int("11"):
                col = "K"
            if col == int("12"):
                col = "L"
            if col == int("13"):
                col = "M"

            
            for row in range(2, wsGehalt.max_row + 1):
                print(row)
                JahrCursorGehalt += round(float(wsGehalt[str(col) + str(row)].value), 2)
                gehaltAll += round(float(wsGehalt[str(col) + str(row)].value), 2)

        GehaltListe.append(JahrCursorGehalt)

    print("---------------")
    print("Zusammen: (Ausgaben)")
    AllValuesTogether = round(AllValuesTogether, 2)
    print(AllValuesTogether)
    print("Zusammen: (Einnahmen)")
    gehaltAll = round(gehaltAll, 2)
    print(gehaltAll)
    print(Jahre)
    print(BeträgeProJahr)
    print(GehaltListe)
    Relation = float(gehaltAll) - float(AllValuesTogether)
    print(Relation)
    Relation = round(Relation, 2)


    global pltTL
    pltTL = plt

    pltTL.rcParams['font.family'] = 'sans-serif'
    pltTL.rcParams['font.sans-serif'] = ['Arial']
    fig, ax = pltTL.subplots(figsize=(16, 9), facecolor='#f0f0f0')
    ax.set_facecolor('#ffffff')
    fig.canvas.manager.set_window_title(f'Mein Finanzüberblick {Jahre}')

    
    # Data
    years = Jahre
    expenses = BeträgeProJahr
    salary = GehaltListe

    
    # Add annotations for highest and lowest points
    """
    max_expense = max(expenses)
    min_expense = min(expenses)

    # Annotation für das Maximum
    ax.annotate(f'Maximum: {max_expense}€', 
                xy=(months[expenses.index(max_expense)], max_expense),
                xytext=(50, 50), textcoords='offset points', 
                ha='left', va='bottom', fontsize=10,
                bbox=dict(boxstyle='round,pad=0.5', fc='yellow', alpha=0.5),
                arrowprops=dict(arrowstyle='->', connectionstyle='arc3,rad=0.3'))

    # Annotation für das Minimum
    ax.annotate(f'Minimum: {min_expense}€', 
                xy=(months[expenses.index(min_expense)], min_expense),
                xytext=(50, -50), textcoords='offset points', 
                ha='left', va='top', fontsize=10,
                bbox=dict(boxstyle='round,pad=0.5', fc='yellow', alpha=0.5),
                arrowprops=dict(arrowstyle='->', connectionstyle='arc3,rad=0.3'))


    max_salary = max(salary)
    min_salary = min(salary)
    # Annotation für das Maximum
    ax.annotate(f'Maximum: {max_salary}€', 
                xy=(months[salary.index(max_salary)], max_salary),
                xytext=(50, 50), textcoords='offset points', 
                ha='left', va='bottom', fontsize=10,
                bbox=dict(boxstyle='round,pad=0.5', fc='purple', alpha=0.5),
                arrowprops=dict(arrowstyle='->', connectionstyle='arc3,rad=0.3'))

    # Annotation für das Minimum
    ax.annotate(f'Minimum: {min_salary}€', 
                xy=(months[salary.index(min_salary)], min_salary),
                xytext=(50, -50), textcoords='offset points', 
                ha='left', va='top', fontsize=10,
                bbox=dict(boxstyle='round,pad=0.5', fc='purple', alpha=0.5),
                arrowprops=dict(arrowstyle='->', connectionstyle='arc3,rad=0.3'))
    """


    # Set the width of each bar
    bar_width = 0.35

    # Set the positions of the bars on the x-axis
    r1 = np.arange(len(years))
    r2 = [x + bar_width for x in r1]

    # Create the bar chart
    expense_bars = ax.bar(r1, expenses, color='#FF6B6B', width=bar_width, label='Ausgaben (€)', alpha=0.8)
    salary_bars = ax.bar(r2, salary, color='#4ECDC4', width=bar_width, label='Einnahmen (€)', alpha=0.8)

    # Max-/Min Einkommen/Ausgaben Legende
    """purple_frame = [0,0,0,0,0,0,0,0,0,0,0,0]
    purple_label = ax.bar(r1, purple_frame, color="#b555b5", width=bar_width, label="Max./Min. Einnahmen", alpha=0.8)
    yellow_frame = [0,0,0,0,0,0,0,0,0,0,0,0]
    yellow_label = ax.bar(r1, yellow_frame, color="#fcf52f", width=bar_width, label="Max./Min. Ausgaben", alpha=0.8)"""

    # Customize the plot
    ax.set_title(f"Monatliche Ausgaben und Einnahmen im Jahresverlauf {years}", fontsize=24, fontweight='bold', pad=20)
    ax.set_xlabel('Monate', fontsize=16, labelpad=15, fontweight="bold")
    ax.set_ylabel('Betrag (€)', fontsize=16, labelpad=15, fontweight="bold")

    # place a text box in upper left in axes coords
    props = dict(boxstyle='round', facecolor='white', alpha=0.7)

    ax.text(0.992, 0.98, "Komplette Einnahmen: " + str(gehaltAll) + '\n' + "Komplette Ausgaben: " + str(AllValuesTogether) + '\n' + '\n' + "Differenz: " + str(Relation), 
            transform=ax.transAxes, 
            fontsize=14, 
            verticalalignment='top', 
            horizontalalignment='right',
            bbox=props)
    
    #Trendlinie
    """# Calculate and plot trendline
    x_numeric = np.arange(len(months))
    z = np.polyfit(x_numeric, expenses, 1)
    p = np.poly1d(z)
    ax.plot(months, p(x_numeric), linestyle='--', color='red', linewidth=2, label='Trendlinie (Ausgaben)')

    x_numeric2 = np.arange(len(months))
    z2 = np.polyfit(x_numeric2, salary, 1)
    p2 = np.poly1d(z2)
    ax.plot(months, p2(x_numeric2), linestyle='--', color='blue', linewidth=2, label='Trendlinie (Gehalt)')"""

    # Set x-axis ticks
    ax.set_xticks([r + bar_width/2 for r in range(len(years))])
    ax.set_xticklabels(years, rotation=45, ha='right', fontsize=12)

    # Customize y-axis
    ax.yaxis.set_major_formatter(FuncFormatter(lambda x, p: f'{x:,.0f}'))
    ax.set_ylim(0, max(max(expenses), max(salary)) * 1.1)

    # Add grid
    ax.grid(True, linestyle='--', alpha=0.7, axis='y')



    # Add value labels on top of each bar
    def add_value_labels(bars):
        for bar in bars:
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2., height,
                    f'{height:,.0f}€',##########################################################Hier mit '2f' die nachkommastellen bearbeiten
                    ha='center', va='bottom', fontsize=10, fontweight='bold')

    add_value_labels(expense_bars)
    add_value_labels(salary_bars)

    # Add a subtle box around the plot area
    for spine in ax.spines.values():
        spine.set_edgecolor('#CCCCCC')

    # Add a background color gradient
    gradient = np.linspace(0, 1, 256).reshape(1, -1)
    gradient = np.vstack((gradient, gradient))
    ax.imshow(gradient, extent=[ax.get_xlim()[0], ax.get_xlim()[1], ax.get_ylim()[0], ax.get_ylim()[1]], 
            aspect='auto', alpha=0.1, cmap='coolwarm')

    # Add some visual enhancements
    for bar in expense_bars + salary_bars:
        bar.set_edgecolor('white')
        bar.set_linewidth(0.5)

    # Add a title box
    title_box = patches.Rectangle((0, 1.02), 1, 0.08, transform=ax.transAxes, fill=True,
                                facecolor='lightgray', edgecolor='none', alpha=0.8)
    ax.add_patch(title_box)


    # Customize legend
    ax.legend(fontsize=14, loc='upper left', frameon=True, facecolor='white', edgecolor='gray')

    # Add a watermark
    fig.text(0.95, 0, 'Jonas Gaiser', fontsize=12, color='gray', ha='right', va='bottom', alpha=0.5)


    # Adjust layout and display
    pltTL.tight_layout()
    pltTL.show()

#-Kateggorien
def getAllYears_Kategorien():

    global Jahre
    Jahre = []
    global KategorienListe_allYears
    KategorienListe_allYears = []
    global KategorienValueListe
    KategorienValueListe = []

    YearsListItem = str(YearsList.__getitem__(1))
    wb = load_workbook("Finanzen" + YearsListItem + ".xlsx")
    ws = wb["Finanzen"]
    for row in range(2, ws.max_row + 1):
            KategorienListe_allYears.append(str(ws[str("A") + str(row)].value))

    #get gehaltAll
    gehaltAll = 0

    testListeKategorien = []
    for length in range(0, YearsList.__len__()):
        print(YearsList.__len__())
        YearsListItem = str(YearsList.__getitem__(length))
        print(YearsList.__getitem__(length))

        wb = load_workbook("Finanzen" + YearsListItem + ".xlsx")
        ws = wb["Finanzen"]

        wbGehalt = load_workbook("Gehalt" + YearsListItem + ".xlsx")
        wsGehalt = wbGehalt["Gehalt"]
        Jahre.append(YearsListItem)
        
        for col in range(2, wsGehalt.max_column + 1):
            if col == int("1"):
                col = "A"
            if col == int("2"):
                col = "B"
            if col == int("3"):
                col = "C"
            if col == int("4"):
                col = "D"
            if col == int("5"):
                col = "E"
            if col == int("6"):
                col = "F"
            if col == int("7"):
                col = "G"
            if col == int("8"):
                col = "H"
            if col == int("9"):
                col = "I"
            if col == int("10"):
                col = "J"
            if col == int("11"):
                col = "K"
            if col == int("12"):
                col = "L"
            if col == int("13"):
                col = "M"

            gehaltDieserMonat = 0
            for row in range(2, wsGehalt.max_row + 1):
                print(row)
                gehaltDieserMonat += round(float(wsGehalt[str(col) + str(row)].value), 2)
            print("Gehalt dieser Monat: " + str(gehaltDieserMonat))

            gehaltAll += gehaltDieserMonat
            gehaltAll = round(gehaltAll, 2)
            if str(gehaltDieserMonat) == "0.0":
                print("Gehalt noch nicht eingetragen für den Monat " + str(wsGehalt[str(col) + str(1)].value))
                #showText("Das Gehalt für den Monat " + str(wsGehalt[str(col) + str(
                    #1)].value) + " wurde noch nicht eingetragen. Gehe dafür bitte zum Hauptfenster zurück und füge dein Gehalt links unten, unter <Gehalt> hinzu.")
            print("Gehalt all:" + str(gehaltAll))

        global AllKategorienValuesTogether
        AllKategorienValuesTogether = 0

        
        KategorieimJahr = 0
        for row in range(2, ws.max_row + 1):
            print(row)
            KategorieValue = 0

            
            for col in range(2, ws.max_column + 1):
                print(col)
                if col == int("1"):
                    col = "A"
                if col == int("2"):
                    col = "B"
                if col == int("3"):
                    col = "C"
                if col == int("4"):
                    col = "D"
                if col == int("5"):
                    col = "E"
                if col == int("6"):
                    col = "F"
                if col == int("7"):
                    col = "G"
                if col == int("8"):
                    col = "H"
                if col == int("9"):
                    col = "I"
                if col == int("10"):
                    col = "J"
                if col == int("11"):
                    col = "K"
                if col == int("12"):
                    col = "L"
                if col == int("13"):
                    col = "M"
                KategorieValue += float(ws[str(col) + str(row)].value)
        
            testListeKategorien.append(KategorieValue)
    

    

    print(testListeKategorien)
    li1 = testListeKategorien #[1, 2, 3, 4, 5, 6, 7, 8]
    midpoint = len(li1) // 2
    first_half = li1[:midpoint]
    second_half = li1[midpoint:]
    li2 = []
    for i in range(len(first_half)):
        li2.append(round(first_half[i] + second_half[i], 2))


    KategorienValueListe = li2
        
    print(Jahre)
    print(KategorienListe_allYears)
    print(KategorienValueListe)
    print(gehaltAll)
    

def getAllYears_Kategorien_List():

    global Jahre
    Jahre = []
    global KategorienListe_allYears
    KategorienListe_allYears = []
    global KategorienValueListe
    KategorienValueListe = []

    YearsListItem = str(YearsList.__getitem__(0))
    wb = load_workbook("Finanzen" + YearsListItem + ".xlsx")
    ws = wb["Finanzen"]
    for row in range(2, ws.max_row + 1):
            KategorienListe_allYears.append(str(ws[str("A") + str(row)].value))

    #get gehaltAll
    gehaltAll = 0
    global AllKategorienValuesTogether
    AllKategorienValuesTogether = 0
    testListeKategorien = []
    for length in range(0, YearsList.__len__()):
        print(YearsList.__len__())
        YearsListItem = str(YearsList.__getitem__(length))
        print(YearsList.__getitem__(length))

        wb = load_workbook("Finanzen" + YearsListItem + ".xlsx")
        ws = wb["Finanzen"]

        wbGehalt = load_workbook("Gehalt" + YearsListItem + ".xlsx")
        wsGehalt = wbGehalt["Gehalt"]
        Jahre.append(YearsListItem)
        
        for col in range(2, wsGehalt.max_column + 1):
            if col == int("1"):
                col = "A"
            if col == int("2"):
                col = "B"
            if col == int("3"):
                col = "C"
            if col == int("4"):
                col = "D"
            if col == int("5"):
                col = "E"
            if col == int("6"):
                col = "F"
            if col == int("7"):
                col = "G"
            if col == int("8"):
                col = "H"
            if col == int("9"):
                col = "I"
            if col == int("10"):
                col = "J"
            if col == int("11"):
                col = "K"
            if col == int("12"):
                col = "L"
            if col == int("13"):
                col = "M"

            gehaltDieserMonat = 0
            for row in range(2, wsGehalt.max_row + 1):
                print(row)
                gehaltDieserMonat += round(float(wsGehalt[str(col) + str(row)].value), 2)
            print("Gehalt dieser Monat: " + str(gehaltDieserMonat))

            gehaltAll += gehaltDieserMonat
            gehaltAll = round(gehaltAll, 2)
            if str(gehaltDieserMonat) == "0.0":
                print("Gehalt noch nicht eingetragen für den Monat " + str(wsGehalt[str(col) + str(1)].value))
                #showText("Das Gehalt für den Monat " + str(wsGehalt[str(col) + str(
                    #1)].value) + " wurde noch nicht eingetragen. Gehe dafür bitte zum Hauptfenster zurück und füge dein Gehalt links unten, unter <Gehalt> hinzu.")
            print("Gehalt all:" + str(gehaltAll))


        
        
        for row in range(2, ws.max_row + 1):
            print(row)
            KategorieValue = 0

            
            for col in range(2, ws.max_column + 1):
                print(col)
                if col == int("1"):
                    col = "A"
                if col == int("2"):
                    col = "B"
                if col == int("3"):
                    col = "C"
                if col == int("4"):
                    col = "D"
                if col == int("5"):
                    col = "E"
                if col == int("6"):
                    col = "F"
                if col == int("7"):
                    col = "G"
                if col == int("8"):
                    col = "H"
                if col == int("9"):
                    col = "I"
                if col == int("10"):
                    col = "J"
                if col == int("11"):
                    col = "K"
                if col == int("12"):
                    col = "L"
                if col == int("13"):
                    col = "M"
                KategorieValue += float(ws[str(col) + str(row)].value)
                AllKategorienValuesTogether += float(ws[str(col) + str(row)].value)
        
            testListeKategorien.append(KategorieValue)
    

    

    print(testListeKategorien)
    li1 = testListeKategorien #[1, 2, 3, 4, 5, 6, 7, 8]
    midpoint = len(li1) // 2
    first_half = li1[:midpoint]
    second_half = li1[midpoint:]
    li2 = []
    for i in range(len(first_half)):
        li2.append(round(first_half[i] + second_half[i], 2))


    KategorienValueListe = li2
        
    print(Jahre)
    print(KategorienListe_allYears)
    print(KategorienValueListe)
    print(gehaltAll)
    print(AllKategorienValuesTogether)


    #WINDOW

    getValueAllKategorien_AllYears_ListWind = tk.Tk()
    getValueAllKategorien_AllYears_ListWind.title("Alle Jahre nach Kategorie ausgelesen: ")


    title = tk.Label(getValueAllKategorien_AllYears_ListWind, text="Alle Kategorien in allen Jahren", font=("Arial", 20))
    title.pack()

    frame = tk.Frame(getValueAllKategorien_AllYears_ListWind)
    frame.pack(side="top")
    KategorieLabel = tk.Entry(frame, font=("Arial", 15))
    KategorieLabel.pack(side="left")
    KategorieLabel.insert(0, "Kategorie:")
    AusgabenLabel = tk.Entry(frame, font=("Arial", 15))
    AusgabenLabel.pack(side="right")
    AusgabenLabel.insert(0, "Ausgaben:")

    


    for length in range(0, KategorienListe_allYears.__len__()):
        print(KategorienListe_allYears.__len__())
        KategorieListItem = str(KategorienListe_allYears.__getitem__(length))
        print(KategorienListe_allYears.__getitem__(length))

        frame = tk.Frame(getValueAllKategorien_AllYears_ListWind)
        frame.pack()
        entry1 = tk.Entry(frame, font=("Arial", 15))
        entry1.pack(side=tk.LEFT)
        entry1.insert(0, str(KategorieListItem) + ": ")

        entry2 = tk.Entry(frame, font=("Arial", 15))
        entry2.pack(side=tk.RIGHT)
        entry2.insert(0,str(round(float(testListeKategorien.__getitem__(length)), 2)))


    entryAll = tk.Entry(getValueAllKategorien_AllYears_ListWind, font=("Arial", 20), width=30)
    entryAll.pack(pady=10, padx=10)
    entryAll.insert(0, round(float(AllKategorienValuesTogether), 2))
    entryAll.insert(0, "Summe Ausgaben: ")



    entryGehalt = tk.Entry(getValueAllKategorien_AllYears_ListWind, font=("Arial", 20), width=30)
    entryGehalt.pack(pady=10, padx=10)
    entryGehalt.insert(0, str(round(float(gehaltAll), 2)))
    entryGehalt.insert(0, "Summe Einnahmen: ")

    Relation = float(gehaltAll) - float(AllKategorienValuesTogether)
    print(Relation)
    Relation = round(Relation, 2)

    if Relation < 0:
        print("Verlust")
        entryRelation = tk.Entry(getValueAllKategorien_AllYears_ListWind, font=("Arial", 20), width=30, fg="red")
        entryRelation.pack(pady=10, padx=10)
        entryRelation.insert(0, str(round(Relation, 2)))
        entryRelation.insert(0, "Verlust von ")
    else:  # Auch wenn Relation == 0
        print("Gewinn")
        entryRelation = tk.Entry(getValueAllKategorien_AllYears_ListWind, font=("Arial", 20), width=30, fg="green")
        entryRelation.pack(pady=10, padx=10)
        entryRelation.insert(0, str(round(Relation, 2)))
        entryRelation.insert(0, "Gewinn von ")



    getValueAllKategorien_AllYears_ListWind.mainloop()


def getAllYears_Kategorien_Diagramm():
    global Jahre
    Jahre = []
    global KategorienListe_allYears
    KategorienListe_allYears = []
    global KategorienValueListe
    KategorienValueListe = []

    YearsListItem = str(YearsList.__getitem__(0))
    wb = load_workbook("Finanzen" + YearsListItem + ".xlsx")
    ws = wb["Finanzen"]
    for row in range(2, ws.max_row + 1):
            KategorienListe_allYears.append(str(ws[str("A") + str(row)].value))

    #get gehaltAll
    gehaltAll = 0
    global AllKategorienValuesTogether
    AllKategorienValuesTogether = 0
    testListeKategorien = []
    for length in range(0, YearsList.__len__()):
        print(YearsList.__len__())
        YearsListItem = str(YearsList.__getitem__(length))
        print(YearsList.__getitem__(length))

        wb = load_workbook("Finanzen" + YearsListItem + ".xlsx")
        ws = wb["Finanzen"]

        wbGehalt = load_workbook("Gehalt" + YearsListItem + ".xlsx")
        wsGehalt = wbGehalt["Gehalt"]
        Jahre.append(YearsListItem)
        
        for col in range(2, wsGehalt.max_column + 1):
            if col == int("1"):
                col = "A"
            if col == int("2"):
                col = "B"
            if col == int("3"):
                col = "C"
            if col == int("4"):
                col = "D"
            if col == int("5"):
                col = "E"
            if col == int("6"):
                col = "F"
            if col == int("7"):
                col = "G"
            if col == int("8"):
                col = "H"
            if col == int("9"):
                col = "I"
            if col == int("10"):
                col = "J"
            if col == int("11"):
                col = "K"
            if col == int("12"):
                col = "L"
            if col == int("13"):
                col = "M"

            gehaltDieserMonat = 0
            for row in range(2, wsGehalt.max_row + 1):
                print(row)
                gehaltDieserMonat += round(float(wsGehalt[str(col) + str(row)].value), 2)
            print("Gehalt dieser Monat: " + str(gehaltDieserMonat))

            gehaltAll += gehaltDieserMonat
            gehaltAll = round(gehaltAll, 2)
            if str(gehaltDieserMonat) == "0.0":
                print("Gehalt noch nicht eingetragen für den Monat " + str(wsGehalt[str(col) + str(1)].value))
                #showText("Das Gehalt für den Monat " + str(wsGehalt[str(col) + str(
                    #1)].value) + " wurde noch nicht eingetragen. Gehe dafür bitte zum Hauptfenster zurück und füge dein Gehalt links unten, unter <Gehalt> hinzu.")
            print("Gehalt all:" + str(gehaltAll))


        
        
        for row in range(2, ws.max_row + 1):
            print(row)
            KategorieValue = 0

            
            for col in range(2, ws.max_column + 1):
                print(col)
                if col == int("1"):
                    col = "A"
                if col == int("2"):
                    col = "B"
                if col == int("3"):
                    col = "C"
                if col == int("4"):
                    col = "D"
                if col == int("5"):
                    col = "E"
                if col == int("6"):
                    col = "F"
                if col == int("7"):
                    col = "G"
                if col == int("8"):
                    col = "H"
                if col == int("9"):
                    col = "I"
                if col == int("10"):
                    col = "J"
                if col == int("11"):
                    col = "K"
                if col == int("12"):
                    col = "L"
                if col == int("13"):
                    col = "M"
                KategorieValue += float(ws[str(col) + str(row)].value)
                AllKategorienValuesTogether += float(ws[str(col) + str(row)].value)
        
            testListeKategorien.append(KategorieValue)
    

    

    print(testListeKategorien)
    li1 = testListeKategorien #[1, 2, 3, 4, 5, 6, 7, 8]
    midpoint = len(li1) // 2
    first_half = li1[:midpoint]
    second_half = li1[midpoint:]
    li2 = []
    for i in range(len(first_half)):
        li2.append(round(first_half[i] + second_half[i], 2))


    KategorienValueListe = li2
        
    print(Jahre)
    print(KategorienListe_allYears)
    print(KategorienValueListe)
    print(gehaltAll)
    print(AllKategorienValuesTogether)
    Relation = float(gehaltAll) - float(AllKategorienValuesTogether)
    print(Relation)
    Relation = round(Relation, 2)


    #For remote control
    global pltKD
    pltKD = plt

    # Konvertieren Sie die Beträge in Floats
    betraege = [float(betrag) for betrag in testListeKategorien]

    # Erstellen Sie das Diagramm
    fig, ax = pltKD.subplots(figsize=(16, 9), facecolor='#f0f0f0')
    ax.set_facecolor('#ffffff')
    fig.canvas.manager.set_window_title(f"Mein Finanzüberblick {Jahre}")

    # Balkenbreite und Positionen
    bar_width = 0.7
    r = np.arange(len(KategorienListe_allYears))

    # Erstellen Sie die Balken
    bars = ax.bar(r, betraege, color='#FF6B6B', width=bar_width, edgecolor='white', linewidth=0.7, alpha=0.8, label="Ausgaben")
    ax.legend(fontsize=14, loc='upper left', frameon=True, facecolor='white', edgecolor='gray')

    # Beschriftungen und Titel
    ax.set_xlabel('Kategorien', fontsize=16, fontweight='bold', labelpad=15)
    ax.set_ylabel('Beträge (€)', fontsize=16, fontweight='bold', labelpad=15)
    ax.set_title(f'Ausgaben aller Kategorien der Jahre {Jahre}', fontsize=24, fontweight='bold', pad=20)

    # Rotieren Sie die x-Achsen-Beschriftungen für bessere Lesbarkeit
    pltKD.xticks(r, KategorienListe_allYears, rotation=45, ha='right', fontsize=12)

    # Formatieren Sie die y-Achse als Währung
    def currency_formatter(x, p):
        return f"{x:,.0f} €" ##########################################################Hier mit '2f' die nachkommastellen bearbeiten
    ax.yaxis.set_major_formatter(FuncFormatter(currency_formatter))

    # Fügen Sie Werte über den Balken hinzu
    for bar in bars:
        height = bar.get_height()
        ax.text(bar.get_x() + bar.get_width()/2., height,
                f'{height:,.0f} €',##########################################################Hier mit '2f' die nachkommastellen bearbeiten
                ha='center', va='bottom', fontsize=10, fontweight='bold')

    # Fügen Sie eine Textbox mit Gesamtinformationen hinzu
    props = dict(boxstyle='round,pad=0.5', facecolor='white', alpha=0.7, edgecolor='gray')
    info_text = (f"Einnahmen: {gehaltAll:,.2f} €\n"
                 f"Ausgaben: {AllKategorienValuesTogether:,.2f} €\n"
                 f"Differenz: {Relation:,.2f} €")
    ax.text(0.992, 0.98, info_text, transform=ax.transAxes, 
            fontsize=14, verticalalignment='top', horizontalalignment='right',
            bbox=props)

    # Fügen Sie ein Gitternetz hinzu
    ax.grid(True, linestyle='--', alpha=0.7, axis='y')

    # Entfernen Sie die obere und rechte Achse
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)

    # Fügen Sie einen Rahmen um den Plotbereich hinzu
    for spine in ax.spines.values():
        spine.set_edgecolor('#CCCCCC')

    # Fügen Sie einen Farbverlauf als Hintergrund hinzu
    gradient = np.linspace(0, 1, 256).reshape(1, -1)
    gradient = np.vstack((gradient, gradient))
    ax.imshow(gradient, extent=[ax.get_xlim()[0], ax.get_xlim()[1], ax.get_ylim()[0], ax.get_ylim()[1]], 
              aspect='auto', alpha=0.1, cmap='coolwarm')

    # Fügen Sie ein Titelfeld hinzu
    title_box = patches.Rectangle((0, 1.02), 1, 0.08, transform=ax.transAxes, fill=True,
                                  facecolor='lightgray', edgecolor='none', alpha=0.8)
    ax.add_patch(title_box)

    # Add a watermark
    fig.text(0.95, 0, 'Jonas Gaiser', fontsize=12, color='gray', ha='right', va='bottom', alpha=0.5)

    # Passen Sie das Layout an
    pltKD.tight_layout()

    # Zeigen Sie das Diagramm an
    pltKD.show()




# Menu
menubar = tk.Menu(root)
root.config(menu=menubar)

# Datei-Menü
file_menu = tk.Menu(menubar, tearoff=0)
file_menu.add_command(label="Leere Einträge Suche", command=CheckForOpenEntrys)
file_menu.add_separator()
file_menu.add_command(label="Beenden", command=root.quit)
menubar.add_cascade(label="Datei", menu=file_menu)

# Info-Menü
info_menu = tk.Menu(menubar, tearoff=0)
info_menu.add_command(label="Farbcodes", command=FarbcodesErklärung)
info_menu.add_command(label="CodeInfo", command=showInfosAboutPythonWorkbook)
info_menu.add_command(label="Interessante Fakten", command=InterestFacts)
info_menu.add_separator()
info_menu.add_command(label="Über", command=showCoder)
menubar.add_cascade(label="Info", menu=info_menu)




# Oberen Bereich erstellen
oben = tk.Frame(root, padx=50)
oben.pack(side='top', fill='both', expand=True)



# Unteren Bereich erstellen
unten = tk.Frame(root, padx=50)
unten.pack(side='bottom', fill='x', expand=True)
#unten.pack(side="bottom", fill="both", expand=True)



#-----------------------------------------------------------------------------------------------------------------------



# Oben Links
oben_links = tk.Frame(oben, bg="#0093AF")
oben_links.pack(side='left', fill='both', padx=7.5)#, pady=7.5)

oben_linksFrame1 = tk.Frame(oben_links, bg="#606c84")
oben_linksFrame1.pack(side='top', fill='both', padx=7.5, pady=7.5, expand=True)

oben_linksFrame2 = tk.Frame(oben_links, bg="#606c84")
oben_linksFrame2.pack(side='bottom', fill='both', padx=7.5, pady=7.5, expand=True)


# Notebook (Tabs)

styleTab1 = ttk.Style()
styleTab1.configure("Tab1.TLabel", foreground="black", background="white")

styleTab2 = ttk.Style()
styleTab2.configure("Tab2.TLabel", foreground="black", background="white")


AuslesenEinzelÜbersicht = tk.Label(oben_linksFrame1, text="Einzelübersicht", font=("Arial", 15))
AuslesenEinzelÜbersicht.pack(fill='both', padx=10, pady=10, side=tk.TOP)
notebook = ttk.Notebook(oben_linksFrame1)
notebook.pack(expand=True, fill='both', padx=7.5, pady=7.5)


#frame2
AuslesenJahresÜbersicht = tk.Label(oben_linksFrame2, text="Jahresübersicht", font=("Arial", 15))
AuslesenJahresÜbersicht.pack(fill='both', padx=10, pady=10)

notebook2 = ttk.Notebook(oben_linksFrame2)
notebook2.pack(expand=True, fill='both', padx=7.5, pady=7.5)

# Erster Tab
tab1_olF2 = ttk.Frame(notebook2, style="Tab1.TLabel")
notebook2.add(tab1_olF2, text='Monate auslesen')

tab1_textYear = tk.Label(tab1_olF2, text="Jahr:", fg="black", bg="white", font=("Arial", 15))
tab1_textYear.pack()

YearsListName = "YearsList.txt"
YearsList = laden_liste(YearsListName)  #Checkt ob Years Liste vorhanden ist

YearvarAllMonthAuslesen = StringVar()
YearvarAllMonthAuslesen = ttk.Combobox(tab1_olF2, textvariable=YearvarAllMonthAuslesen, font=("Arial", 15))
YearvarAllMonthAuslesen.bind('<<ComboboxSelected>>')
YearvarAllMonthAuslesen['values'] = (YearsList)
YearvarAllMonthAuslesen.state(["readonly"])
YearvarAllMonthAuslesen.pack(padx=10, pady=10)



Btn = tk.Button(tab1_olF2, text="Monate auslesen", font=("Arial", 15), bg="#F25E24", fg="black", command= lambda: BereiteGetAllMonthsVor())
Btn.pack(fill='x', padx=7.5, pady=7.5, side=tk.BOTTOM)



def BereiteGetAllMonthsVor():
    if YearvarAllMonthAuslesen.get() == "":
        print("Leeres Feld vorhanden.")
        empty_fields()
    else:
        GetAllMonths(str(YearvarAllMonthAuslesen.get()))



# Zweiter Tab
tab2_olF2 = ttk.Frame(notebook2, style="Tab2.TLabel")
notebook2.add(tab2_olF2, text='Kategorien auslesen')

tab2_textYear = tk.Label(tab2_olF2, text="Jahr:", fg="black", bg="white", font=("Arial", 15))
tab2_textYear.pack()

YearsListName = "YearsList.txt"
YearsList = laden_liste(YearsListName)  #Checkt ob Years Liste vorhanden ist

YearvarAllKategorieAuslesen = StringVar()
YearvarAllKategorieAuslesen = ttk.Combobox(tab2_olF2, textvariable=YearvarAllKategorieAuslesen, font=("Arial", 15))
YearvarAllKategorieAuslesen.bind('<<ComboboxSelected>>')
YearvarAllKategorieAuslesen['values'] = (YearsList)
YearvarAllKategorieAuslesen.state(["readonly"])
YearvarAllKategorieAuslesen.pack(padx=10, pady=10)

Btn2 = tk.Button(tab2_olF2, text="Kategorien auslesen", font=("Arial", 15), bg="#F25E24", fg="black", command= lambda: BereiteGetAllKategorienVor())
Btn2.pack(fill='both', padx=7.5, pady=7.5, side=tk.BOTTOM)

def BereiteGetAllKategorienVor():
    if YearvarAllKategorieAuslesen.get() == "":
        print("Leeres Feld vorhanden.")
        empty_fields()
    else:
        GetAllKategories(str(YearvarAllKategorieAuslesen.get()))




# Frame1
# Erster Tab


tab1_ol = ttk.Frame(notebook, style="Tab1.TLabel")
notebook.add(tab1_ol, text='Monat auslesen')


tab1_textYear = tk.Label(tab1_ol, text="Jahr:", fg="black", bg="white", font=("Arial", 15))
tab1_textYear.pack()

YearsListName = "YearsList.txt"
YearsList = laden_liste(YearsListName)  #Checkt ob Years Liste vorhanden ist

YearvarMonthAuslesen = StringVar()
YearvarMonthAuslesen = ttk.Combobox(tab1_ol, textvariable=YearvarMonthAuslesen, font=("Arial", 10))
YearvarMonthAuslesen.bind('<<ComboboxSelected>>')
YearvarMonthAuslesen['values'] = (YearsList)
YearvarMonthAuslesen.state(["readonly"])
YearvarMonthAuslesen.pack()

tab1_textMonth = tk.Label(tab1_ol, text="Monat:", fg="black", bg="white", font=("Arial", 15))
tab1_textMonth.pack()

MonthvarMonthAuslesen = StringVar()
MonthvarMonthAuslesen = ttk.Combobox(tab1_ol, textvariable=MonthvarMonthAuslesen, font=("Arial", 10))
MonthvarMonthAuslesen.bind('<<ComboboxSelected>>')
MonthvarMonthAuslesen['values'] = ("Januar", "Februar", "März", "April", "Mai", "Juni", "Juli", "August", "September", "Oktober", "November", "Dezember")
MonthvarMonthAuslesen.state(["readonly"])
MonthvarMonthAuslesen.pack()


tab1_MonthBtn = tk.Button(tab1_ol, command= lambda: BereitegetValueOfMonthAllKategoriesVor(), text="Monat auslesen", bg = "#F25E24", fg="black", font=("Arial", 15), pady=10)
tab1_MonthBtn.pack(pady=10)

def BereitegetValueOfMonthAllKategoriesVor():
    if str(YearvarMonthAuslesen.get()) == "" or str(MonthvarMonthAuslesen.get()) == "":
        print("Leeres Feld vorhanden")
        empty_fields()
    else:
        getValueOfMonthAllKategories(str(YearvarMonthAuslesen.get()), str(MonthvarMonthAuslesen.get()))





# Zweiter Tab
tab2_ol = ttk.Frame(notebook, style="Tab2.TLabel")
notebook.add(tab2_ol, text="Kategorie auslesen")

tab2_textYear = tk.Label(tab2_ol, text="Jahr:", fg="black", bg="white", font=("Arial", 15))
tab2_textYear.pack()

YearsListName = "YearsList.txt"
YearsList = laden_liste(YearsListName)  #Checkt ob Years Liste vorhanden ist

YearvarKategorieAuslesen = StringVar()
YearvarKategorieAuslesen = ttk.Combobox(tab2_ol, textvariable=YearvarKategorieAuslesen, font=("Arial", 10))
YearvarKategorieAuslesen.bind('<<ComboboxSelected>>')
YearvarKategorieAuslesen['values'] = (YearsList)
YearvarKategorieAuslesen.state(["readonly"])
YearvarKategorieAuslesen.pack()

tab2_textKategorie = tk.Label(tab2_ol, text="Kategorie:", fg="black", bg="white", font=("Arial", 15))
tab2_textKategorie.pack()

KategorieListName = "KategorieList.txt"
KategorieList = laden_liste(KategorieListName)

KategorievarKategorieAuslesen = StringVar()
KategorievarKategorieAuslesen = ttk.Combobox(tab2_ol, textvariable=KategorievarKategorieAuslesen, font=("Arial", 10))
KategorievarKategorieAuslesen.bind('<<ComboboxSelected>>')
KategorievarKategorieAuslesen['values'] = (KategorieList)
KategorievarKategorieAuslesen.state(["readonly"])
KategorievarKategorieAuslesen.pack()


tab2_KategorieBtn = tk.Button(tab2_ol, command= lambda: BereitegetValueOfKategorieAllMonthVor(), text="Kategorie auslesen", bg = "#F25E24", fg="black", font=("Arial", 15), pady=10)
tab2_KategorieBtn.pack(pady=10)

def BereitegetValueOfKategorieAllMonthVor():
    if str(YearvarKategorieAuslesen.get()) == "" or str(KategorievarKategorieAuslesen.get()) == "":
        print("Leeres Feld vorhanden")
        empty_fields()
    else:
        getValueOfKategorieAllMonths(str(YearvarKategorieAuslesen.get()), str(KategorievarKategorieAuslesen.get()))




#-----------------------------------------------------------------------------------------------------------------------




# Oben Rechts
oben_rechts = tk.Frame(oben, bg="#0093AF")
oben_rechts.pack(side='right', fill='x', padx=7.5, expand=True) # pady=7.5

notebook = ttk.Notebook(oben_rechts)
notebook.pack(expand=True, fill='x', padx=7.5, pady=7.5)



#Ausgabentabelle

tab1_fullViewAusgaben = ttk.Frame(notebook, style="Tab1.TLabel")
notebook.add(tab1_fullViewAusgaben, text='Ausgaben-Tabelle')


frameUmrandung = tk.Frame(tab1_fullViewAusgaben, bg="#606c84")
frameUmrandung.pack(fill="x", padx=7.5, pady=7.5)


oben_settingsFrame = tk.Frame(frameUmrandung, bg="#9c1c78")
oben_settingsFrame.pack(side='top', fill="x", padx=7.5, pady=7.5)

oben_settings = tk.Frame(oben_settingsFrame, bg="white")
oben_settings.pack(side='top', fill="x", padx=7.5, pady=7.5)


YearsListName = "YearsList.txt"
YearsList = laden_liste(YearsListName)  #Checkt ob Years Liste vorhanden ist

SubmitBtnSettings = tk.Button(oben_settings, text="Neu Laden", fg="black", bg="#F25E24", font=("Arial", 15), command= lambda: BereiteGetYearForViewVor(str(YearvarAll.get())))
SubmitBtnSettings.pack(side="right", expand=True)


YearvarAll = StringVar()
YearvarAll = ttk.Combobox(oben_settings, textvariable=YearvarAll, font=("Arial", 10))
YearvarAll.bind('<<ComboboxSelected>>')
YearvarAll['values'] = (YearsList)
YearvarAll.state(["readonly"])
YearvarAll.pack(pady=5, side="right", expand=True)


tab1_textYear = tk.Label(oben_settings, text="Ausgaben aus dem Jahr:", fg="black", bg="white", font=("Arial", 15))
tab1_textYear.pack(side="right", expand=True)




def BereiteGetYearForViewVor(year):
    """This Function will open the file explorer and assign the chosen file path to label_file"""
    #filename = "Finanzen" + str(YearvarAll.get()) + ".xlsx"
    global filename
    filename = "Finanzen" + str(year) + ".xlsx"
    global label_file
    label_file["text"] = filename
    if str(year) == "":
        empty_fields()
    else:
        Load_excel_data()




unten_viewFrame2 = tk.Frame(frameUmrandung, bg="#9c1c78")
unten_viewFrame2.pack(side="bottom", fill="both", padx=7.5, pady=7.5)
unten_viewFrame = tk.Frame(unten_viewFrame2, bg="white")
unten_viewFrame.pack(side="bottom", fill="both", padx=7.5, pady=7.5)

unten_view = tk.Frame(unten_viewFrame, bg="white")
unten_view.pack(side='bottom', fill="both", padx=7.5, pady=7.5)

# Frame for TreeView
frame1 = tk.LabelFrame(unten_viewFrame, text="Excel Data")
frame1.pack(side="top", expand=True, padx=7.5, pady=7.5)

# Frame for open file dialog
file_frame = tk.LabelFrame(frameUmrandung, text="Open File")
file_frame.pack(side="bottom", expand=True, padx=7.5,pady=7.5)


# The file/file path text
label_file = ttk.Label(file_frame, text="No File Selected")
label_file.pack(side="top", expand=True, padx=7.5)



## Treeview Widget
tv1 = ttk.Treeview(frame1)
tv1.pack(side="top", expand=True, padx=7.5, pady=7.5) # set the height and width of the widget to 100% of its container (frame1).

treescrolly = tk.Scrollbar(frame1, orient="vertical", command=tv1.yview) # command means update the yaxis view of the widget
treescrollx = tk.Scrollbar(frame1, orient="horizontal", command=tv1.xview) # command means update the xaxis view of the widget
tv1.configure(xscrollcommand=treescrollx.set, yscrollcommand=treescrolly.set) # assign the scrollbars to the Treeview Widget
treescrollx.pack(side="bottom", fill="x") # make the scrollbar fill the x axis of the Treeview widget
treescrolly.pack(side="right", fill="y") # make the scrollbar fill the y axis of the Treeview widget





def Load_excel_data():
    """If the file selected is valid this will load the file into the Treeview"""
    file_path = label_file["text"]
    try:
        excel_filename = r"{}".format(file_path)
        if excel_filename[-4:] == ".csv":
            df = pd.read_csv(excel_filename)
        else:
            df = pd.read_excel(excel_filename)

    except ValueError:
        tk.messagebox.showerror("Information", "The file you have chosen is invalid")
        return None
    except FileNotFoundError:
        tk.messagebox.showerror("Information", f"No such file as {file_path}")
        return None

    clear_data()
    tv1["column"] = list(df.columns)
    tv1["show"] = "headings"
    for column in tv1["columns"]:
        tv1.heading(column, text=column) # let the column heading = column name

    df_rows = df.to_numpy().tolist() # turns the dataframe into a list of lists
    for row in df_rows:
        tv1.insert("", "end", values=row) # inserts each list into the treeview. For parameters see https://docs.python.org/3/library/tkinter.ttk.html#tkinter.ttk.Treeview.insert
    return None


def clear_data():
    tv1.delete(*tv1.get_children())
    return None




## Einnahme Tabelle



tab1_fullViewGehalt = ttk.Frame(notebook, style="Tab1.TLabel")
notebook.add(tab1_fullViewGehalt, text='Einnahmen-Tabelle')


frameUmrandung = tk.Frame(tab1_fullViewGehalt, bg="#606c84")
frameUmrandung.pack(fill="both", padx=7.5, pady=7.5)


oben_settingsFrameGe = tk.Frame(frameUmrandung, bg="#208844")
oben_settingsFrameGe.pack(side='top', fill="both", padx=7.5, pady=7.5)

oben_settingsGe = tk.Frame(oben_settingsFrameGe, bg="white")
oben_settingsGe.pack(side='top', fill="both", padx=7.5, pady=7.5)


YearsListName = "YearsList.txt"
YearsList = laden_liste(YearsListName)  #Checkt ob Years Liste vorhanden ist

SubmitBtnSettings = tk.Button(oben_settingsGe, text="Neu Laden", fg="black", bg="#F25E24", font=("Arial", 15), command= lambda: BereiteGetYearForViewVorGe(str(YearvarAllGehalt.get())))
SubmitBtnSettings.pack(side="right", expand=True)


YearvarAllGehalt = StringVar()
YearvarAllGehalt = ttk.Combobox(oben_settingsGe, textvariable=YearvarAllGehalt, font=("Arial", 10))
YearvarAllGehalt.bind('<<ComboboxSelected>>')
YearvarAllGehalt['values'] = (YearsList)
YearvarAllGehalt.state(["readonly"])
YearvarAllGehalt.pack(pady=5, side="right", expand=True)


tab1_textYear = tk.Label(oben_settingsGe, text="Einnahmen aus dem Jahr:", fg="black", bg="white", font=("Arial", 15))
tab1_textYear.pack(side="right", expand=True)




def BereiteGetYearForViewVorGe(year):
    """This Function will open the file explorer and assign the chosen file path to label_file"""
    #filename = "Finanzen" + str(YearvarAll.get()) + ".xlsx"
    global filenameGe
    filenameGe = "Gehalt" + str(year) + ".xlsx"
    global label_fileGe
    label_fileGe["text"] = filenameGe
    if str(year) == "":
        empty_fields()
    else:
        Load_excel_dataGe()





unten_viewFrame2Ge = tk.Frame(frameUmrandung, bg="#208844")
unten_viewFrame2Ge.pack(side="bottom", fill="both", padx=7.5, pady=7.5)
unten_viewFrameGe = tk.Frame(unten_viewFrame2Ge, bg="white")
unten_viewFrameGe.pack(side="bottom", fill="both", padx=7.5, pady=7.5)

unten_viewGe = tk.Frame(unten_viewFrameGe, bg="white")
unten_viewGe.pack(side='bottom', fill="both", padx=7.5, pady=7.5)

# Frame for TreeView
frame1Ge = tk.LabelFrame(unten_viewFrameGe, text="Excel Data")
frame1Ge.pack(side="top", expand=True, padx=7.5, pady=7.5)

# Frame for open file dialog
file_frameGe = tk.LabelFrame(frameUmrandung, text="Open File")
file_frameGe.pack(side="bottom", expand=True, padx=7.5,pady=7.5)



# The file/file path text
label_fileGe = ttk.Label(file_frameGe, text="No File Selected")
label_fileGe.pack(side="top", expand=True, padx=7.5)


## Treeview Widget
tv1Ge = ttk.Treeview(frame1Ge)
tv1Ge.pack(side="top", expand=True, padx=7.5, pady=7.5) # set the height and width of the widget to 100% of its container (frame1).

treescrollyGe = tk.Scrollbar(frame1Ge, orient="vertical", command=tv1Ge.yview) # command means update the yaxis view of the widget
treescrollxGe = tk.Scrollbar(frame1Ge, orient="horizontal", command=tv1Ge.xview) # command means update the xaxis view of the widget
tv1Ge.configure(xscrollcommand=treescrollxGe.set, yscrollcommand=treescrollyGe.set) # assign the scrollbars to the Treeview Widget
treescrollxGe.pack(side="bottom", fill="x") # make the scrollbar fill the x axis of the Treeview widget
treescrollyGe.pack(side="right", fill="y") # make the scrollbar fill the y axis of the Treeview widget






def Load_excel_dataGe():
    """If the file selected is valid this will load the file into the Treeview"""
    file_pathGe = label_fileGe["text"]
    try:
        excel_filenameGe = r"{}".format(file_pathGe)
        if excel_filenameGe[-4:] == ".csv":
            df = pd.read_csv(excel_filenameGe)
        else:
            df = pd.read_excel(excel_filenameGe)

    except ValueError:
        tk.messagebox.showerror("Information", "The file you have chosen is invalid")
        return None
    except FileNotFoundError:
        tk.messagebox.showerror("Information", f"No such file as {file_pathGe}")
        return None

    clear_dataGe()
    tv1Ge["column"] = list(df.columns)
    tv1Ge["show"] = "headings"
    for columnGe in tv1Ge["columns"]:
        tv1Ge.heading(columnGe, text=columnGe) # let the column heading = column name

    df_rowsGe = df.to_numpy().tolist() # turns the dataframe into a list of lists
    for rowGe in df_rowsGe:
        tv1Ge.insert("", "end", values=rowGe) # inserts each list into the treeview. For parameters see https://docs.python.org/3/library/tkinter.ttk.html#tkinter.ttk.Treeview.insert
    return None


def clear_dataGe():
    tv1Ge.delete(*tv1Ge.get_children())
    return None





# Diagrams

tab3_diagrams = ttk.Frame(notebook, style="Tab1.TLabel")
notebook.add(tab3_diagrams, text='Diagramme')

titleFrame = tk.Frame(tab3_diagrams)
titleFrame.pack(fill="x")  # Füllt die gesamte Breite aus

# Erstellen Sie drei Unterframes innerhalb des titleFrame
leftFrame = tk.Frame(titleFrame, width=100, bg="white")  # Platzhalter links
middleFrame = tk.Frame(titleFrame, bg="white")
rightFrame = tk.Frame(titleFrame, width=100, bg="white")  # Platzhalter rechts

leftFrame.pack(side="left", fill="y")
middleFrame.pack(side="left", expand=True, fill="both")
rightFrame.pack(side="right", fill="y")

# Titel in der Mitte
titleDiagrams = tk.Label(middleFrame, text="Diagramme", font=("Arial", 25), bg="white", fg="black")
titleDiagrams.pack(expand=True)

# Button ganz rechts
InfoDiagramsBtn = tk.Button(rightFrame, text="i", command=lambda: showInfoDiagrams(), fg="black", bg="lightblue", font=("Arial", 25), width=4, height=1)
InfoDiagramsBtn.pack(side="right", padx=5, pady=5)  

def showInfoDiagrams():
    showText(
            "Wichtig: Die Diagramme dienen zur übersichtlichen Darstellung der Werte. " \
            "Daher werden die Ausgaben/Einnahmen in ganzen Zahlen dargestellt. "\
            "Für nähere Details, nutze bitte die anderen Funktionen dieser Software "\
            "oder beachte die Informations-Box oben rechts in den Diagrammen.")



notebook_diagrams = ttk.Notebook(tab3_diagrams)
notebook_diagrams.pack(expand=True, fill='both', padx=7.5, pady=7.5)



# Farben für die Abteile
colors = ["#606c84", "#606c84", "#606c84", "#606c84"]


# Erstellen der äußeren Frames für die horizontale Aufteilung
diagrams1_top_frame = tk.Frame(notebook_diagrams)
diagrams1_bottom_frame = tk.Frame(notebook_diagrams)

diagrams1_top_frame.pack(side="top", fill="both", expand=True)
diagrams1_bottom_frame.pack(side="bottom", fill="both", expand=True)

# Erstellen und Packen der inneren Frames für die vertikale Aufteilung
diagrams1_top_left_frame = tk.Frame(diagrams1_top_frame, bg=colors[0])
diagrams1_top_right_frame = tk.Frame(diagrams1_top_frame, bg=colors[1])
diagrams1_bottom_left_frame = tk.Frame(diagrams1_bottom_frame, bg=colors[2])
diagrams1_bottom_right_frame = tk.Frame(diagrams1_bottom_frame, bg=colors[3])

diagrams1_top_left_frame.pack(side="left", fill="both", expand=True)
diagrams1_top_right_frame.pack(side="right", fill="both", expand=True)
diagrams1_bottom_left_frame.pack(side="left", fill="both", expand=True)
diagrams1_bottom_right_frame.pack(side="right", fill="both", expand=True)

# Einrichten um die Ränder zu erzeugen (eigentliche frames)
diagrams1_top_left = tk.Frame(diagrams1_top_left_frame, bg="white")
diagrams1_top_right = tk.Frame(diagrams1_top_right_frame, bg="white")
diagrams1_bottom_left = tk.Frame(diagrams1_bottom_left_frame, bg="white")
diagrams1_bottom_right = tk.Frame(diagrams1_bottom_right_frame, bg="white")

diagrams1_top_left.pack(side="left", fill="both", expand=True, padx=7.5, pady=7.5)
diagrams1_top_right.pack(side="right", fill="both", expand=True, padx=7.5, pady=7.5)
diagrams1_bottom_left.pack(side="left", fill="both", expand=True, padx=7.5, pady=7.5)
diagrams1_bottom_right.pack(side="right", fill="both", expand=True, padx=7.5, pady=7.5)



### Top-Left # Jahresübersicht Monate

# Widgets

diagrams1_top_left_title = tk.Label(diagrams1_top_left, text="Jahresübersicht (Monate)", bg="white", fg="green", font=("Arial", 20))
diagrams1_top_left_title.pack()

YearsListName = "YearsList.txt"
YearsList = laden_liste(YearsListName)  #Checkt ob Years Liste vorhanden ist

diagrams1_top_left_textYear = tk.Label(diagrams1_top_left, text="Jahr:", fg="black", bg="white", font=("Arial", 15))
diagrams1_top_left_textYear.pack()

diagrams1_top_leftYearvar = StringVar()
diagrams1_top_leftYearvar = ttk.Combobox(diagrams1_top_left, textvariable=diagrams1_top_leftYearvar, font=("Arial", 10))
diagrams1_top_leftYearvar.bind('<<ComboboxSelected>>')
diagrams1_top_leftYearvar['values'] = (YearsList)
diagrams1_top_leftYearvar.state(["readonly"])
diagrams1_top_leftYearvar.pack(pady=7.5)


diagrams1_top_leftbutton = tk.Button(diagrams1_top_left, font=("Arial", 15), text="Laden", bg="#F25E24", fg="black", command= lambda: BereitePlotMonthsVor(str(diagrams1_top_leftYearvar.get())))
diagrams1_top_leftbutton.pack()



def BereitePlotMonthsVor(year):
    if year == "":
        print("Leeres Feld vorhanden.")
        empty_fields()
    else:
        plot_Months(year)





### Top-Right # Jahreübersicht Kategorien

# Widgets

diagrams1_top_right_title = tk.Label(diagrams1_top_right, text="Jahresübersicht (Kategorien)", bg="white", fg="green", font=("Arial", 20))
diagrams1_top_right_title.pack()

YearsListName = "YearsList.txt"
YearsList = laden_liste(YearsListName)  #Checkt ob Years Liste vorhanden ist

diagrams1_top_right_textYear = tk.Label(diagrams1_top_right, text="Jahr:", fg="black", bg="white", font=("Arial", 15))
diagrams1_top_right_textYear.pack()

diagrams1_top_rightYearvar = StringVar()
diagrams1_top_rightYearvar = ttk.Combobox(diagrams1_top_right, textvariable=diagrams1_top_rightYearvar, font=("Arial", 10))
diagrams1_top_rightYearvar.bind('<<ComboboxSelected>>')
diagrams1_top_rightYearvar['values'] = (YearsList)
diagrams1_top_rightYearvar.state(["readonly"])
diagrams1_top_rightYearvar.pack(pady=7.5)


diagrams1_top_rightbutton = tk.Button(diagrams1_top_right, font=("Arial", 15), text="Laden", bg="#F25E24", fg="black", command= lambda: BereitePlotKategorienVor(str(diagrams1_top_rightYearvar.get())))
diagrams1_top_rightbutton.pack()

def BereitePlotKategorienVor(year):
    if year == "":
        print("Leeres Feld vorhanden.")
        empty_fields()
    else:
        plot_kategorien(year)


### Bottom-Left # Einzelübersicht Monat

#Widgets

diagrams1_bottom_left_title = tk.Label(diagrams1_bottom_left, text="Einzelübersicht (Monat)", bg="white", fg="green", font=("Arial", 20))
diagrams1_bottom_left_title.pack()

diagrams1_bottom_left_YearsListName = "YearsList.txt"
diagrams1_bottom_left_YearsList = laden_liste(YearsListName)  #Checkt ob Years Liste vorhanden ist

diagrams1_bottom_left_tab1_textYear = tk.Label(diagrams1_bottom_left, text="Jahr:", fg="black", bg="white", font=("Arial", 15))
diagrams1_bottom_left_tab1_textYear.pack()

diagrams1_bottom_left_Yearvar = StringVar()
diagrams1_bottom_left_Yearvar = ttk.Combobox(diagrams1_bottom_left, textvariable=diagrams1_bottom_left_Yearvar, font=("Arial", 10))
diagrams1_bottom_left_Yearvar.bind('<<ComboboxSelected>>')
diagrams1_bottom_left_Yearvar['values'] = (YearsList)
diagrams1_bottom_left_Yearvar.state(["readonly"])
diagrams1_bottom_left_Yearvar.pack()

diagrams1_bottom_left_tab1_textMonth = tk.Label(diagrams1_bottom_left, text="Monat:", fg="black", bg="white", font=("Arial", 15))
diagrams1_bottom_left_tab1_textMonth.pack()

diagrams1_bottom_left_Monthvar = StringVar()
diagrams1_bottom_left_Monthvar = ttk.Combobox(diagrams1_bottom_left, textvariable=diagrams1_bottom_left_Monthvar, font=("Arial", 10))
diagrams1_bottom_left_Monthvar.bind('<<ComboboxSelected>>')
diagrams1_bottom_left_Monthvar['values'] = ("Januar", "Februar", "März", "April", "Mai", "Juni", "Juli", "August", "September", "Oktober", "November", "Dezember")
diagrams1_bottom_left_Monthvar.state(["readonly"])
diagrams1_bottom_left_Monthvar.pack()

diagrams1_bottom_left_button = tk.Button(diagrams1_bottom_left, font=("Arial", 15), text="Laden", bg="#F25E24", fg="black", command= lambda: BereitePlot_kategorienMonthVor(str(diagrams1_bottom_left_Yearvar.get()), str(diagrams1_bottom_left_Monthvar.get())))
diagrams1_bottom_left_button.pack()

def BereitePlot_kategorienMonthVor(year, month):
    if year == "" or month == "":
        print("Leeres Feld vorhanden.")
        empty_fields()
    else:
        plot_kategorienMonth(year, month)



### Bottom-Right # Einzelübersicht Kategorie

diagrams1_bottom_right_title = tk.Label(diagrams1_bottom_right, text="Einzelübersicht (Kategorie)", bg="white", fg="green", font=("Arial", 20))
diagrams1_bottom_right_title.pack()

diagrams1_bottom_right_YearsListName = "YearsList.txt"
diagrams1_bottom_right_YearsList = laden_liste(diagrams1_bottom_right_YearsListName)  #Checkt ob Years Liste vorhanden ist

diagrams1_bottom_right_textYear = tk.Label(diagrams1_bottom_right, text="Jahr:", fg="black", bg="white", font=("Arial", 15))
diagrams1_bottom_right_textYear.pack()

diagrams1_bottom_right_Yearvar = StringVar()
diagrams1_bottom_right_Yearvar = ttk.Combobox(diagrams1_bottom_right, textvariable=diagrams1_bottom_right_Yearvar, font=("Arial", 10))
diagrams1_bottom_right_Yearvar.bind('<<ComboboxSelected>>')
diagrams1_bottom_right_Yearvar['values'] = (YearsList)
diagrams1_bottom_right_Yearvar.state(["readonly"])
diagrams1_bottom_right_Yearvar.pack()


diagrams1_bottom_right_KategorieListName = "KategorieList.txt"
diagrams1_bottom_right_KategorieList = laden_liste(diagrams1_bottom_right_KategorieListName)  #Checkt ob Years Liste vorhanden ist

diagrams1_bottom_right_textKategorie = tk.Label(diagrams1_bottom_right, text="Kategorie: ", fg="black", bg="white", font=("Arial", 15))
diagrams1_bottom_right_textKategorie.pack()

diagrams1_bottom_right_Kategorievar = StringVar()
diagrams1_bottom_right_Kategorievar = ttk.Combobox(diagrams1_bottom_right, textvariable=diagrams1_bottom_right_Kategorievar, font=("Arial", 10))
diagrams1_bottom_right_Kategorievar.bind('<<ComboboxSelected>>')
diagrams1_bottom_right_Kategorievar['values'] = (KategorieList)
diagrams1_bottom_right_Kategorievar.state(["readonly"])
diagrams1_bottom_right_Kategorievar.pack()

diagrams1_bottom_right_button = tk.Button(diagrams1_bottom_right, font=("Arial", 15), text="Laden", bg="#F25E24", fg="black", command= lambda: Bereiteplot_eineKategorieVor(str(diagrams1_bottom_right_Yearvar.get()), str(diagrams1_bottom_right_Kategorievar.get())))
diagrams1_bottom_right_button.pack()

def Bereiteplot_eineKategorieVor(year, kategorie):
    if year == "" or kategorie == "" :
        print("Leeres Feld vorhanden.")
        empty_fields()
    else:
        plot_eineKategorie(year, kategorie)



# mehrfache Jahresübersicht

tab4_allYears = ttk.Frame(notebook, style="Tab1.TLabel")
notebook.add(tab4_allYears, text='Alle Jahre Übersicht')

AllYears_left_frame = tk.Frame(tab4_allYears, bg="#606c84")
AllYears_right_frame = tk.Frame(tab4_allYears, bg="#606c84")

AllYears_left_frame.pack(side="left", fill="both", expand=True)
AllYears_right_frame.pack(side="right", fill="both", expand=True)

AllYears_left = tk.Frame(AllYears_left_frame, bg="white")
AllYears_left.pack(fill="both", expand=True, padx=7.5, pady=7.5)

AllYears_right = tk.Frame(AllYears_right_frame, bg="white")
AllYears_right.pack(fill="both", expand=True, padx=7.5, pady=7.5)

# left

AllYears_title = tk.Label(AllYears_left, text="Alle Jahre", font=("Arial", 23), bg="white")
AllYears_title.pack(side="top")
AllYears_title2 = tk.Label(AllYears_left, text="(Einzeln)", font=("Arial", 18), bg="white")
AllYears_title2.pack(side="top")

AllYears_listBtnFrame2 = tk.Frame(AllYears_left, bg="#0093AF")
AllYears_listBtnFrame2.pack(side="bottom", padx=7.5, pady=7.5, expand=True)
AllYears_listBtnFrame = tk.Frame(AllYears_listBtnFrame2, bg="white")
AllYears_listBtnFrame.pack(padx=7.5, pady=7.5)
AllYears_list_title = tk.Label(AllYears_listBtnFrame, text="     Alle Jahre auslesen (Liste)    ", font=("Arial", 25), bg="white")
AllYears_list_title.pack(padx=7.5, pady=7.5)
AllYears_listBtn = tk.Button(AllYears_listBtnFrame, text="Laden", font=("Arial", 25), bg="#F25E24", command=lambda: getValueAll_AllYears_List())
AllYears_listBtn.pack(side="bottom", padx=40, pady=40, fill="both", expand=True)



AllYears_diagrammBtnFrame2 = tk.Frame(AllYears_left, bg="#0093AF")
AllYears_diagrammBtnFrame2.pack(side="bottom", padx=7.5, pady=7.5, expand=True)
AllYears_diagrammBtnFrame = tk.Frame(AllYears_diagrammBtnFrame2, bg="white")
AllYears_diagrammBtnFrame.pack(padx=7.5, pady=7.5)
AllYears_diagramm_title = tk.Label(AllYears_diagrammBtnFrame, text="Alle Jahre auslesen (Diagramm)", font=("Arial", 25), bg="white")
AllYears_diagramm_title.pack(padx=7.5, pady=7.5)
AllYears_diagramBtn = tk.Button(AllYears_diagrammBtnFrame, text="Laden", font=("Arial", 25), bg="#F25E24", command=lambda: getValueAll_AllYears_Diagram())
AllYears_diagramBtn.pack(side="bottom", padx=40, pady=40, fill="both",expand=True)


# right
AllYears_Kategories_title = tk.Label(AllYears_right, text="Kategorien", font=("Arial", 23), bg="white")
AllYears_Kategories_title.pack(side="top")
AllYears_Kategories_title2 = tk.Label(AllYears_right, text="(Alle Jahre)", font=("Arial", 18), bg="white")
AllYears_Kategories_title2.pack(side="top")

AllYears_Kategories_listBtnFrame2 = tk.Frame(AllYears_right, bg="#0093AF")
AllYears_Kategories_listBtnFrame2.pack(side="bottom", padx=7.5, pady=7.5, expand=True)
AllYears_Kategories_listBtnFrame = tk.Frame(AllYears_Kategories_listBtnFrame2, bg="white")
AllYears_Kategories_listBtnFrame.pack(padx=7.5, pady=7.5)
AllYears_Kategories_list_title = tk.Label(AllYears_Kategories_listBtnFrame, text="     Alle Kategorien auslesen (Liste)    ", font=("Arial", 25), bg="white")
AllYears_Kategories_list_title.pack(padx=7.5, pady=7.5)
AllYears_Kategories_listBtn = tk.Button(AllYears_Kategories_listBtnFrame, text="Laden", font=("Arial", 25), bg="#F25E24", command=lambda: getAllYears_Kategorien_List())
AllYears_Kategories_listBtn.pack(side="bottom", padx=40, pady=40, fill="both", expand=True)



AllYears_Kategories_diagrammBtnFrame2 = tk.Frame(AllYears_right, bg="#0093AF")
AllYears_Kategories_diagrammBtnFrame2.pack(side="bottom", padx=7.5, pady=7.5, expand=True)
AllYears_Kategories_diagrammBtnFrame = tk.Frame(AllYears_Kategories_diagrammBtnFrame2, bg="white")
AllYears_Kategories_diagrammBtnFrame.pack(padx=7.5, pady=7.5)
AllYears_Kategories_diagramm_title = tk.Label(AllYears_Kategories_diagrammBtnFrame, text="Alle Kategorien auslesen (Diagramm)", font=("Arial", 25), bg="white")
AllYears_Kategories_diagramm_title.pack(padx=7.5, pady=7.5)
AllYears_Kategories_diagramBtn = tk.Button(AllYears_Kategories_diagrammBtnFrame, text="Laden", font=("Arial", 25), bg="#F25E24", command=lambda: getAllYears_Kategorien_Diagramm())
AllYears_Kategories_diagramBtn.pack(side="bottom", padx=40, pady=40, fill="both",expand=True)
#-----------------------------------------------------------------------------------------------------------------------






# Unten Links
unten_links = tk.Frame(unten, bg="#9c1c78")
unten_links.pack(side='left', fill="both", padx=7.5) # pady=7.5

#unten_links = tk.Frame(unten, bg=blau1)
#unten_links.pack(side='left', fill='both', expand=True, padx=7.5, pady=7.5)



# Notebook (Tabs)

styleTab1 = ttk.Style()
styleTab1.configure("Tab1.TLabel", foreground="black", background="white")

styleTab2 = ttk.Style()
styleTab2.configure("Tab2.TLabel", foreground="black", background="white")

styleTab3 = ttk.Style()
styleTab3.configure("Tab3.TLabel", foreground="black", background="white")

styleTab4 = ttk.Style()
styleTab4.configure("Tab4.TLabel", foreground="black", background="white")

notebook = ttk.Notebook(unten_links)
notebook.pack(expand=True, fill='both', padx=7.5, pady=7.5)


# Erster Tab
tab1 = ttk.Frame(notebook, style="Tab1.TLabel")
notebook.add(tab1, text='Ausgaben Hinzufügen')

# Zweiter Tab
tab2 = ttk.Frame(notebook, style="Tab2.TLabel")
notebook.add(tab2, text='Ausgaben Löschen')

# Dritter Tab
tab3 = ttk.Frame(notebook, style="Tab3.TLabel")
notebook.add(tab3, text="Ausgaben-Kategorien Hinzufügen")

# Vierter Tab
tab4 = ttk.Frame(notebook, style="Tab4.TLabel")
notebook.add(tab4, text="Ausgaben-Kategorien Löschen")




# Widgets für den ersten Tab


tab1_title = tk.Label(tab1, text="Ausgaben Hinzufügen", bg="white", fg="#9c1c78", font=("Arial", 25))
tab1_title.pack()

YearsListName = "YearsList.txt"
YearsList = laden_liste(YearsListName)  #Checkt ob Years Liste vorhanden ist

tab1_textYear = tk.Label(tab1, text="Jahr:", fg="black", bg="white", font=("Arial", 15))
tab1_textYear.pack()

Yearvar = StringVar()
Yearvar = ttk.Combobox(tab1, textvariable=Yearvar, font=("Arial", 10))
Yearvar.bind('<<ComboboxSelected>>')
Yearvar['values'] = (YearsList)
Yearvar.state(["readonly"])
Yearvar.pack()

tab1_textMonth = tk.Label(tab1, text="Monat:", fg="black", bg="white", font=("Arial", 15))
tab1_textMonth.pack()

Monthvar = StringVar()
Monthvar = ttk.Combobox(tab1, textvariable=Monthvar, font=("Arial", 10))
Monthvar.bind('<<ComboboxSelected>>')
Monthvar['values'] = ("Januar", "Februar", "März", "April", "Mai", "Juni", "Juli", "August", "September", "Oktober", "November", "Dezember")
Monthvar.state(["readonly"])
Monthvar.pack()

KategorieListName = "KategorieList.txt"
KategorieList = laden_liste(KategorieListName)  #Checkt ob Years Liste vorhanden ist

tab1_textKategorie = tk.Label(tab1, text="Kategorie: ", fg="black", bg="white", font=("Arial", 15))
tab1_textKategorie.pack()

Kategorievar = StringVar()
Kategorievar = ttk.Combobox(tab1, textvariable=Kategorievar, font=("Arial", 10))
Kategorievar.bind('<<ComboboxSelected>>')
Kategorievar['values'] = (KategorieList)
Kategorievar.state(["readonly"])
Kategorievar.pack()

tab1_textBetrag = tk.Label(tab1, text="Betrag: ", fg="black", bg="white", font=("Arial", 15))
tab1_textBetrag.pack()

tab1_entry = tk.Entry(tab1, font=("Arial", 10))
tab1_entry.pack()

tab1_textaddOrreset = tk.Label(tab1, text="Addieren oder Ersetzen:", fg="black", bg="white", font=("Arial", 15))
tab1_textaddOrreset.pack()

AddOrResetvar = StringVar()
AddOrResetvar = ttk.Combobox(tab1, textvariable=AddOrResetvar, font=("Arial", 10))
AddOrResetvar.bind('<<ComboboxSelected>>')
AddOrResetvar['values'] = ("Addieren", "Ersetzen")
AddOrResetvar.state(["readonly"])
AddOrResetvar.pack()

button_tab1 = tk.Button(tab1, font=("Arial", 15), text="Hinzufügen", bg="#F25E24", fg="black", command= lambda: BereiteValueInsertVor(str(Yearvar.get()), str(Monthvar.get()), str(Kategorievar.get()), str(tab1_entry.get()), str(AddOrResetvar.get())))
button_tab1.pack()

def BereiteValueInsertVor(year, month, kategorie, value, addOrreset):
    if year == "" or month == "" or kategorie == "" or value == "" or addOrreset == "":
        print("Leeres Feld vorhanden.")
        empty_fields()
    if is_valid_number(value) == True:
        print("Valid")
    if is_valid_number(value) == False:
        print("Invalid")
        showError("Es ist etwas schiefgelaufen. Bitte überprüfe die Eingabe auf Richtigkeit. Es dürfen keine Sonderzeichen auftreten. Für Nachkommastellen, nutze bitte einen Punkt.")
    else:
        Insert(year, month, kategorie, value, addOrreset)
        



# Widgets für den zweiten Tab




tab2_title = tk.Label(tab2, text="Ausgaben Löschen:", bg="white", fg="#9c1c78", font=("Arial", 25))
tab2_title.pack()

YearsListName = "YearsList.txt"
YearsList = laden_liste(YearsListName)  #Checkt ob Years Liste vorhanden ist

tab2_textYear = tk.Label(tab2, text="Jahr:", fg="black", bg="white", font=("Arial", 15))
tab2_textYear.pack()

YearvarTab2 = StringVar()
YearvarTab2 = ttk.Combobox(tab2, textvariable=YearvarTab2, font=("Arial", 10))
YearvarTab2.bind('<<ComboboxSelected>>')
YearvarTab2['values'] = (YearsList)
YearvarTab2.state(["readonly"])
YearvarTab2.pack()

tab2_textMonth = tk.Label(tab2, text="Monat:", fg="black", bg="white", font=("Arial", 15))
tab2_textMonth.pack()

MonthvarTab2 = StringVar()
MonthvarTab2 = ttk.Combobox(tab2, textvariable=MonthvarTab2, font=("Arial", 10))
MonthvarTab2.bind('<<ComboboxSelected>>')
MonthvarTab2['values'] = ("Januar", "Februar", "März", "April", "Mai", "Juni", "Juli", "August", "September", "Oktober", "November", "Dezember")
MonthvarTab2.state(["readonly"])
MonthvarTab2.pack()

KategorieListName = "KategorieList.txt"
KategorieList = laden_liste(KategorieListName)  #Checkt ob Years Liste vorhanden ist

tab2_textKategorie = tk.Label(tab2, text="Kategorie: ", fg="black", bg="white", font=("Arial", 15))
tab2_textKategorie.pack()

KategorievarTab2 = StringVar()
KategorievarTab2 = ttk.Combobox(tab2, textvariable=KategorievarTab2, font=("Arial", 10))
KategorievarTab2.bind('<<ComboboxSelected>>')
KategorievarTab2['values'] = (KategorieList)
KategorievarTab2.state(["readonly"])
KategorievarTab2.pack()

tab2_textBetrag = tk.Label(tab2, text="Betrag: ", fg="black", bg="white", font=("Arial", 15))
tab2_textBetrag.pack()




button_tab2 = tk.Button(tab2, font=("Arial", 15), text="Löschen", bg="#F25E24", fg="black", command= lambda: BereiteValueDeleteVor(str(YearvarTab2.get()), str(MonthvarTab2.get()), str(KategorievarTab2.get())))
button_tab2.pack()



def BereiteValueDeleteVor(year, month, kategorie):
    if year == "" or month == "" or kategorie == "":
        print("Leeres Feld vorhanden.")
        empty_fields()
    else:
        deleteValue(year, month, kategorie)




# Widgets für den dritten Tab
tab3_title = tk.Label(tab3, text="Ausgaben-Kategorien Hinzufügen", fg="#9c1c78", bg="white", font=("Arial", 25))
tab3_title.pack()

YearsListName = "YearsList.txt"
YearsList = laden_liste(YearsListName)  #Checkt ob Years Liste vorhanden ist


tab3_textEntry = tk.Label(tab3, text="Kategoriename:", bg = "white", fg="black", font=("Arial", 15), pady=10)
tab3_textEntry.pack()

tab3_entry = tk.Entry(tab3, bg= "white",fg="black", font=("Arial", 15))
tab3_entry.pack()


tab3_entryBtn = tk.Button(tab3, command= lambda: BereiteKategorieHinzufügenVor(), text="Hinzufügen", bg = "#F25E24", fg="black", font=("Arial", 15), pady=10)
tab3_entryBtn.pack()

def BereiteKategorieHinzufügenVor():
    if str(tab3_entry.get()) == "":
        print("Leerer Kategoriename")
        empty_fields()
    else:
        element = str(tab3_entry.get())
        NewKategorie(element)

# Widgets für den vierten Tab

tab4_title = tk.Label(tab4, text="Ausgaben-Kategorien Löschen", fg="#9c1c78", bg="white", font=("Arial", 25))
tab4_title.pack()

tab4_textEntry = tk.Label(tab4, text="Kategorie:", bg = "white", fg="black", font=("Arial", 15), pady=10)
tab4_textEntry.pack()


DeleteKategorieVar = StringVar()
DeleteKategorieVar = ttk.Combobox(tab4, textvariable=DeleteKategorieVar, font=("Arial", 10))
DeleteKategorieVar.bind('<<ComboboxSelected>>')
DeleteKategorieVar['values'] = (KategorieList)
DeleteKategorieVar.state(["readonly"])
DeleteKategorieVar.pack()

tab3_DeleteBtn = tk.Button(tab4, command= lambda: BereiteKategorieDeleteVor(str(DeleteKategorieVar.get())), text="Löschen", bg = "#F25E24", fg="black", font=("Arial", 15), pady=10)
tab3_DeleteBtn.pack()

def BereiteKategorieDeleteVor(KategorieName):
    if str(DeleteKategorieVar.get()) == "":
        print("Leeres Feld vorhanden")
        empty_fields()
    else:
        deleteKategorie(KategorieName)




#-----------------------------------------------------------------------------------------------------------------------





# Unten Rechts
unten_rechts = tk.Frame(unten, bg="#208844")
unten_rechts.pack(side='right', fill='both', expand=True, padx=7.5)#, pady=7.5)


notebookUR = ttk.Notebook(unten_rechts)
notebookUR.pack(expand=True, fill='both', padx=7.5, pady=7.5)

styleTab1 = ttk.Style()
styleTab1.configure("Tab1.TLabel", foreground="black", background="white")

styleTab2 = ttk.Style()
styleTab2.configure("Tab2.TLabel", foreground="black", background="white")

styleTab3 = ttk.Style()
styleTab3.configure("Tab3.TLabel", foreground="black", background="white")

styleTab4 = ttk.Style()
styleTab4.configure("Tab4.TLabel", foreground="black", background="white")

# Erster Tab
tab1UR = ttk.Frame(notebookUR, style="Tab1.TLabel")
notebookUR.add(tab1UR, text='Einnahmen')

tab1URTitle = tk.Label(tab1UR, text="Einnahmen", font=("Arial", 20), fg="#208844", bg="white")
tab1URTitle.pack()

GehaltMonatHinzufügenFrame = tk.Frame(tab1UR, bg="#208844")
GehaltMonatHinzufügenFrame.pack(side='left', fill='both', expand=True, padx=7.5, pady=7.5)

GehaltMonatLöschenFrame = tk.Frame(tab1UR, bg="#208844")
GehaltMonatLöschenFrame.pack(side='right', fill='both', expand=True, padx=7.5, pady=7.5)

GehaltMonatHinzufügen = tk.Frame(GehaltMonatHinzufügenFrame, bg="white")
GehaltMonatHinzufügen.pack(side='left', fill='both', expand=True, padx=10, pady=10)

GehaltMonatLöschen = tk.Frame(GehaltMonatLöschenFrame, bg="white")
GehaltMonatLöschen.pack(side='right', fill='both', expand=True, padx=10, pady=10)

#Gehalt hinzufügen widgets
GehaltMonatHinzufügenTitle = tk.Label(GehaltMonatHinzufügen, text="Einnahmen eintragen:", font=("Arial", 15), bg="white", fg="green")
GehaltMonatHinzufügenTitle.pack()

YearsListName = "YearsList.txt"
YearsList = laden_liste(YearsListName)  #Checkt ob Years Liste vorhanden ist

GehaltHinzufügen_textYear = tk.Label(GehaltMonatHinzufügen, text="Jahr:", fg="black", bg="white", font=("Arial", 15))
GehaltHinzufügen_textYear.pack()

GehaltHinzufügenYearvar = StringVar()
GehaltHinzufügenYearvar = ttk.Combobox(GehaltMonatHinzufügen, textvariable=GehaltHinzufügenYearvar, font=("Arial", 10))
GehaltHinzufügenYearvar.bind('<<ComboboxSelected>>')
GehaltHinzufügenYearvar['values'] = (YearsList)
GehaltHinzufügenYearvar.state(["readonly"])
GehaltHinzufügenYearvar.pack()

tab1_textMonth = tk.Label(GehaltMonatHinzufügen, text="Monat:", fg="black", bg="white", font=("Arial", 15))
tab1_textMonth.pack()

GehaltHinzufügenMonthvar = StringVar()
GehaltHinzufügenMonthvar = ttk.Combobox(GehaltMonatHinzufügen, textvariable=GehaltHinzufügenMonthvar, font=("Arial", 10))
GehaltHinzufügenMonthvar.bind('<<ComboboxSelected>>')
GehaltHinzufügenMonthvar['values'] = ("Januar", "Februar", "März", "April", "Mai", "Juni", "Juli", "August", "September", "Oktober", "November", "Dezember")
GehaltHinzufügenMonthvar.state(["readonly"])
GehaltHinzufügenMonthvar.pack()


tab1_textEQ = tk.Label(GehaltMonatHinzufügen, text="Einnahmen-Quelle", fg="black", bg="white", font=("Arial", 15))
tab1_textEQ.pack()

EinnahmeQuellenListName = "EinnahmequelleList.txt"
EinnahmeQuelleList = laden_liste(EinnahmeQuellenListName)

EinnahmeQuelleVar = StringVar()
EinnahmeQuelleVar = ttk.Combobox(GehaltMonatHinzufügen, textvariable=EinnahmeQuelleVar, font=("Arial", 10))
EinnahmeQuelleVar.bind('<<ComboboxSelected>>')
EinnahmeQuelleVar['values'] = (EinnahmeQuelleList)
EinnahmeQuelleVar.state(["readonly"])
EinnahmeQuelleVar.pack()

GehaltHinzufügen_textBetrag = tk.Label(GehaltMonatHinzufügen, text="Betrag: ", fg="black", bg="white", font=("Arial", 15))
GehaltHinzufügen_textBetrag.pack()

GehaltHinzufügen_entry = tk.Entry(GehaltMonatHinzufügen, font=("Arial", 10))
GehaltHinzufügen_entry.pack()


button_GehaltHinzufügen = tk.Button(GehaltMonatHinzufügen, font=("Arial", 15), text="Hinzufügen", bg="#F25E24", fg="black", command= lambda: BereiteGehaltInsertVor(str(GehaltHinzufügenYearvar.get()), str(GehaltHinzufügenMonthvar.get()), str(EinnahmeQuelleVar.get()), str(GehaltHinzufügen_entry.get())))
button_GehaltHinzufügen.pack()

def BereiteGehaltInsertVor(year, month, quelle, value):
    if str(year) == "" or str(month) == "" or str(quelle) == "" or str(value) == "":
        print("Leeres Feld vorhanden.")
        empty_fields()
    if is_valid_number(value) == True:
        print("Valid")
    if is_valid_number(value) == False:
        print("Invalid")
        showError("Es ist etwas schiefgelaufen. Bitte überprüfe die Eingabe auf Richtigkeit. Es dürfen keine Sonderzeichen auftreten. Für Nachkommastellen, nutze bitte einen Punkt.")
    else:
        GehaltInsert(year, month, quelle, value)


#gehalt löschen widgets
GehaltMonatLöschenTitle = tk.Label(GehaltMonatLöschen, text="Einnahmen löschen:", font=("Arial", 15), bg="white", fg="#208844")
GehaltMonatLöschenTitle.pack()

YearsListName = "YearsList.txt"
YearsList = laden_liste(YearsListName)  #Checkt ob Years Liste vorhanden ist

GehaltLöschen_textYear = tk.Label(GehaltMonatLöschen, text="Jahr:", fg="black", bg="white", font=("Arial", 15))
GehaltLöschen_textYear.pack()

GehaltLöschenYearvar = StringVar()
GehaltLöschenYearvar = ttk.Combobox(GehaltMonatLöschen, textvariable=GehaltLöschenYearvar, font=("Arial", 10))
GehaltLöschenYearvar.bind('<<ComboboxSelected>>')
GehaltLöschenYearvar['values'] = (YearsList)
GehaltLöschenYearvar.state(["readonly"])
GehaltLöschenYearvar.pack()

tab1_textMonth = tk.Label(GehaltMonatLöschen, text="Monat:", fg="black", bg="white", font=("Arial", 15))
tab1_textMonth.pack()

GehaltLöschenMonthvar = StringVar()
GehaltLöschenMonthvar = ttk.Combobox(GehaltMonatLöschen, textvariable=GehaltLöschenMonthvar, font=("Arial", 10))
GehaltLöschenMonthvar.bind('<<ComboboxSelected>>')
GehaltLöschenMonthvar['values'] = ("Januar", "Februar", "März", "April", "Mai", "Juni", "Juli", "August", "September", "Oktober", "November", "Dezember")
GehaltLöschenMonthvar.state(["readonly"])
GehaltLöschenMonthvar.pack()

tab1_textEQL = tk.Label(GehaltMonatLöschen, text="Einnahmen-Quelle:", fg="black", bg="white", font=("Arial", 15))
tab1_textEQL.pack()

EinnahmeQuellenListName = "EinnahmequelleList.txt"
EinnahmeQuelleList = laden_liste(EinnahmeQuellenListName)

EinnahmeQuelleVarDel = StringVar()
EinnahmeQuelleVarDel = ttk.Combobox(GehaltMonatLöschen, textvariable=EinnahmeQuelleVarDel, font=("Arial", 10))
EinnahmeQuelleVarDel.bind('<<ComboboxSelected>>')
EinnahmeQuelleVarDel['values'] = (EinnahmeQuelleList)
EinnahmeQuelleVarDel.state(["readonly"])
EinnahmeQuelleVarDel.pack()


button_GehaltLöschen = tk.Button(GehaltMonatLöschen, font=("Arial", 15), text="Löschen", bg="#F25E24", fg="black", command= lambda: BereiteGehaltDeleteVor(str(GehaltLöschenYearvar.get()), str(GehaltLöschenMonthvar.get()), str(EinnahmeQuelleVarDel.get())))
button_GehaltLöschen.pack()

def BereiteGehaltDeleteVor(year, month, quelle):
    print(year, month, quelle)
    if str(year) == "" or str(month) == "" or str(quelle) == "":
        print("Leeres Feld vorhanden.")
        empty_fields()
    else:
        GehaltDelete(year, month, quelle)



# Zweiter Tab
tab2UR = ttk.Frame(notebookUR, style="Tab2.TLabel")
notebookUR.add(tab2UR, text='Einnahmen-Quellen')


tab2URTitle = tk.Label(tab2UR, text="Einnahmen-Quellen", font=("Arial", 20), fg="#208844", bg="white")
tab2URTitle.pack()

EinnahmeQuellenHinzufügenFrame = tk.Frame(tab2UR, bg="#208844")
EinnahmeQuellenHinzufügenFrame.pack(side='left', fill='both', expand=True, padx=7.5, pady=7.5)

EinnahmeQuellenLöschenFrame = tk.Frame(tab2UR, bg="#208844")
EinnahmeQuellenLöschenFrame.pack(side='right', fill='both', expand=True, padx=7.5, pady=7.5)

EinnahmeQuellenHinzufügen = tk.Frame(EinnahmeQuellenHinzufügenFrame, bg="white")
EinnahmeQuellenHinzufügen.pack(side='left', fill='both', expand=True, padx=10, pady=10)

EinnahmeQuellenLöschen = tk.Frame(EinnahmeQuellenLöschenFrame, bg="white")
EinnahmeQuellenLöschen.pack(side='right', fill='both', expand=True, padx=10, pady=10)

#EinnahmeQuellen hinzufügen widgets
EinnahmeQuellenHinzufügenTitle = tk.Label(EinnahmeQuellenHinzufügen, text="Einnahme-Quellen hinzufügen:", font=("Arial", 15), bg="white", fg="#208844")
EinnahmeQuellenHinzufügenTitle.pack()


tab2_textEinQuelle = tk.Label(EinnahmeQuellenHinzufügen, text="Quellename:", fg="black", bg="white", font=("Arial", 15))
tab2_textEinQuelle.pack()

EinnahmeQuellenEntry = tk.Entry(EinnahmeQuellenHinzufügen, font=("Arial", 10))
EinnahmeQuellenEntry.pack()


button_EinnahmequellenHinzufügen = tk.Button(EinnahmeQuellenHinzufügen, font=("Arial", 15), text="Hinzufügen", bg="#F25E24", fg="black", command= lambda: BereiteEinnahmeQuellenInsertVor(EinnahmeQuellenEntry.get()))
button_EinnahmequellenHinzufügen.pack()

def BereiteEinnahmeQuellenInsertVor(name):
    if str(name) == "":
        print("Leeres Feld vorhanden.")
        empty_fields()
    else:
        NewEinnahmeQuelle(name)


#EinnahmeQuellen löschen widgets
EinnahmeQuellenLöschenTitle = tk.Label(EinnahmeQuellenLöschen, text="Einnahmen-Quellen löschen:", font=("Arial", 15), bg="white", fg="#208844")
EinnahmeQuellenLöschenTitle.pack()

EinnahmeQuellenListName = "EinnahmequelleList.txt"
EinnahmeQuelleList = laden_liste(EinnahmeQuellenListName)

EinnahmeQuelleLöschen_textYear = tk.Label(EinnahmeQuellenLöschen, text="Quellename:", fg="black", bg="white", font=("Arial", 15))
EinnahmeQuelleLöschen_textYear.pack()

EinnahmeQuelleLöschenYearvar = StringVar()
EinnahmeQuelleLöschenYearvar = ttk.Combobox(EinnahmeQuellenLöschen, textvariable=EinnahmeQuelleLöschenYearvar, font=("Arial", 10))
EinnahmeQuelleLöschenYearvar.bind('<<ComboboxSelected>>')
EinnahmeQuelleLöschenYearvar['values'] = (EinnahmeQuelleList)
EinnahmeQuelleLöschenYearvar.state(["readonly"])
EinnahmeQuelleLöschenYearvar.pack()

button_EinnahmequellenLöschen = tk.Button(EinnahmeQuellenLöschen, font=("Arial", 15), text="Löschen", bg="#F25E24", fg="black", command= lambda: BereiteEinnahmeQuelleDeleteVor(EinnahmeQuelleLöschenYearvar.get()))
button_EinnahmequellenLöschen.pack()

def BereiteEinnahmeQuelleDeleteVor(name):
    if str(name) == "":
        print("Leeres Feld vorhanden.")
        empty_fields()
    else:
        deleteEinnahmequelle(name)


root.mainloop()
