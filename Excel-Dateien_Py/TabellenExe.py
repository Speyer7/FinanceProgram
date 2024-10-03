from openpyxl import Workbook, load_workbook

wb = load_workbook("Finanzen2024.xlsx") # Erstellt NEUES, bei nur hinzufügen/abändern nutze load_workbook
ws = wb["Finanzen"]


def ReadOutContentOfRow():
    row = input("Enter a row: ")

    print(" " + str(ws["A" + str(row)].value))
    print("Januar: " + str(ws["B" + str(row)].value))
    print("Februar: " + str(ws["C" + str(row)].value))
    print("März: " + str(ws["D" + str(row)].value))
    print("April: " + str(ws["E" + str(row)].value))
    print("Mai: " + str(ws["F" + str(row)].value))
    print("Juni: " + str(ws["G" + str(row)].value))
    print("Juli: " + str(ws["H" + str(row)].value))
    print("August: " + str(ws["I" + str(row)].value))
    print("September: " + str(ws["J" + str(row)].value))
    print("Oktober: " + str(ws["K" + str(row)].value))
    print("November: " + str(ws["L" + str(row)].value))
    print("Dezember: " + str(ws["M" + str(row)].value))


def ReadOutContentOfKategorie():
    KategorieInp = input("Enter a Kategorie: ")


    for row in range(1, ws.max_row + 1):
        print(row)
        if KategorieInp == ws.cell(row=row, column=1).value:
            print("Gefunden")
            global KategorieRowRead
            KategorieRowRead = row


    print("----------------------------------")
    print(str(KategorieRowRead) + " ist die gesuchte Zeile.")
    print("----------------------------------")

    print(" " + str(ws["A" + str(KategorieRowRead)].value))
    print("Januar: " + str(ws["B" + str(KategorieRowRead)].value))
    print("Februar: " + str(ws["C" + str(KategorieRowRead)].value))
    print("März: " + str(ws["D" + str(KategorieRowRead)].value))
    print("April: " + str(ws["E" + str(KategorieRowRead)].value))
    print("Mai: " + str(ws["F" + str(KategorieRowRead)].value))
    print("Juni: " + str(ws["G" + str(KategorieRowRead)].value))
    print("Juli: " + str(ws["H" + str(KategorieRowRead)].value))
    print("August: " + str(ws["I" + str(KategorieRowRead)].value))
    print("September: " + str(ws["J" + str(KategorieRowRead)].value))
    print("Oktober: " + str(ws["K" + str(KategorieRowRead)].value))
    print("November: " + str(ws["L" + str(KategorieRowRead)].value))
    print("Dezember: " + str(ws["M" + str(KategorieRowRead)].value))

    KategorieInp = input("Of which Kategorie do you want to see all values together? ")

    for row in range(2, ws.max_row + 1):
        print(row)
        if KategorieInp == ws.cell(row=row, column=1).value:
            print("Gefunden")
            global KategorieRow
            KategorieRow = row

    test = KategorieRow
    print("zeile der Kategorie: " + str(test))



def ReadOutContentOfMonth(): # Zahlen weg!
    MonthInp = input("Enter a Month: ")

    for column in range(1, ws.max_column + 1):
        print(column)
        if MonthInp == ws.cell(row=1, column=column).value:
            print("Gefunden")
            global MonthColRead
            MonthColRead = column
    



    print("----------------------------------")
    print(str(MonthColRead) + " ist die gesuchte Zeile.")
    print("----------------------------------")

    print(MonthInp + ":")
    print("---------")

    for i in range(2, ws.max_row + 1):
        #print(i), print(ws["A" + str(i)].value)

        if MonthColRead == int("1"):
            MonthColRead = "A"
        if MonthColRead == int("2"):
            MonthColRead = "B"
        if MonthColRead == int("3"):
            MonthColRead = "C"
        if MonthColRead == int("4"):
            MonthColRead = "D"
        if MonthColRead == int("5"):
            MonthColRead = "E"
        if MonthColRead == int("6"):
            MonthColRead = "F"
        if MonthColRead == int("7"):
            MonthColRead = "G"
        if MonthColRead == int("8"):
            MonthColRead = "H"
        if MonthColRead == int("9"):
            MonthColRead = "I"
        if MonthColRead == int("10"):
            MonthColRead = "J"
        if MonthColRead == int("11"):
            MonthColRead = "K"
        if MonthColRead == int("12"):
            MonthColRead = "L"
        if MonthColRead == int("13"):
            MonthColRead = "M"

        print(ws["A" + str(i)].value + ": ")
        print(ws[str(MonthColRead) + str(i)].value) # # zahlen weg!


def NewKategorie():
    NewKategorieInp = input("New Kategorie: ")

    ws["A" + str((int(ws.max_row) + 1))].value = NewKategorieInp

    clearTabel()


def deleteKategorie():
    DeleteKategorieInp = input("To Delete Kategorie: ")

    for row in range(1, ws.max_row + 1):
        print(row)
        if DeleteKategorieInp == ws.cell(row=row, column=1).value:
            print("Gefunden")
            global KategorieRowDelete
            KategorieRowDelete = row

    ws.delete_rows(KategorieRowDelete)

#!!!
def deleteValue():
    DeleteValueOfKategorie = input("Of which Kategorie do you want to delete a value? ")
    DeleteValueOfMonth = input("Of which Month do you want to delete a value? ")

    for row in range(1, ws.max_row + 1):
        print(row)
        if DeleteValueOfKategorie == ws.cell(row=row, column=1).value:
            print("Gefunden")
            global DeleteValueOfKategorieRow
            DeleteValueOfKategorieRow = row

    for col in range(1, ws.max_column + 1):
        print(col)
        if DeleteValueOfMonth == ws.cell(row=1, column=col).value:
            print("Gefunden")
            global DeleteValueOfMonthCol
            DeleteValueOfMonthCol = col

    if DeleteValueOfMonthCol == int("1"):
        DeleteValueOfMonthCol = "A"
    if DeleteValueOfMonthCol == int("2"):
        DeleteValueOfMonthCol = "B"
    if DeleteValueOfMonthCol == int("3"):
        DeleteValueOfMonthCol = "C"
    if DeleteValueOfMonthCol == int("4"):
        DeleteValueOfMonthCol = "D"
    if DeleteValueOfMonthCol == int("5"):
        DeleteValueOfMonthCol = "E"
    if DeleteValueOfMonthCol == int("6"):
        DeleteValueOfMonthCol = "F"
    if DeleteValueOfMonthCol == int("7"):
        DeleteValueOfMonthCol = "G"
    if DeleteValueOfMonthCol == int("8"):
        DeleteValueOfMonthCol = "H"
    if DeleteValueOfMonthCol == int("9"):
        DeleteValueOfMonthCol = "I"
    if DeleteValueOfMonthCol == int("10"):
        DeleteValueOfMonthCol = "J"
    if DeleteValueOfMonthCol == int("11"):
        DeleteValueOfMonthCol = "K"
    if DeleteValueOfMonthCol == int("12"):
        DeleteValueOfMonthCol = "L"
    if DeleteValueOfMonthCol == int("13"):
        DeleteValueOfMonthCol = "M"

    ws[DeleteValueOfMonthCol + str(DeleteValueOfKategorieRow)].value = 0


def Insert():
    InsertToKategorie = input("Enter a Kategorie: ")
    InsertToMonth = input("Enter a Month: ")
    InsertPrice = input("Enter a Price: ")
    InsertOrAdd = input("Want to <<add>> oder <<reset>>?: ")

    for row in range(1, ws.max_row + 1):
        print(row)
        if InsertToKategorie == ws.cell(row=row, column=1).value:
            print("Gefunden")
            global KategorieRowInsert
            KategorieRowInsert = row

    for col in range(1, ws.max_column + 1):
        print(col)
        if InsertToMonth == ws.cell(row=1, column=col).value:
            print("Gefunden")
            global MonthRowInsert
            MonthRowInsert = col

    print(str(KategorieRowInsert) + " ist die Zeile der Kategorie " + InsertToKategorie)
    print(str(MonthRowInsert) + " ist die Zeile des Monats " + InsertToMonth)

    if MonthRowInsert == int("1"):
        MonthRowInsert = "A"
    if MonthRowInsert == int("2"):
        MonthRowInsert = "B"
    if MonthRowInsert == int("3"):
        MonthRowInsert = "C"
    if MonthRowInsert == int("4"):
        MonthRowInsert = "D"
    if MonthRowInsert == int("5"):
        MonthRowInsert = "E"
    if MonthRowInsert == int("6"):
        MonthRowInsert = "F"
    if MonthRowInsert == int("7"):
        MonthRowInsert = "G"
    if MonthRowInsert == int("8"):
        MonthRowInsert = "H"
    if MonthRowInsert == int("9"):
        MonthRowInsert = "I"
    if MonthRowInsert == int("10"):
        MonthRowInsert = "J"
    if MonthRowInsert == int("11"):
        MonthRowInsert = "K"
    if MonthRowInsert == int("12"):
        MonthRowInsert = "L"
    if MonthRowInsert == int("13"):
        MonthRowInsert = "M"

    print(str(MonthRowInsert) + " ist der Buchstabe des Monats " + InsertToMonth)

    if InsertOrAdd == "add":
        PriceNow = ws[MonthRowInsert + str(KategorieRowInsert)].value
        if str(PriceNow) == "None":
            ws[MonthRowInsert + str(KategorieRowInsert)].value = 0
        Price = int(ws[MonthRowInsert + str(KategorieRowInsert)].value) + int(InsertPrice)
        ws[MonthRowInsert + str(KategorieRowInsert)].value = Price
    if InsertOrAdd == "reset":
        ws[MonthRowInsert + str(KategorieRowInsert)].value = InsertPrice


def getValueOfMonthAllKategories():
    MonthInp = input("Of which Month do you want to see all values together? ")

    for col in range(1, ws.max_column + 1):
        print(col)
        if MonthInp == ws.cell(row=1, column=col).value:
            print("Gefunden")
            global MonthCol
            MonthCol = col

    if MonthCol == int("1"):
        MonthCol = "A"
    if MonthCol == int("2"):
        MonthCol = "B"
    if MonthCol == int("3"):
        MonthCol = "C"
    if MonthCol == int("4"):
        MonthCol = "D"
    if MonthCol == int("5"):
        MonthCol = "E"
    if MonthCol == int("6"):
        MonthCol = "F"
    if MonthCol == int("7"):
        MonthCol = "G"
    if MonthCol == int("8"):
        MonthCol = "H"
    if MonthCol == int("9"):
        MonthCol = "I"
    if MonthCol == int("10"):
        MonthCol = "J"
    if MonthCol == int("11"):
        MonthCol = "K"
    if MonthCol == int("12"):
        MonthCol = "L"
    if MonthCol == int("13"):
        MonthCol = "M"

    AllValuesTogether = 0


    for colMonth in range(2, ws.max_row + 1):
        print(ws["A" + str(colMonth)].value)
        if str(ws[MonthCol + str(colMonth)].value) == "None":
            ws[MonthCol + str(colMonth)].value = 0
        print(ws[MonthCol + str(colMonth)].value)
        AllValuesTogether += int(ws[MonthCol + str(colMonth)].value)

    print("---------------")
    print("Zusammen:")
    print(AllValuesTogether)

    #for row in range(2, ws.max_row + 1):
        #Value = ws[MonthCol + str(row)].value
        #if str(Value) == "None":
            #Value = 0
        #AllValuesTogether = int(AllValuesTogether) + int(Value)
        #print(AllValuesTogether)


def getValueOfKategorieAllMonths():
    KategorieInp = input("Of which Kategorie do you want to see all values together? ")

    for row in range(2, ws.max_row + 1):
        print(row)
        if KategorieInp == ws.cell(row=row, column=1).value:
            print("Gefunden")
            global KategorieRow
            KategorieRow = row

    test = KategorieRow
    print("zeile der Kategorie: " + str(test))
    AllValuesTogether = 0

    for col in range(2, ws.max_column + 1):
        if col == int("1"):
            col = "A"
        if col == int("2"):
            col = "B"
        if col == int("3"):
            col = "C"
        if col == int("4"):
            col = "D"
        if col == int("5"):
            col = "E"
        if col == int("6"):
            col = "F"
        if col == int("7"):
            col = "G"
        if col == int("8"):
            col = "H"
        if col == int("9"):
            col = "I"
        if col == int("10"):
            col = "J"
        if col == int("11"):
            col = "K"
        if col == int("12"):
            col = "L"
        if col == int("13"):
            col = "M"



        if str(ws[str(col) + str(KategorieRow)].value) == "None":
            ws[str(col) + str(KategorieRow)].value = 0

        print(str(ws[str(col) + "1"].value) + ":")
        print(ws[str(col) + str(KategorieRow)].value)
        AllValuesTogether += int(ws[col + str(KategorieRow)].value)

    print("---------------")
    print("Zusammen:")
    print(AllValuesTogether)

#!!!
def getValueAll():
    global AllValuesTogether
    AllValuesTogether = 0
    for col in range(2, ws.max_column + 1):
        print("-----Col: " + str(col))
        if col == int("1"):
            col = "A"
        if col == int("2"):
            col = "B"
        if col == int("3"):
            col = "C"
        if col == int("4"):
            col = "D"
        if col == int("5"):
            col = "E"
        if col == int("6"):
            col = "F"
        if col == int("7"):
            col = "G"
        if col == int("8"):
            col = "H"
        if col == int("9"):
            col = "I"
        if col == int("10"):
            col = "J"
        if col == int("11"):
            col = "K"
        if col == int("12"):
            col = "L"
        if col == int("13"):
            col = "M"

        for row in range(2, ws.max_row + 1):
            print(row)


            AllValuesTogether += int(ws[str(col) + str(row)].value)

    print("---------------")
    print("Zusammen:")
    print(AllValuesTogether)

def clearTabel():
    print("clearing Tabel...")
    for col in range(1, ws.max_column + 1):
        print("---Col: " + str(col))
        if col == int("1"):
            col = "A"
        if col == int("2"):
            col = "B"
        if col == int("3"):
            col = "C"
        if col == int("4"):
            col = "D"
        if col == int("5"):
            col = "E"
        if col == int("6"):
            col = "F"
        if col == int("7"):
            col = "G"
        if col == int("8"):
            col = "H"
        if col == int("9"):
            col = "I"
        if col == int("10"):
            col = "J"
        if col == int("11"):
            col = "K"
        if col == int("12"):
            col = "L"
        if col == int("13"):
            col = "M"
        for row in range(2, ws.max_row + 1):
            print(row)
            if str(ws[str(col) + str(row)].value) == "None":
                ws[str(col) + str(row)].value = 0
                print("Changend")

getValueAll()

wb.save("Finanzen2024.xlsx")