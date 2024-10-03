from openpyxl import Workbook, load_workbook

wb = Workbook() # Erstellt NEUES, bei nur hinzufügen/abändern nutze load_workbook
ws = wb.active
ws.title = "Finanzen"


ws.append(["", "Januar", "Februar", "März", "April", "Mai", "Juni", "Juli", "August", "September", "Oktober", "November", "Dezember"]) #Fügt eine reihe hinzu

ws['A2'].value = "Essen"
ws["A3"].value = "Klamotten"
ws["A4"].value = "Kinder"
ws["A5"].value = "Haus"

ws.append(["Strom", "80"])




wb.save("Finanzen2024.xlsx")